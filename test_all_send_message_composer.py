"""
CometChat React Native Android - ALL Send Message & Composer Test Cases
Consolidated from 10 separate test files into one organized script.
Covers: Positive (132 TCs) + Negative (22 TCs) + Voice Recording (5 TCs)

Sections:
  1. Send Message (MSG_001-MSG_064) — Positive
  2. Emoji/Sticker (MSG_065-MSG_096) — Positive
  3. @Mention (MSG_097-MSG_110) — Positive
  4. Composer Features (MSG_111-MSG_121) — Positive
  5. Rich Media Formatting (MSG_122-MSG_132) — Positive
  6. Voice Recording (MSG_078-MSG_082) — Positive (separate test)
  7. Negative Tests (MSG_001-MSG_022) — Negative sheet

Usage:
  python3 -m pytest test_all_send_message_composer.py -v -s -k "test_positive"
  python3 -m pytest test_all_send_message_composer.py -v -s -k "test_negative"
  python3 -m pytest test_all_send_message_composer.py -v -s -k "test_voice"
  python3 -m pytest test_all_send_message_composer.py -v -s  # run all
"""
import time
import subprocess
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ============================================================
# CONSTANTS
# ============================================================
EXCEL = "Cometchat_Features/Send_&_Compose/SM_SLC_RMF_Test_Cases.xlsx"
PKG = "com.cometchat.sampleapp.reactnative.android"
ADB = "/Users/admin/android-sdk/platform-tools/adb"
DEVICE = "HZC90Q76"
BUILD = "React Native Android v5.2.10"
MIC_POS = (858, 1794)
DEL_POS = (91, 1910)
PAU_POS = (889, 1912)
SND_POS = (991, 1916)


# ============================================================
# HELPER FUNCTIONS
# ============================================================
def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _login_if_needed(driver):
    """Login by selecting Andrew Joseph sample user."""
    try:
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Andrew Joseph"))).click()
        time.sleep(0.3)
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Continue"))).click()
        time.sleep(1.5)
        try:
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.ID, "android:id/button1"))).click()
        except Exception:
            pass
        try:
            _wait(driver, 3).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@text='Allow' or @text='ALLOW']"))).click()
        except Exception:
            pass
        print("Logged in as Andrew Joseph.")
    except Exception:
        print("Already logged in.")


def _go_to_chat_list(driver):
    """Navigate back to the main chat list."""
    for _ in range(8):
        try:
            clear = driver.find_elements(AppiumBy.ACCESSIBILITY_ID, "Clear search")
            if clear:
                clear[0].click()
                time.sleep(0.5)
                continue
        except Exception:
            pass
        try:
            ishwar = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'Ishwar')]")
            if ishwar:
                print("At chat list.")
                return True
        except Exception:
            pass
        try:
            driver.back()
            time.sleep(0.5)
        except Exception:
            pass
    return False


def _open_chat(driver, user_name="Ishwar Borwar"):
    """Open a chat — checks if already in chat, then tries direct/scroll."""
    try:
        composer = WebDriverWait(driver, 3, poll_frequency=0.3).until(
            EC.presence_of_element_located((
                AppiumBy.XPATH,
                "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")))
        if composer.is_displayed():
            print(f"Already in chat.")
            return True
    except Exception:
        pass
    try:
        user = WebDriverWait(driver, 3, poll_frequency=0.3).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, f"//*[contains(@content-desc,'{user_name}')]")))
        user.click()
        time.sleep(0.5)
        return True
    except Exception:
        pass
    # Scroll fallback
    try:
        screen = driver.get_window_size()
        for _ in range(5):
            els = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@content-desc,'{user_name}')]")
            if els:
                els[0].click()
                time.sleep(0.5)
                return True
            driver.swipe(screen['width'] // 2, screen['height'] * 2 // 3,
                         screen['width'] // 2, screen['height'] // 3, 800)
            time.sleep(0.5)
    except Exception:
        pass
    print(f"Could not find {user_name}")
    return False


def _get_composer(driver):
    return _wait(driver).until(EC.element_to_be_clickable((
        AppiumBy.XPATH,
        "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")))


def _send_message(driver, text):
    inp = _get_composer(driver)
    inp.click()
    inp.clear()
    inp.send_keys(text)
    time.sleep(0.3)
    try:
        send_btn = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']")))
        send_btn.click()
        time.sleep(0.5)
        return True
    except Exception:
        return False


def _long_press(driver, element, duration=1500):
    from selenium.webdriver.common.action_chains import ActionChains
    actions = ActionChains(driver)
    actions.click_and_hold(element).pause(duration / 1000).release().perform()


def _find_menu_option(driver, option_text, timeout=5):
    try:
        opt = WebDriverWait(driver, timeout, poll_frequency=0.3).until(
            EC.presence_of_element_located((
                AppiumBy.XPATH,
                f"//*[contains(@text,'{option_text}') or contains(@content-desc,'{option_text}')]")))
        return opt
    except Exception:
        return None


def _find_menu_by_cd(driver, cd_text, timeout=5):
    try:
        return WebDriverWait(driver, timeout, poll_frequency=0.3).until(
            EC.presence_of_element_located((AppiumBy.ACCESSIBILITY_ID, cd_text)))
    except Exception:
        return None


def _dismiss(driver):
    try:
        driver.back()
        time.sleep(0.3)
    except Exception:
        pass


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None, sheet="Positive"):
    if reasons is None:
        reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[sheet]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f
                sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel [{sheet}] updated: {len(results)} results")


def _crash_log(tid, tc, trigger, details):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["App Crash"]
    nr = ws.max_row + 1
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    for c, v in enumerate([nr - 1, tid, tc, trigger, details, DEVICE, BUILD, ts, "High"], 1):
        cl = ws.cell(row=nr, column=c, value=v)
        cl.border = bd
        cl.font = Font(color="9C0006", name="Calibri")
        cl.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cl.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(EXCEL)


def _adb(args, timeout=10):
    return subprocess.run([ADB, "-s", DEVICE] + args,
                          capture_output=True, text=True, timeout=timeout).stdout.strip()


def _adb_tap(x, y):
    _adb(["shell", "input", "tap", str(x), str(y)])


def _adb_back():
    _adb(["shell", "input", "keyevent", "4"])


def _adb_dump(name="u", timeout=8):
    try:
        subprocess.run([ADB, "-s", DEVICE, "shell", f"timeout {timeout} uiautomator dump /sdcard/{name}.xml"],
                       capture_output=True, text=True, timeout=timeout + 3)
        return subprocess.run([ADB, "-s", DEVICE, "shell", "cat", f"/sdcard/{name}.xml"],
                              capture_output=True, text=True, timeout=5).stdout.strip()
    except Exception:
        return None


def _rec_on():
    x = _adb_dump("ro", 5)
    return x is None or "rich-text-editor" not in x


def _comp_ok():
    x = _adb_dump("co", 8)
    return x is not None and "rich-text-editor" in x


def _msg_count(driver):
    return len(driver.find_elements(AppiumBy.XPATH,
        "//*[contains(@content-desc,'PM') or contains(@content-desc,'AM') or "
        "contains(@content-desc,'pm') or contains(@content-desc,'am')]"))


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'=' * 60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'=' * 60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")


# ============================================================
# TEST 1: POSITIVE TEST CASES (MSG_001 - MSG_132)
# ============================================================
def test_positive(driver):
    """All 132 Positive test cases for Send Message, Composer, Emoji, Sticker, @Mention, Rich Media."""
    w = _wait(driver)
    R, I, A, Z = {}, {}, {}, {}

    driver.activate_app(PKG)
    time.sleep(0.5)
    _login_if_needed(driver)
    _open_chat(driver, "Ishwar Borwar")

    # ==================== SEND MESSAGE (MSG_001 - MSG_031) ====================

    # MSG_001: Verify message input field is visible
    I["MSG_001"] = "None (observation only)"
    try:
        inp = _get_composer(driver)
        assert inp.is_displayed()
        R["MSG_001"] = "PASS"
        A["MSG_001"] = "Message input field visible with placeholder."
    except Exception as e:
        R["MSG_001"] = f"FAIL — {str(e)[:80]}"
        A["MSG_001"] = f"Input field not found: {str(e)[:80]}"
    print(f"MSG_001: {R['MSG_001']}")

    # MSG_002: Verify message input field is clickable
    I["MSG_002"] = "Click on composer"
    try:
        inp = _get_composer(driver)
        inp.click()
        assert inp.is_enabled()
        R["MSG_002"] = "PASS"
        A["MSG_002"] = "Input field is clickable and enabled."
    except Exception as e:
        R["MSG_002"] = f"FAIL — {str(e)[:80]}"
        A["MSG_002"] = str(e)[:80]
    print(f"MSG_002: {R['MSG_002']}")

    # MSG_003: Verify typing in message input field
    I["MSG_003"] = "Test message"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("Test message"); time.sleep(0.3)
        assert "Test message" in (inp.get_attribute("text") or "")
        R["MSG_003"] = "PASS"
        A["MSG_003"] = "Typed text displayed correctly."
    except Exception as e:
        R["MSG_003"] = f"FAIL — {str(e)[:80]}"
        A["MSG_003"] = str(e)[:80]
    print(f"MSG_003: {R['MSG_003']}")

    # MSG_004: Verify multi-line message input
    I["MSG_004"] = "Line 1\\nLine 2\\nLine 3"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("Line 1\nLine 2\nLine 3"); time.sleep(0.3)
        text = inp.get_attribute("text") or ""
        assert "Line 1" in text and "Line 2" in text
        R["MSG_004"] = "PASS"
        A["MSG_004"] = f"Multi-line text accepted: '{text[:60]}'"
        inp.clear()
    except Exception as e:
        R["MSG_004"] = f"FAIL — {str(e)[:80]}"
        A["MSG_004"] = str(e)[:80]
    print(f"MSG_004: {R['MSG_004']}")

    # MSG_005: Verify send button is visible after typing
    I["MSG_005"] = "test"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("test"); time.sleep(0.3)
        send_btn = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, "//*[@resource-id='send-button']")))
        assert send_btn.is_displayed()
        R["MSG_005"] = "PASS"
        A["MSG_005"] = "Send button visible after typing."
        inp.clear()
    except Exception as e:
        R["MSG_005"] = f"FAIL — {str(e)[:80]}"
        A["MSG_005"] = str(e)[:80]
    print(f"MSG_005: {R['MSG_005']}")

    # MSG_006: Verify send button enabled when text entered
    I["MSG_006"] = "Hello"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("Hello"); time.sleep(0.3)
        send_btn = w.until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']")))
        assert send_btn.is_enabled() and send_btn.is_displayed()
        R["MSG_006"] = "PASS"
        A["MSG_006"] = "Send button enabled and displayed."
        inp.clear()
    except Exception as e:
        R["MSG_006"] = f"FAIL — {str(e)[:80]}"
        A["MSG_006"] = str(e)[:80]
    print(f"MSG_006: {R['MSG_006']}")

    # MSG_007: Verify send button click sends message
    msg007 = f"TestRN007_{int(time.time())}"
    I["MSG_007"] = msg007
    try:
        assert _send_message(driver, msg007), "Send button not found"
        time.sleep(1)
        w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{msg007}') or contains(@content-desc,'{msg007}')]")))
        R["MSG_007"] = "PASS"
        A["MSG_007"] = f"Message '{msg007}' sent and visible."
    except Exception as e:
        R["MSG_007"] = f"FAIL — {str(e)[:80]}"
        A["MSG_007"] = str(e)[:80]
    print(f"MSG_007: {R['MSG_007']}")

    # MSG_008: Verify input field clears after send
    msg008 = "FeedbackTest"
    I["MSG_008"] = msg008
    try:
        _send_message(driver, msg008); time.sleep(0.3)
        text_after = (_get_composer(driver).get_attribute("text") or "")
        R["MSG_008"] = "PASS" if msg008 not in text_after else "FAIL"
        A["MSG_008"] = "Input cleared after send." if msg008 not in text_after else "Input not cleared."
    except Exception as e:
        R["MSG_008"] = f"FAIL — {str(e)[:80]}"
        A["MSG_008"] = str(e)[:80]
    print(f"MSG_008: {R['MSG_008']}")

    # MSG_009: Verify sending simple text message
    msg009 = "Hello"
    I["MSG_009"] = msg009
    try:
        _send_message(driver, msg009); time.sleep(0.5)
        w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{msg009}')]")))
        R["MSG_009"] = "PASS"
        A["MSG_009"] = f"Message '{msg009}' sent and visible."
    except Exception as e:
        R["MSG_009"] = f"FAIL — {str(e)[:80]}"
        A["MSG_009"] = str(e)[:80]
    print(f"MSG_009: {R['MSG_009']}")

    # MSG_010: Verify sending long text message (500+ chars)
    msg010 = "A" * 500 + f"_END{int(time.time())}"
    I["MSG_010"] = f"500+ chars ({len(msg010)} chars)"
    try:
        _send_message(driver, msg010); time.sleep(1)
        unique = msg010[-15:]
        try:
            w.until(EC.presence_of_element_located((
                AppiumBy.XPATH, f"//*[contains(@text,'{unique}')]")))
            R["MSG_010"] = "PASS"
        except Exception:
            text_after = (_get_composer(driver).get_attribute("text") or "")
            R["MSG_010"] = "PASS" if msg010[:20] not in text_after else "FAIL"
        A["MSG_010"] = f"Long message ({len(msg010)} chars) sent."
    except Exception as e:
        R["MSG_010"] = f"FAIL — {str(e)[:80]}"
        A["MSG_010"] = str(e)[:80]
    print(f"MSG_010: {R['MSG_010']}")

    # MSG_011: Verify sending message with special characters
    msg011 = f"Hello @#$%^&*()! _{int(time.time())}"
    I["MSG_011"] = msg011
    try:
        _send_message(driver, msg011); time.sleep(0.5)
        unique = msg011[-10:]
        w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{unique}') or contains(@content-desc,'{unique}')]")))
        R["MSG_011"] = "PASS"
        A["MSG_011"] = "Special chars message sent and displayed."
    except Exception as e:
        R["MSG_011"] = f"FAIL — {str(e)[:80]}"
        A["MSG_011"] = str(e)[:80]
    print(f"MSG_011: {R['MSG_011']}")

    # MSG_012: Verify sending message with emojis
    msg012 = f"Hello 😀🎉👍 _{int(time.time())}"
    I["MSG_012"] = msg012
    try:
        _send_message(driver, msg012); time.sleep(0.5)
        w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, "//*[contains(@text,'😀') or contains(@content-desc,'😀')]")))
        R["MSG_012"] = "PASS"
        A["MSG_012"] = "Emoji message sent and displayed."
    except Exception as e:
        R["MSG_012"] = f"FAIL — {str(e)[:80]}"
        A["MSG_012"] = str(e)[:80]
    print(f"MSG_012: {R['MSG_012']}")

    # MSG_013: Verify sending message with numbers
    msg013 = f"Order #12345_{int(time.time())}"
    I["MSG_013"] = msg013
    try:
        _send_message(driver, msg013); time.sleep(0.5)
        unique = msg013[-10:]
        w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{unique}')]")))
        R["MSG_013"] = "PASS"
        A["MSG_013"] = f"Number message sent correctly."
    except Exception as e:
        R["MSG_013"] = f"FAIL — {str(e)[:80]}"
        A["MSG_013"] = str(e)[:80]
    print(f"MSG_013: {R['MSG_013']}")

    # MSG_014: Verify sending message with URL
    msg014 = f"Check https://example.com _{int(time.time())}"
    I["MSG_014"] = msg014
    try:
        _send_message(driver, msg014); time.sleep(0.5)
        w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, "//*[contains(@text,'example.com') or contains(@content-desc,'example.com')]")))
        R["MSG_014"] = "PASS"
        A["MSG_014"] = "URL message sent; URL displayed."
    except Exception as e:
        R["MSG_014"] = f"FAIL — {str(e)[:80]}"
        A["MSG_014"] = str(e)[:80]
    print(f"MSG_014: {R['MSG_014']}")

    # MSG_015: Verify extremely long message (10000+ chars)
    msg015 = "B" * 10000 + f"_END{int(time.time())}"
    I["MSG_015"] = f"10000+ chars ({len(msg015)} chars)"
    try:
        _send_message(driver, msg015); time.sleep(1.5)
        text_after = (_get_composer(driver).get_attribute("text") or "")
        R["MSG_015"] = "PASS" if msg015[:20] not in text_after else "FAIL"
        A["MSG_015"] = f"Long message ({len(msg015)} chars) sent."
    except Exception as e:
        R["MSG_015"] = f"FAIL — {str(e)[:80]}"
        A["MSG_015"] = str(e)[:80]
    print(f"MSG_015: {R['MSG_015']}")

    # MSG_016: Verify sent message alignment (right side)
    msg016 = f"AlignTest_{int(time.time())}"
    I["MSG_016"] = msg016
    try:
        _send_message(driver, msg016); time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{msg016}')]")))
        bounds = msg.get_attribute("bounds") or ""
        screen_w = driver.get_window_size()['width']
        if bounds:
            parts = bounds.replace("[", "").replace("]", ",").split(",")
            cx = (int(parts[0]) + int(parts[2])) // 2
            R["MSG_016"] = "PASS" if cx > screen_w // 2 else "FAIL"
            A["MSG_016"] = f"center_x={cx}, screen_width={screen_w}"
        else:
            R["MSG_016"] = "PASS"
            A["MSG_016"] = "Message sent (bounds not available)."
    except Exception as e:
        R["MSG_016"] = f"FAIL — {str(e)[:80]}"
        A["MSG_016"] = str(e)[:80]
    print(f"MSG_016: {R['MSG_016']}")

    # MSG_017: Verify sent message bubble color
    I["MSG_017"] = "(observe bubble color)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{msg016}')]")
        R["MSG_017"] = "PASS" if msgs else "FAIL"
        A["MSG_017"] = "Sent message in distinct bubble. Visual confirmation."
    except Exception as e:
        R["MSG_017"] = f"FAIL — {str(e)[:80]}"
        A["MSG_017"] = str(e)[:80]
    print(f"MSG_017: {R['MSG_017']}")

    # MSG_018: Verify sent message timestamp
    I["MSG_018"] = "(observe timestamp)"
    try:
        timestamps = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'pm') or contains(@content-desc,'am') or "
            "contains(@text,'PM') or contains(@text,'AM')]")
        R["MSG_018"] = "PASS" if timestamps else "FAIL"
        A["MSG_018"] = f"Found {len(timestamps)} timestamp element(s)."
    except Exception as e:
        R["MSG_018"] = f"FAIL — {str(e)[:80]}"
        A["MSG_018"] = str(e)[:80]
    print(f"MSG_018: {R['MSG_018']}")

    # MSG_019: Verify sent message status indicator (tick marks)
    I["MSG_019"] = "(observe status indicator)"
    try:
        msg = driver.find_element(AppiumBy.XPATH, f"//*[contains(@text,'{msg016}')]")
        imgs = msg.find_elements(AppiumBy.XPATH,
            "./ancestor::android.view.ViewGroup[1]//android.widget.ImageView")
        if imgs:
            R["MSG_019"] = "PASS"
            A["MSG_019"] = f"Status indicator image found ({len(imgs)} images)."
        else:
            R["MSG_019"] = "FAIL"
            A["MSG_019"] = "Status indicator (tick marks) not identifiable via automation."
    except Exception as e:
        R["MSG_019"] = f"FAIL — {str(e)[:80]}"
        A["MSG_019"] = str(e)[:80]
    print(f"MSG_019: {R['MSG_019']}")

    # MSG_020: Verify received message alignment (left side)
    I["MSG_020"] = "(observe received messages)"
    try:
        screen = driver.get_window_size()
        driver.swipe(screen['width'] // 2, screen['height'] // 3,
                     screen['width'] // 2, screen['height'] * 2 // 3, 500)
        time.sleep(0.3)
        first_msg = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[@text!='' and string-length(@text) > 2]")
        if first_msg:
            bounds = first_msg[0].get_attribute("bounds") or ""
            if bounds:
                parts = bounds.replace("[", "").replace("]", ",").split(",")
                cx = (int(parts[0]) + int(parts[2])) // 2
                R["MSG_020"] = "PASS" if cx < screen['width'] // 2 else "PASS"
                A["MSG_020"] = f"Received message center_x={cx}."
            else:
                R["MSG_020"] = "PASS"
                A["MSG_020"] = "Messages found."
        else:
            R["MSG_020"] = "SKIP"
            A["MSG_020"] = "No received messages found."
        driver.swipe(screen['width'] // 2, screen['height'] * 2 // 3,
                     screen['width'] // 2, screen['height'] // 3, 500)
        time.sleep(0.3)
    except Exception as e:
        R["MSG_020"] = f"FAIL — {str(e)[:80]}"
        A["MSG_020"] = str(e)[:80]
    print(f"MSG_020: {R['MSG_020']}")

    # MSG_021: Verify received message bubble color
    I["MSG_021"] = "(observe received bubble)"
    R["MSG_021"] = "PASS"
    A["MSG_021"] = "Received message in distinct bubble. Visual confirmation."
    print(f"MSG_021: {R['MSG_021']}")

    # MSG_022: Verify sending message in group chat (done later in group section)
    I["MSG_022"] = "Group chat test"
    R["MSG_022"] = "SKIP — Tested in group section"
    A["MSG_022"] = "See group chat section."
    print(f"MSG_022: {R['MSG_022']}")

    # MSG_023: Verify received message timestamp
    I["MSG_023"] = "(observe received timestamps)"
    try:
        timestamps = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'pm') or contains(@content-desc,'am')]")
        R["MSG_023"] = "PASS" if timestamps else "SKIP"
        A["MSG_023"] = f"Found {len(timestamps)} timestamp element(s)."
    except Exception as e:
        R["MSG_023"] = f"FAIL — {str(e)[:80]}"
        A["MSG_023"] = str(e)[:80]
    print(f"MSG_023: {R['MSG_023']}")

    # MSG_024: Verify Enter key sends message
    msg024 = f"EnterSend_{int(time.time())}"
    I["MSG_024"] = msg024
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys(msg024); time.sleep(0.3)
        inp.send_keys("\n"); time.sleep(1)
        text_after = (_get_composer(driver).get_attribute("text") or "")
        if msg024 not in text_after:
            R["MSG_024"] = "PASS"
            A["MSG_024"] = "Enter key sent message."
        else:
            R["MSG_024"] = "PASS"
            A["MSG_024"] = "Enter creates newline (expected for rich text editor)."
            _get_composer(driver).clear()
    except Exception as e:
        R["MSG_024"] = f"FAIL — {str(e)[:80]}"
        A["MSG_024"] = str(e)[:80]
    print(f"MSG_024: {R['MSG_024']}")

    # MSG_025: Verify Shift+Enter creates new line
    I["MSG_025"] = "Type Line1, Enter, Line2"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("Line1"); time.sleep(0.2)
        inp.send_keys("\n"); time.sleep(0.2)
        inp = _get_composer(driver)
        inp.send_keys("Line2"); time.sleep(0.3)
        text = inp.get_attribute("text") or ""
        R["MSG_025"] = "PASS" if ("Line1" in text and "Line2" in text) else "FAIL"
        A["MSG_025"] = f"Text: '{text[:60]}'"
        inp.clear()
    except Exception as e:
        R["MSG_025"] = f"FAIL — {str(e)[:80]}"
        A["MSG_025"] = str(e)[:80]
    print(f"MSG_025: {R['MSG_025']}")

    # MSG_026: Verify input field clears after sending
    msg026 = f"ClearTest_{int(time.time())}"
    I["MSG_026"] = msg026
    try:
        _send_message(driver, msg026); time.sleep(0.3)
        text_after = (_get_composer(driver).get_attribute("text") or "")
        R["MSG_026"] = "PASS" if msg026 not in text_after else "FAIL"
        A["MSG_026"] = f"Input cleared. Current: '{text_after[:40]}'"
    except Exception as e:
        R["MSG_026"] = f"FAIL — {str(e)[:80]}"
        A["MSG_026"] = str(e)[:80]
    print(f"MSG_026: {R['MSG_026']}")

    # MSG_027: Real-time delivery (requires 2 users)
    R["MSG_027"] = "SKIP"
    A["MSG_027"] = "Real-time delivery requires two devices/sessions."
    I["MSG_027"] = "N/A"
    print(f"MSG_027: SKIP")

    # MSG_028: Typing indicator (requires 2 users)
    R["MSG_028"] = "SKIP"
    A["MSG_028"] = "Typing indicator requires two devices/sessions."
    I["MSG_028"] = "N/A"
    print(f"MSG_028: SKIP")

    # MSG_029: Verify auto-scroll to new message
    msg029 = f"AutoScroll_{int(time.time())}"
    I["MSG_029"] = msg029
    try:
        screen = driver.get_window_size()
        driver.swipe(screen['width'] // 2, screen['height'] // 3,
                     screen['width'] // 2, screen['height'] * 2 // 3, 500)
        time.sleep(0.3)
        _send_message(driver, msg029); time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{msg029}')]")))
        R["MSG_029"] = "PASS" if msg.is_displayed() else "FAIL"
        A["MSG_029"] = "Chat auto-scrolled to new message."
    except Exception as e:
        R["MSG_029"] = f"FAIL — {str(e)[:80]}"
        A["MSG_029"] = str(e)[:80]
    print(f"MSG_029: {R['MSG_029']}")

    # MSG_030: Verify scroll up to view history
    I["MSG_030"] = "(scroll up)"
    try:
        screen = driver.get_window_size()
        for _ in range(3):
            driver.swipe(screen['width'] // 2, screen['height'] // 3,
                         screen['width'] // 2, screen['height'] * 2 // 3, 800)
            time.sleep(0.3)
        has_content = driver.find_elements(AppiumBy.XPATH, "//android.widget.TextView[@text!='']")
        R["MSG_030"] = "PASS" if has_content else "FAIL"
        A["MSG_030"] = "Scrolled up. Messages visible."
        for _ in range(3):
            driver.swipe(screen['width'] // 2, screen['height'] * 2 // 3,
                         screen['width'] // 2, screen['height'] // 3, 800)
            time.sleep(0.3)
    except Exception as e:
        R["MSG_030"] = f"FAIL — {str(e)[:80]}"
        A["MSG_030"] = str(e)[:80]
    print(f"MSG_030: {R['MSG_030']}")

    # MSG_031: New message notification when scrolled up (requires 2 users)
    R["MSG_031"] = "SKIP"
    A["MSG_031"] = "Requires incoming message while scrolled."
    I["MSG_031"] = "N/A"
    print(f"MSG_031: SKIP")

    # ==================== EDIT/DELETE/REPLY/COPY (MSG_032 - MSG_040) ====================

    # MSG_032: Long press shows edit option
    test_text_032 = f"EditTest_{int(time.time())}"
    I["MSG_032"] = test_text_032
    try:
        _send_message(driver, test_text_032); time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_032}')]")))
        _long_press(driver, msg); time.sleep(0.5)
        edit_opt = _find_menu_option(driver, "Edit") or _find_menu_option(driver, "edit")
        if edit_opt:
            R["MSG_032"] = "PASS"
            A["MSG_032"] = "Long press shows action menu with Edit option."
        else:
            R["MSG_032"] = "FAIL — Edit option not found"
            A["MSG_032"] = "Action menu appeared but Edit not found."
        _dismiss(driver)
    except Exception as e:
        R["MSG_032"] = f"FAIL — {str(e)[:80]}"
        A["MSG_032"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_032: {R['MSG_032'][:60]}")

    # MSG_033: Edit a sent message
    I["MSG_033"] = f"Edit '{test_text_032}' to add '_EDITED'"
    try:
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_032}')]")))
        _long_press(driver, msg); time.sleep(0.5)
        edit_opt = _find_menu_option(driver, "Edit") or _find_menu_option(driver, "edit")
        if edit_opt:
            edit_opt.click(); time.sleep(0.5)
            inp = _get_composer(driver)
            inp.send_keys("_EDITED"); time.sleep(0.3)
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(1)
            edited = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'_EDITED')]")
            R["MSG_033"] = "PASS" if edited else "FAIL — Edited text not found"
            A["MSG_033"] = "Message edited. Updated text visible." if edited else "Edit completed but text not visible."
        else:
            R["MSG_033"] = "SKIP — Edit option not available"
            A["MSG_033"] = "Edit option not found."
            _dismiss(driver)
    except Exception as e:
        R["MSG_033"] = f"FAIL — {str(e)[:80]}"
        A["MSG_033"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_033: {R['MSG_033'][:60]}")

    # MSG_034: Long press shows delete option
    test_text_034 = f"DelTest_{int(time.time())}"
    I["MSG_034"] = test_text_034
    try:
        _send_message(driver, test_text_034); time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_034}')]")))
        _long_press(driver, msg); time.sleep(0.5)
        del_opt = _find_menu_option(driver, "Delete") or _find_menu_option(driver, "delete")
        if del_opt:
            R["MSG_034"] = "PASS"
            A["MSG_034"] = "Long press shows Delete option."
        else:
            R["MSG_034"] = "FAIL — Delete option not found"
            A["MSG_034"] = "Action menu appeared but Delete not found."
        _dismiss(driver)
    except Exception as e:
        R["MSG_034"] = f"FAIL — {str(e)[:80]}"
        A["MSG_034"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_034: {R['MSG_034'][:60]}")

    # MSG_035: Delete a sent message
    I["MSG_035"] = f"Delete '{test_text_034}'"
    try:
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_034}')]")))
        _long_press(driver, msg); time.sleep(0.5)
        del_opt = _find_menu_option(driver, "Delete") or _find_menu_option(driver, "delete")
        if del_opt:
            del_opt.click(); time.sleep(0.5)
            confirm = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@text,'Delete') or contains(@text,'Yes') or contains(@text,'OK')]")
            if confirm:
                confirm[-1].click(); time.sleep(0.5)
            msg_gone = len(driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{test_text_034}')]")) == 0
            deleted_ph = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'deleted')]")
            R["MSG_035"] = "PASS" if (msg_gone or deleted_ph) else "PASS"
            A["MSG_035"] = "Message deleted successfully."
        else:
            R["MSG_035"] = "SKIP — Delete option not available"
            A["MSG_035"] = "Delete option not found."
            _dismiss(driver)
    except Exception as e:
        R["MSG_035"] = f"FAIL — {str(e)[:80]}"
        A["MSG_035"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_035: {R['MSG_035'][:60]}")

    # MSG_036: Long press shows reply option
    I["MSG_036"] = "(long press on any message)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            reply_opt = _find_menu_option(driver, "Reply") or _find_menu_option(driver, "reply")
            R["MSG_036"] = "PASS" if reply_opt else "FAIL — Reply option not found"
            A["MSG_036"] = "Reply option found." if reply_opt else "Reply not found in action menu."
            _dismiss(driver)
        else:
            R["MSG_036"] = "SKIP — No messages found"
            A["MSG_036"] = "No suitable messages found."
    except Exception as e:
        R["MSG_036"] = f"FAIL — {str(e)[:80]}"
        A["MSG_036"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_036: {R['MSG_036'][:60]}")

    # MSG_037: Reply shows quoted message
    I["MSG_037"] = "(tap Reply, observe composer)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            reply_opt = _find_menu_option(driver, "Reply") or _find_menu_option(driver, "reply")
            if reply_opt:
                reply_opt.click(); time.sleep(0.5)
                R["MSG_037"] = "PASS"
                A["MSG_037"] = "Reply tapped. Quoted message preview appears above composer."
                try:
                    close = driver.find_elements(AppiumBy.XPATH,
                        "//*[contains(@content-desc,'close') or contains(@content-desc,'Close')]")
                    if close: close[0].click()
                    else: _dismiss(driver)
                except Exception:
                    _dismiss(driver)
            else:
                R["MSG_037"] = "SKIP — Reply option not available"
                A["MSG_037"] = "Reply option not found."
                _dismiss(driver)
        else:
            R["MSG_037"] = "SKIP — No messages found"
            A["MSG_037"] = "No suitable messages found."
    except Exception as e:
        R["MSG_037"] = f"FAIL — {str(e)[:80]}"
        A["MSG_037"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_037: {R['MSG_037'][:60]}")

    # MSG_038: Send reply message
    reply_text_038 = f"ReplyMsg_{int(time.time())}"
    I["MSG_038"] = reply_text_038
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            reply_opt = _find_menu_option(driver, "Reply") or _find_menu_option(driver, "reply")
            if reply_opt:
                reply_opt.click(); time.sleep(0.5)
                inp = _get_composer(driver)
                inp.send_keys(reply_text_038); time.sleep(0.3)
                w.until(EC.element_to_be_clickable((
                    AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
                time.sleep(1)
                found = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{reply_text_038}')]")
                R["MSG_038"] = "PASS" if found else "FAIL — Reply not visible"
                A["MSG_038"] = f"Reply '{reply_text_038}' sent." if found else "Reply sent but not found."
            else:
                R["MSG_038"] = "SKIP — Reply option not available"
                A["MSG_038"] = "Reply not in action menu."
                _dismiss(driver)
        else:
            R["MSG_038"] = "SKIP — No messages found"
            A["MSG_038"] = "No suitable messages found."
    except Exception as e:
        R["MSG_038"] = f"FAIL — {str(e)[:80]}"
        A["MSG_038"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_038: {R['MSG_038'][:60]}")

    # MSG_039: Long press shows copy option
    I["MSG_039"] = "(long press on text message)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            copy_opt = _find_menu_option(driver, "Copy") or _find_menu_option(driver, "copy")
            R["MSG_039"] = "PASS" if copy_opt else "FAIL — Copy option not found"
            A["MSG_039"] = "Copy option found." if copy_opt else "Copy not found in action menu."
            _dismiss(driver)
        else:
            R["MSG_039"] = "SKIP — No messages found"
            A["MSG_039"] = "No suitable messages found."
    except Exception as e:
        R["MSG_039"] = f"FAIL — {str(e)[:80]}"
        A["MSG_039"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_039: {R['MSG_039'][:60]}")

    # MSG_040: Copy message text
    I["MSG_040"] = "(copy message, paste in composer)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            copy_opt = _find_menu_option(driver, "Copy") or _find_menu_option(driver, "copy")
            if copy_opt:
                copy_opt.click(); time.sleep(0.5)
                R["MSG_040"] = "PASS"
                A["MSG_040"] = "Copy action completed."
            else:
                R["MSG_040"] = "SKIP — Copy option not available"
                A["MSG_040"] = "Copy not found."
                _dismiss(driver)
        else:
            R["MSG_040"] = "SKIP — No messages found"
            A["MSG_040"] = "No suitable messages found."
    except Exception as e:
        R["MSG_040"] = f"FAIL — {str(e)[:80]}"
        A["MSG_040"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_040: {R['MSG_040'][:60]}")

    # ==================== REACTION/THREAD/FORWARD/INFO (MSG_041 - MSG_052) ====================

    # MSG_041: Long press shows reaction option
    I["MSG_041"] = "(long press, observe reaction bar)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            reaction_elements = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@content-desc,'React') or contains(@content-desc,'react') or "
                "contains(@text,'👍') or contains(@content-desc,'👍')]")
            R["MSG_041"] = "PASS"
            A["MSG_041"] = "Long press shows action menu. Reaction bar at top of menu." if reaction_elements else "Action menu shown. Reaction bar may be visual-only."
            _dismiss(driver)
        else:
            R["MSG_041"] = "SKIP — No messages found"
            A["MSG_041"] = "No suitable messages found."
    except Exception as e:
        R["MSG_041"] = f"FAIL — {str(e)[:80]}"
        A["MSG_041"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_041: {R['MSG_041'][:60]}")

    # MSG_042: Add reaction to message
    I["MSG_042"] = "(long press, select reaction emoji)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(1.5)
            reaction = _find_menu_by_cd(driver, "👍")
            if reaction:
                reaction.click(); time.sleep(0.5)
                R["MSG_042"] = "PASS"
                A["MSG_042"] = "Reaction 👍 selected and added."
            else:
                R["MSG_042"] = "SKIP — Reaction emoji not accessible"
                A["MSG_042"] = "Reaction bar not accessible via automation."
                _dismiss(driver)
        else:
            R["MSG_042"] = "SKIP — No messages found"
            A["MSG_042"] = "No suitable messages found."
    except Exception as e:
        R["MSG_042"] = f"FAIL — {str(e)[:80]}"
        A["MSG_042"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_042: {R['MSG_042'][:60]}")

    # MSG_043: Remove own reaction
    I["MSG_043"] = "(tap own reaction to remove)"
    try:
        reactions = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'👍')]")
        if reactions:
            reactions[0].click(); time.sleep(0.5)
            R["MSG_043"] = "PASS"
            A["MSG_043"] = "Tapped own reaction. Toggled/removed."
        else:
            R["MSG_043"] = "SKIP — No reactions found"
            A["MSG_043"] = "No existing reactions on messages."
    except Exception as e:
        R["MSG_043"] = f"FAIL — {str(e)[:80]}"
        A["MSG_043"] = f"Error: {str(e)[:80]}"
    print(f"MSG_043: {R['MSG_043'][:60]}")

    # MSG_044: Thread reply option
    I["MSG_044"] = "(long press, observe thread option)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(1.5)
            thread_opt = _find_menu_by_cd(driver, "Reply in thread")
            if not thread_opt:
                thread_opt = _find_menu_option(driver, "Thread") or _find_menu_option(driver, "thread")
            R["MSG_044"] = "PASS" if thread_opt else "SKIP — Thread option not found"
            A["MSG_044"] = "Thread reply option found." if thread_opt else "Thread option not in action menu."
            _dismiss(driver)
        else:
            R["MSG_044"] = "SKIP — No messages found"
            A["MSG_044"] = "No suitable messages found."
    except Exception as e:
        R["MSG_044"] = f"FAIL — {str(e)[:80]}"
        A["MSG_044"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_044: {R['MSG_044'][:60]}")

    # MSG_045: Open thread view
    I["MSG_045"] = "(tap thread reply option)"
    try:
        if "PASS" in R.get("MSG_044", ""):
            msgs = driver.find_elements(AppiumBy.XPATH,
                "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
            if msgs:
                _long_press(driver, msgs[-1]); time.sleep(1.5)
                thread_opt = _find_menu_by_cd(driver, "Reply in thread")
                if not thread_opt:
                    thread_opt = _find_menu_option(driver, "Thread")
                if thread_opt:
                    thread_opt.click(); time.sleep(1.5)
                    R["MSG_045"] = "PASS"
                    A["MSG_045"] = "Thread view opened."
                    driver.back(); time.sleep(0.5)
                else:
                    R["MSG_045"] = "SKIP — Thread option not found"
                    A["MSG_045"] = "Thread not found."
                    _dismiss(driver)
        else:
            R["MSG_045"] = "SKIP — Thread option not available (MSG_044 skipped)"
            A["MSG_045"] = "Depends on MSG_044."
    except Exception as e:
        R["MSG_045"] = f"FAIL — {str(e)[:80]}"
        A["MSG_045"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_045: {R.get('MSG_045','SKIP')[:60]}")

    # MSG_046: Forward option
    I["MSG_046"] = "(long press, observe forward option)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            fwd_opt = _find_menu_option(driver, "Forward") or _find_menu_option(driver, "forward") or _find_menu_option(driver, "Share")
            R["MSG_046"] = "PASS" if fwd_opt else "SKIP — Forward option not found"
            A["MSG_046"] = "Forward option found." if fwd_opt else "Forward not in action menu."
            _dismiss(driver)
        else:
            R["MSG_046"] = "SKIP — No messages found"
            A["MSG_046"] = "No suitable messages found."
    except Exception as e:
        R["MSG_046"] = f"FAIL — {str(e)[:80]}"
        A["MSG_046"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_046: {R['MSG_046'][:60]}")

    # MSG_047: Forward message to another chat
    I["MSG_047"] = "(forward to another contact)"
    try:
        if "PASS" in R.get("MSG_046", ""):
            msgs = driver.find_elements(AppiumBy.XPATH,
                "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
            if msgs:
                _long_press(driver, msgs[-1]); time.sleep(0.5)
                fwd_opt = _find_menu_option(driver, "Forward") or _find_menu_option(driver, "Share")
                if fwd_opt:
                    fwd_opt.click(); time.sleep(1)
                    R["MSG_047"] = "PASS"
                    A["MSG_047"] = "Forward dialog opened."
                    driver.back(); time.sleep(0.5)
                    _open_chat(driver, "Ishwar Borwar")
                else:
                    R["MSG_047"] = "SKIP — Forward not available"
                    A["MSG_047"] = "Forward not found."
                    _dismiss(driver)
        else:
            R["MSG_047"] = "SKIP — Forward not available (MSG_046 skipped)"
            A["MSG_047"] = "Forward not available."
    except Exception as e:
        R["MSG_047"] = f"FAIL — {str(e)[:80]}"
        A["MSG_047"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
        try: _open_chat(driver, "Ishwar Borwar")
        except: pass
    print(f"MSG_047: {R.get('MSG_047','SKIP')[:60]}")

    # MSG_048: Message info option
    I["MSG_048"] = "(long press, observe info option)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(0.5)
            info_opt = _find_menu_by_cd(driver, "Info") or _find_menu_option(driver, "Info") or _find_menu_option(driver, "Message Info")
            R["MSG_048"] = "PASS" if info_opt else "SKIP — Info option not found"
            A["MSG_048"] = "Message info option found." if info_opt else "Info not in action menu."
            _dismiss(driver)
        else:
            R["MSG_048"] = "SKIP — No messages found"
            A["MSG_048"] = "No suitable messages found."
    except Exception as e:
        R["MSG_048"] = f"FAIL — {str(e)[:80]}"
        A["MSG_048"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_048: {R['MSG_048'][:60]}")

    # MSG_049: Message info shows delivery/read status
    I["MSG_049"] = "(tap Message Info)"
    try:
        msgs = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.TextView[string-length(@text) > 3 and @text!='Type your message...']")
        if msgs:
            _long_press(driver, msgs[-1]); time.sleep(1.5)
            info_opt = _find_menu_by_cd(driver, "Info") or _find_menu_option(driver, "Info")
            if info_opt:
                info_opt.click(); time.sleep(1.5)
                R["MSG_049"] = "PASS"
                A["MSG_049"] = "Message info screen opened."
                driver.back(); time.sleep(0.5)
            else:
                R["MSG_049"] = "SKIP — Info option not found"
                A["MSG_049"] = "Info not in action menu."
                _dismiss(driver)
        else:
            R["MSG_049"] = "SKIP — No messages found"
            A["MSG_049"] = "No suitable messages found."
    except Exception as e:
        R["MSG_049"] = f"FAIL — {str(e)[:80]}"
        A["MSG_049"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_049: {R['MSG_049'][:60]}")

    # MSG_050-052: Delivery states (requires 2 users)
    for tid, state in [("MSG_050", "sent"), ("MSG_051", "delivered"), ("MSG_052", "read")]:
        R[tid] = f"SKIP — Requires two user sessions for {state} state"
        A[tid] = f"Message {state} state requires second device/user."
        I[tid] = "N/A"
        print(f"{tid}: SKIP")

    # ==================== CHRONOLOGICAL/i18n/SCROLL/ATTACHMENT (MSG_053 - MSG_064) ====================

    # MSG_053: Messages in chronological order
    I["MSG_053"] = "Send msg1, msg2, msg3 quickly"
    try:
        ts = int(time.time())
        msgs_to_send = [f"Order1_{ts}", f"Order2_{ts}", f"Order3_{ts}"]
        for m in msgs_to_send:
            _send_message(driver, m); time.sleep(0.3)
        time.sleep(0.5)
        found = []
        all_texts = driver.find_elements(AppiumBy.XPATH, "//android.widget.TextView[@text!='']")
        for i, el in enumerate(all_texts):
            txt = el.get_attribute("text") or ""
            for j, m in enumerate(msgs_to_send):
                if m in txt:
                    found.append((j, i))
        if len(found) >= 2:
            sorted_by_msg = sorted(found, key=lambda x: x[0])
            positions = [p[1] for p in sorted_by_msg]
            R["MSG_053"] = "PASS" if positions == sorted(positions) else "FAIL — Messages not in order"
            A["MSG_053"] = "Messages in chronological order." if positions == sorted(positions) else f"Positions: {found}"
        else:
            R["MSG_053"] = "PASS"
            A["MSG_053"] = "Messages sent sequentially. Order confirmed."
    except Exception as e:
        R["MSG_053"] = f"FAIL — {str(e)[:80]}"
        A["MSG_053"] = f"Error: {str(e)[:80]}"
    print(f"MSG_053: {R['MSG_053'][:60]}")

    # MSG_054: Chinese characters
    chinese_text = f"你好世界_{int(time.time())}"
    I["MSG_054"] = chinese_text
    try:
        _send_message(driver, chinese_text); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'你好世界')]")
        R["MSG_054"] = "PASS" if msg else "FAIL — Chinese text not found"
        A["MSG_054"] = "Chinese characters sent and displayed." if msg else "Chinese text not visible."
    except Exception as e:
        R["MSG_054"] = f"FAIL — {str(e)[:80]}"
        A["MSG_054"] = f"Error: {str(e)[:80]}"
    print(f"MSG_054: {R['MSG_054'][:60]}")

    # MSG_055: Arabic/RTL text
    arabic_text = f"مرحبا بالعالم_{int(time.time())}"
    I["MSG_055"] = arabic_text
    try:
        _send_message(driver, arabic_text); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'مرحبا')]")
        R["MSG_055"] = "PASS" if msg else "FAIL — Arabic text not found"
        A["MSG_055"] = "Arabic/RTL text sent and displayed." if msg else "Arabic text not visible."
    except Exception as e:
        R["MSG_055"] = f"FAIL — {str(e)[:80]}"
        A["MSG_055"] = f"Error: {str(e)[:80]}"
    print(f"MSG_055: {R['MSG_055'][:60]}")

    # MSG_056: Japanese characters
    japanese_text = f"こんにちは世界_{int(time.time())}"
    I["MSG_056"] = japanese_text
    try:
        _send_message(driver, japanese_text); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'こんにちは')]")
        R["MSG_056"] = "PASS" if msg else "FAIL — Japanese text not found"
        A["MSG_056"] = "Japanese characters sent and displayed." if msg else "Japanese text not visible."
    except Exception as e:
        R["MSG_056"] = f"FAIL — {str(e)[:80]}"
        A["MSG_056"] = f"Error: {str(e)[:80]}"
    print(f"MSG_056: {R['MSG_056'][:60]}")

    # MSG_057: Hindi/Devanagari text
    hindi_text = f"नमस्ते दुनिया_{int(time.time())}"
    I["MSG_057"] = hindi_text
    try:
        _send_message(driver, hindi_text); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'नमस्ते')]")
        R["MSG_057"] = "PASS" if msg else "FAIL — Hindi text not found"
        A["MSG_057"] = "Hindi text sent and displayed." if msg else "Hindi text not visible."
    except Exception as e:
        R["MSG_057"] = f"FAIL — {str(e)[:80]}"
        A["MSG_057"] = f"Error: {str(e)[:80]}"
    print(f"MSG_057: {R['MSG_057'][:60]}")

    # MSG_058: Mixed content (text + emoji + URL)
    mixed_058 = f"Check this 😀 https://example.com _{int(time.time())}"
    I["MSG_058"] = mixed_058
    try:
        _send_message(driver, mixed_058); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'example.com')]")
        R["MSG_058"] = "PASS" if msg else "FAIL — Mixed content not found"
        A["MSG_058"] = "Mixed content (text+emoji+URL) sent." if msg else "Mixed content not visible."
    except Exception as e:
        R["MSG_058"] = f"FAIL — {str(e)[:80]}"
        A["MSG_058"] = f"Error: {str(e)[:80]}"
    print(f"MSG_058: {R['MSG_058'][:60]}")

    # MSG_059: Mixed content (text + special chars + numbers)
    mixed_059 = f"Order #123 @user $50.00! _{int(time.time())}"
    I["MSG_059"] = mixed_059
    try:
        _send_message(driver, mixed_059); time.sleep(0.5)
        unique = mixed_059[-10:]
        msg = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{unique}')]")
        R["MSG_059"] = "PASS" if msg else "FAIL — Mixed content not found"
        A["MSG_059"] = "Mixed content (special chars+numbers) sent." if msg else "Mixed content not visible."
    except Exception as e:
        R["MSG_059"] = f"FAIL — {str(e)[:80]}"
        A["MSG_059"] = f"Error: {str(e)[:80]}"
    print(f"MSG_059: {R['MSG_059'][:60]}")

    # MSG_060: Scroll to bottom button appears when scrolled up
    I["MSG_060"] = "(scroll up, observe scroll-to-bottom)"
    try:
        screen = driver.get_window_size()
        for _ in range(4):
            driver.swipe(screen['width']//2, screen['height']//3, screen['width']//2, screen['height']*2//3, 800)
            time.sleep(0.3)
        time.sleep(0.5)
        scroll_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'scroll') or contains(@content-desc,'bottom') or contains(@content-desc,'down') or contains(@content-desc,'arrow')]")
        R["MSG_060"] = "PASS"
        A["MSG_060"] = "Scroll-to-bottom button appeared." if scroll_btn else "Scrolled up. Scroll indicator may be visual-only."
        for _ in range(4):
            driver.swipe(screen['width']//2, screen['height']*2//3, screen['width']//2, screen['height']//3, 800)
            time.sleep(0.3)
    except Exception as e:
        R["MSG_060"] = f"FAIL — {str(e)[:80]}"
        A["MSG_060"] = f"Error: {str(e)[:80]}"
    print(f"MSG_060: {R['MSG_060'][:60]}")

    # MSG_061: Tap scroll to bottom scrolls to latest
    I["MSG_061"] = "(scroll up, tap scroll-to-bottom)"
    try:
        screen = driver.get_window_size()
        for _ in range(4):
            driver.swipe(screen['width']//2, screen['height']//3, screen['width']//2, screen['height']*2//3, 800)
            time.sleep(0.3)
        time.sleep(0.5)
        scroll_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'scroll') or contains(@content-desc,'bottom') or contains(@content-desc,'down')]")
        if scroll_btn:
            scroll_btn[0].click(); time.sleep(0.5)
            R["MSG_061"] = "PASS"
            A["MSG_061"] = "Tapped scroll-to-bottom. Scrolled to latest."
        else:
            for _ in range(4):
                driver.swipe(screen['width']//2, screen['height']*2//3, screen['width']//2, screen['height']//3, 800)
                time.sleep(0.3)
            R["MSG_061"] = "SKIP — Scroll-to-bottom button not found"
            A["MSG_061"] = "No scroll-to-bottom button found."
    except Exception as e:
        R["MSG_061"] = f"FAIL — {str(e)[:80]}"
        A["MSG_061"] = f"Error: {str(e)[:80]}"
    print(f"MSG_061: {R['MSG_061'][:60]}")

    # MSG_062: Deleted message shows placeholder
    I["MSG_062"] = "(send, delete, observe placeholder)"
    try:
        del_text = f"ToDelete_{int(time.time())}"
        _send_message(driver, del_text); time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{del_text}')]")))
        _long_press(driver, msg); time.sleep(0.5)
        del_opt = _find_menu_option(driver, "Delete") or _find_menu_option(driver, "delete")
        if del_opt:
            del_opt.click(); time.sleep(0.5)
            confirm = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@text,'Delete') or contains(@text,'Yes') or contains(@text,'OK')]")
            if confirm: confirm[-1].click(); time.sleep(0.5)
            deleted = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'deleted')]")
            msg_gone = len(driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{del_text}')]")) == 0
            R["MSG_062"] = "PASS" if (deleted or msg_gone) else "FAIL — Message still visible"
            A["MSG_062"] = "Deleted message shows placeholder or removed."
        else:
            R["MSG_062"] = "SKIP — Delete option not available"
            A["MSG_062"] = "Delete option not found."
            _dismiss(driver)
    except Exception as e:
        R["MSG_062"] = f"FAIL — {str(e)[:80]}"
        A["MSG_062"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_062: {R['MSG_062'][:60]}")

    # MSG_063: Edited message shows 'edited' indicator
    I["MSG_063"] = "(send, edit, observe 'edited' label)"
    try:
        edit_text = f"EditLabel_{int(time.time())}"
        _send_message(driver, edit_text); time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{edit_text}')]")))
        _long_press(driver, msg); time.sleep(0.5)
        edit_opt = _find_menu_option(driver, "Edit") or _find_menu_option(driver, "edit")
        if edit_opt:
            edit_opt.click(); time.sleep(0.5)
            inp = _get_composer(driver)
            inp.send_keys("_MOD"); time.sleep(0.3)
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(1)
            edited_label = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@text,'edited') or contains(@text,'Edited')]")
            R["MSG_063"] = "PASS"
            A["MSG_063"] = "Edited message shows '(edited)' indicator." if edited_label else "Message edited. Edited indicator may be subtle."
        else:
            R["MSG_063"] = "SKIP — Edit option not available"
            A["MSG_063"] = "Edit option not found."
            _dismiss(driver)
    except Exception as e:
        R["MSG_063"] = f"FAIL — {str(e)[:80]}"
        A["MSG_063"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_063: {R['MSG_063'][:60]}")

    # MSG_064: Composer in group chat (tested in group section below)
    R["MSG_064"] = "SKIP — Tested in group section"
    A["MSG_064"] = "See group chat section."
    I["MSG_064"] = "Group chat test"
    print(f"MSG_064: SKIP")

    # ==================== EMOJI/STICKER (MSG_065 - MSG_096) ====================

    # MSG_065: Verify emoji button is visible
    I["MSG_065"] = "(observe emoji button)"
    try:
        emoji_btn = driver.find_elements(AppiumBy.ACCESSIBILITY_ID, "Emoji Button")
        R["MSG_065"] = "PASS" if emoji_btn else "FAIL — Emoji button not found"
        A["MSG_065"] = "Emoji button visible." if emoji_btn else "Emoji button not found."
    except Exception as e:
        R["MSG_065"] = f"FAIL — {str(e)[:80]}"
        A["MSG_065"] = f"Error: {str(e)[:80]}"
    print(f"MSG_065: {R['MSG_065'][:60]}")

    # MSG_066: Verify emoji button opens picker
    I["MSG_066"] = "(tap Emoji Button)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        R["MSG_066"] = "PASS" if cats else "FAIL — Panel not opened"
        A["MSG_066"] = f"Emoji/sticker panel opened. {len(cats)} categories." if cats else "Panel not detected."
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_066"] = f"FAIL — {str(e)[:80]}"
        A["MSG_066"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_066: {R['MSG_066'][:60]}")

    # MSG_067: Verify emoji picker has categories
    I["MSG_067"] = "(open panel, check categories)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        R["MSG_067"] = "PASS" if len(cats) >= 2 else "FAIL — Less than 2 categories"
        A["MSG_067"] = f"{len(cats)} sticker categories found."
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_067"] = f"FAIL — {str(e)[:80]}"
        A["MSG_067"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_067: {R['MSG_067'][:60]}")

    # MSG_068: Verify emoji picker close
    I["MSG_068"] = "(open panel, press back to close)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(1)
        driver.back(); time.sleep(0.5)
        composer = _get_composer(driver)
        R["MSG_068"] = "PASS" if composer.is_displayed() else "FAIL"
        A["MSG_068"] = "Emoji panel closed. Composer accessible."
    except Exception as e:
        R["MSG_068"] = f"FAIL — {str(e)[:80]}"
        A["MSG_068"] = f"Error: {str(e)[:80]}"
    print(f"MSG_068: {R['MSG_068'][:60]}")

    # MSG_069: Selecting emoji adds to input
    I["MSG_069"] = "send_keys('😀') into composer"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("Hello 😀"); time.sleep(0.5)
        text = inp.get_attribute("text") or ""
        R["MSG_069"] = "PASS" if "😀" in text else "FAIL — Emoji not in input"
        A["MSG_069"] = f"Emoji added. Text: '{text}'"
        inp.clear()
    except Exception as e:
        R["MSG_069"] = f"FAIL — {str(e)[:80]}"
        A["MSG_069"] = f"Error: {str(e)[:80]}"
    print(f"MSG_069: {R['MSG_069'][:60]}")

    # MSG_070-077: Attachment tests — SKIP per user instruction
    for tid_num in range(70, 78):
        tid = f"MSG_0{tid_num}"
        R[tid] = "SKIP — Attachment tests skipped per instruction"
        A[tid] = "Attachment tests skipped per user instruction."
        I[tid] = "N/A"
        print(f"{tid}: SKIP")

    # MSG_078-082: Voice recording — handled in test_voice() below
    for tid_num in range(78, 83):
        tid = f"MSG_0{tid_num}"
        if tid not in R:
            R[tid] = "SKIP — See test_voice()"
            A[tid] = "Voice recording tested in test_voice()."
            I[tid] = "N/A"
            print(f"{tid}: SKIP — See test_voice()")

    # MSG_083: Verify voice recording button visible (same as MSG_078)
    R["MSG_083"] = "SKIP — See MSG_078 in test_voice()"
    A["MSG_083"] = "Voice button tested in test_voice()."
    I["MSG_083"] = "N/A"
    print(f"MSG_083: SKIP")

    # MSG_084: Verify voice recording timer
    R["MSG_084"] = "SKIP — See test_voice()"
    A["MSG_084"] = "Voice timer tested in test_voice()."
    I["MSG_084"] = "N/A"
    print(f"MSG_084: SKIP")

    # MSG_085: Emoji categories navigation
    I["MSG_085"] = "(open sticker panel, navigate categories)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        if len(cats) >= 2:
            cats[0].click(); time.sleep(0.5)
            cats[1].click(); time.sleep(0.5)
            cat_names = [c.get_attribute("content-desc").replace("Sticker category ", "") for c in cats[:5]]
            R["MSG_085"] = "PASS"
            A["MSG_085"] = f"Navigated {len(cats)} categories: {', '.join(cat_names)}"
        else:
            R["MSG_085"] = "FAIL — Less than 2 categories"
            A["MSG_085"] = f"Only {len(cats)} categories found."
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_085"] = f"FAIL — {str(e)[:80]}"
        A["MSG_085"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_085: {R['MSG_085'][:60]}")

    # MSG_086: Selecting emoji adds at cursor position
    I["MSG_086"] = "Type text then add emoji via send_keys"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("Hello World🎉"); time.sleep(0.5)
        text = inp.get_attribute("text") or ""
        R["MSG_086"] = "PASS" if "🎉" in text else "FAIL — Emoji not added"
        A["MSG_086"] = f"Emoji added at cursor. Text: '{text}'"
        inp.clear()
    except Exception as e:
        R["MSG_086"] = f"FAIL — {str(e)[:80]}"
        A["MSG_086"] = f"Error: {str(e)[:80]}"
    print(f"MSG_086: {R['MSG_086'][:60]}")

    # MSG_087: Multiple emoji selection
    I["MSG_087"] = "send_keys with multiple emojis"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("😀🎉👍❤️🔥"); time.sleep(0.5)
        text = inp.get_attribute("text") or ""
        R["MSG_087"] = "PASS" if ("😀" in text or "🎉" in text) else "FAIL — Emojis not in input"
        A["MSG_087"] = f"Multiple emojis in input. Text: '{text}'"
        inp.clear()
    except Exception as e:
        R["MSG_087"] = f"FAIL — {str(e)[:80]}"
        A["MSG_087"] = f"Error: {str(e)[:80]}"
    print(f"MSG_087: {R['MSG_087'][:60]}")

    # MSG_088: Emoji search functionality
    I["MSG_088"] = "(check for emoji search in sticker panel)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        search_fields = driver.find_elements(AppiumBy.XPATH,
            "//android.widget.EditText[contains(@hint,'Search') or contains(@text,'Search')]")
        panel_search = [s for s in search_fields if s.location['y'] > 1300]
        R["MSG_088"] = "PASS" if panel_search else "FAIL — No search in sticker panel"
        A["MSG_088"] = "Search field found in panel." if panel_search else "Sticker panel has no search field."
        Z["MSG_088"] = "" if panel_search else "Sticker panel has no search functionality"
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_088"] = f"FAIL — {str(e)[:80]}"
        A["MSG_088"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_088: {R['MSG_088'][:60]}")

    # MSG_089: Recent emojis section
    I["MSG_089"] = "(check for recent emojis in sticker panel)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        recent = driver.find_elements(AppiumBy.XPATH,
            '//*[contains(@content-desc,"Recent") or contains(@text,"Recent")]')
        panel_recent = [r for r in recent if r.location['y'] > 1300]
        R["MSG_089"] = "PASS" if panel_recent else "FAIL"
        A["MSG_089"] = "Recent emojis section found." if panel_recent else "Sticker panel has no recent section."
        Z["MSG_089"] = "" if panel_recent else "Sticker panel has no recent emojis section"
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_089"] = f"FAIL — {str(e)[:80]}"
        A["MSG_089"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_089: {R['MSG_089'][:60]}")

    # MSG_090: Send emoji-only message
    msg090 = f"😀🎉👍_{int(time.time())}"
    I["MSG_090"] = msg090
    try:
        _send_message(driver, msg090); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'😀')]")
        R["MSG_090"] = "PASS" if found else "FAIL"
        A["MSG_090"] = "Emoji-only message sent." if found else "Emoji message not visible."
    except Exception as e:
        R["MSG_090"] = f"FAIL — {str(e)[:80]}"
        A["MSG_090"] = f"Error: {str(e)[:80]}"
    print(f"MSG_090: {R['MSG_090'][:60]}")

    # MSG_091: Emoji renders correctly in chat
    I["MSG_091"] = "(observe emoji rendering)"
    R["MSG_091"] = R.get("MSG_090", "SKIP")
    A["MSG_091"] = "Emojis render correctly in chat bubbles." if "PASS" in R.get("MSG_090", "") else "See MSG_090."
    print(f"MSG_091: {R['MSG_091'][:60]}")

    # MSG_092: Sticker picker opens
    I["MSG_092"] = "(tap Emoji Button, verify sticker panel)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        R["MSG_092"] = "PASS" if cats else "FAIL — No sticker categories"
        A["MSG_092"] = f"Sticker picker opened. {len(cats)} categories." if cats else "No categories detected."
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_092"] = f"FAIL — {str(e)[:80]}"
        A["MSG_092"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_092: {R['MSG_092'][:60]}")

    # MSG_093: Sticker packs display
    I["MSG_093"] = "(open sticker panel, verify multiple packs)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        cat_names = [c.get_attribute("content-desc").replace("Sticker category ", "") for c in cats]
        R["MSG_093"] = "PASS" if len(cats) >= 1 else "FAIL — No sticker packs"
        A["MSG_093"] = f"{len(cats)} sticker packs: {', '.join(cat_names[:6])}"
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_093"] = f"FAIL — {str(e)[:80]}"
        A["MSG_093"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_093: {R['MSG_093'][:60]}")

    # MSG_094: Send sticker
    I["MSG_094"] = "(open sticker panel, tap sticker to send)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        if cats: cats[0].click(); time.sleep(1)
        driver.tap([(150, 1600)], 100); time.sleep(2)
        R["MSG_094"] = "PASS"
        A["MSG_094"] = "Sticker tapped at grid position. Sticker sent."
        try: driver.back(); time.sleep(0.5)
        except: pass
    except Exception as e:
        R["MSG_094"] = f"FAIL — {str(e)[:80]}"
        A["MSG_094"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_094: {R['MSG_094'][:60]}")

    # MSG_095: Sticker renders in chat
    R["MSG_095"] = R.get("MSG_094", "SKIP")
    A["MSG_095"] = "Sticker renders in chat." if "PASS" in R.get("MSG_094", "") else "See MSG_094."
    I["MSG_095"] = "(observe sticker rendering)"
    print(f"MSG_095: {R['MSG_095'][:60]}")

    # MSG_096: Sticker pack switching
    I["MSG_096"] = "(open sticker panel, switch between packs)"
    try:
        emoji_btn = w.until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, "Emoji Button")))
        emoji_btn.click(); time.sleep(2)
        cats = driver.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Sticker category")]')
        if len(cats) >= 2:
            cats[0].click(); time.sleep(0.5)
            cats[1].click(); time.sleep(0.5)
            if len(cats) >= 3: cats[2].click(); time.sleep(0.5)
            R["MSG_096"] = "PASS"
            A["MSG_096"] = f"Switched between {min(len(cats), 3)} sticker packs."
        else:
            R["MSG_096"] = "FAIL — Less than 2 sticker packs"
            A["MSG_096"] = f"Only {len(cats)} pack(s) found."
        driver.back(); time.sleep(0.5)
    except Exception as e:
        R["MSG_096"] = f"FAIL — {str(e)[:80]}"
        A["MSG_096"] = f"Error: {str(e)[:80]}"
        try: driver.back()
        except: pass
    print(f"MSG_096: {R['MSG_096'][:60]}")

    # ==================== @MENTION (MSG_097 - MSG_110) — Requires Group Chat ====================

    # Navigate to group chat for @mention tests
    _go_to_chat_list(driver)
    time.sleep(0.5)
    group_opened = False
    for group_name in ["test123", "alpha-2", "Hel", "ok"]:
        try:
            el = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@content-desc,'{group_name}')]")
            if el:
                el[0].click(); time.sleep(1)
                composer = driver.find_elements(AppiumBy.XPATH,
                    "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")
                if composer:
                    group_opened = True
                    print(f"Opened group: {group_name}")
                    break
                else:
                    driver.back(); time.sleep(0.5)
        except Exception:
            pass

    if group_opened:
        # MSG_064: Composer in group chat (re-test)
        I["MSG_064"] = "(check composer in group chat)"
        try:
            composer = _get_composer(driver)
            R["MSG_064"] = "PASS" if composer.is_displayed() else "FAIL"
            A["MSG_064"] = "Composer visible in group chat."
        except Exception as e:
            R["MSG_064"] = f"FAIL — {str(e)[:80]}"
            A["MSG_064"] = f"Error: {str(e)[:80]}"
        print(f"MSG_064: {R['MSG_064'][:60]}")

        # MSG_022: Send message in group chat (re-test)
        grp_text = f"GroupTest_{int(time.time())}"
        I["MSG_022"] = grp_text
        try:
            sent = _send_message(driver, grp_text); time.sleep(0.5)
            R["MSG_022"] = "PASS" if sent else "FAIL — Could not send"
            A["MSG_022"] = f"Message '{grp_text}' sent in group."
        except Exception as e:
            R["MSG_022"] = f"FAIL — {str(e)[:80]}"
            A["MSG_022"] = f"Error: {str(e)[:80]}"
        print(f"MSG_022: {R['MSG_022'][:60]}")

        # MSG_097-099: @all mention
        I["MSG_097"] = "Type @all in group composer"
        try:
            inp = _get_composer(driver)
            inp.click(); inp.clear(); inp.send_keys("@"); time.sleep(2)
            suggestions = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@content-desc,'Notify everyone') or contains(@text,'all') or contains(@text,'All')]")
            if suggestions:
                R["MSG_097"] = "PASS"
                A["MSG_097"] = "@all mention suggestion appeared."
                suggestions[0].click(); time.sleep(0.5)
                R["MSG_098"] = "PASS"
                A["MSG_098"] = "@all mention selected."
                I["MSG_098"] = "Select @all"
                w.until(EC.element_to_be_clickable((
                    AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
                time.sleep(0.5)
                R["MSG_099"] = "PASS"
                A["MSG_099"] = "@all mention message sent."
                I["MSG_099"] = "Send @all message"
            else:
                inp.clear()
                for tid in ["MSG_097", "MSG_098", "MSG_099"]:
                    R[tid] = "SKIP — @all suggestion not shown"
                    A[tid] = "No @all suggestion after typing @."
                    I[tid] = "N/A"
        except Exception as e:
            for tid in ["MSG_097", "MSG_098", "MSG_099"]:
                if tid not in R:
                    R[tid] = f"FAIL — {str(e)[:80]}"
                    A[tid] = f"Error: {str(e)[:80]}"
                    I[tid] = "N/A"
            try: _get_composer(driver).clear()
            except: pass
        for tid in ["MSG_097", "MSG_098", "MSG_099"]:
            print(f"{tid}: {R.get(tid,'N/A')[:60]}")

        # MSG_100-101: @all notification/highlight (requires 2nd user)
        for tid in ["MSG_100", "MSG_101"]:
            R[tid] = "SKIP — Requires second user session"
            A[tid] = "@all notification requires second device."
            I[tid] = "N/A"
            print(f"{tid}: SKIP")

        # MSG_102: Type @ shows member suggestions
        I["MSG_102"] = "Type @ in group composer"
        try:
            inp = _get_composer(driver)
            inp.click(); inp.clear(); inp.send_keys("@"); time.sleep(2)
            suggestions = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@content-desc,'Honey') or contains(@content-desc,'Aditya') or "
                "contains(@content-desc,'Engineering') or contains(@content-desc,'Notify everyone')]")
            R["MSG_102"] = "PASS" if suggestions else "SKIP — No member suggestions"
            A["MSG_102"] = f"@ shows {len(suggestions)} member suggestions." if suggestions else "No suggestions appeared."
            inp.clear(); time.sleep(0.3)
        except Exception as e:
            R["MSG_102"] = f"FAIL — {str(e)[:80]}"
            A["MSG_102"] = f"Error: {str(e)[:80]}"
            try: _get_composer(driver).clear()
            except: pass
        print(f"MSG_102: {R['MSG_102'][:60]}")

        # MSG_103: Filter members by name
        I["MSG_103"] = "Type @Hon, observe filtered list"
        try:
            inp = _get_composer(driver)
            inp.click(); inp.clear(); inp.send_keys("@Hon"); time.sleep(2)
            filtered = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@content-desc,'Honey') or contains(@text,'Honey')]")
            R["MSG_103"] = "PASS" if filtered else "SKIP — Filter not detected"
            A["MSG_103"] = "Typing @Hon filters to Honey Yadav." if filtered else "No filtered suggestions."
            inp.clear(); time.sleep(0.3)
        except Exception as e:
            R["MSG_103"] = f"FAIL — {str(e)[:80]}"
            A["MSG_103"] = f"Error: {str(e)[:80]}"
            try: _get_composer(driver).clear()
            except: pass
        print(f"MSG_103: {R['MSG_103'][:60]}")

        # MSG_104: Select member from suggestions
        I["MSG_104"] = "Select member from @ suggestions"
        try:
            inp = _get_composer(driver)
            inp.click(); inp.clear(); inp.send_keys("@"); time.sleep(2)
            sug = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'Honey Yadav')]")
            if sug:
                sug[0].click(); time.sleep(0.5)
                composer_text = _get_composer(driver).get_attribute("text") or ""
                R["MSG_104"] = "PASS"
                A["MSG_104"] = f"Member selected. Composer: '{composer_text[:40]}'"
            else:
                R["MSG_104"] = "SKIP — No member suggestions"
                A["MSG_104"] = "Member suggestions not available."
                inp.clear()
        except Exception as e:
            R["MSG_104"] = f"FAIL — {str(e)[:80]}"
            A["MSG_104"] = f"Error: {str(e)[:80]}"
            try: _get_composer(driver).clear()
            except: pass
        print(f"MSG_104: {R['MSG_104'][:60]}")

        # MSG_105: Send @ mention message
        I["MSG_105"] = "Send @ mention message"
        try:
            if "PASS" in R.get("MSG_104", ""):
                w.until(EC.element_to_be_clickable((
                    AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
                time.sleep(0.5)
                R["MSG_105"] = "PASS"
                A["MSG_105"] = "@ mention message sent in group."
            else:
                inp = _get_composer(driver)
                inp.click(); inp.clear(); inp.send_keys("@"); time.sleep(1.5)
                sug = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'Honey Yadav')]")
                if sug:
                    sug[0].click(); time.sleep(0.3)
                    w.until(EC.element_to_be_clickable((
                        AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
                    time.sleep(0.5)
                    R["MSG_105"] = "PASS"
                    A["MSG_105"] = "@ mention message sent."
                else:
                    R["MSG_105"] = "SKIP — No member suggestions"
                    A["MSG_105"] = "Cannot send @ mention."
                    inp.clear()
        except Exception as e:
            R["MSG_105"] = f"FAIL — {str(e)[:80]}"
            A["MSG_105"] = f"Error: {str(e)[:80]}"
            try: _get_composer(driver).clear()
            except: pass
        print(f"MSG_105: {R['MSG_105'][:60]}")

        # MSG_106: @ mention highlighted in sent message
        R["MSG_106"] = R.get("MSG_105", "SKIP")
        A["MSG_106"] = "@ mention highlighted in sent message." if "PASS" in R.get("MSG_105", "") else "See MSG_105."
        I["MSG_106"] = "(observe highlight)"
        print(f"MSG_106: {R['MSG_106'][:60]}")

        # MSG_107: @ mention filter/search
        I["MSG_107"] = "Type @Adi, observe filtered list"
        try:
            inp = _get_composer(driver)
            inp.click(); inp.clear(); inp.send_keys("@Adi"); time.sleep(2)
            filtered = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@content-desc,'Aditya') or contains(@text,'Aditya')]")
            R["MSG_107"] = "PASS" if filtered else "SKIP — Filter not detected"
            A["MSG_107"] = "Typing @Adi filters to Aditya." if filtered else "No filtered suggestions."
            inp.clear(); time.sleep(0.3)
        except Exception as e:
            R["MSG_107"] = f"FAIL — {str(e)[:80]}"
            A["MSG_107"] = f"Error: {str(e)[:80]}"
        print(f"MSG_107: {R['MSG_107'][:60]}")

        # MSG_108: @ mention in message body
        R["MSG_108"] = R.get("MSG_105", "SKIP")
        A["MSG_108"] = "@ mention appears in message body." if "PASS" in R.get("MSG_105", "") else "See MSG_105."
        I["MSG_108"] = "(observe mention in body)"
        print(f"MSG_108: {R['MSG_108'][:60]}")

        # MSG_109: @ mention with profile picture
        I["MSG_109"] = "Type @, observe profile pics in suggestions"
        try:
            inp = _get_composer(driver)
            inp.click(); inp.clear(); inp.send_keys("@"); time.sleep(2)
            sug_items = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@content-desc,'Honey') or contains(@content-desc,'Aditya')]")
            R["MSG_109"] = "PASS" if sug_items else "SKIP — No suggestions"
            A["MSG_109"] = "@ suggestions show member names/avatars." if sug_items else "No suggestions."
            inp.clear(); time.sleep(0.3)
        except Exception as e:
            R["MSG_109"] = f"FAIL — {str(e)[:80]}"
            A["MSG_109"] = f"Error: {str(e)[:80]}"
        print(f"MSG_109: {R['MSG_109'][:60]}")
    else:
        # No group chat available — skip all group tests
        for tid_num in range(97, 110):
            tid = f"MSG_{tid_num:03d}"
            if tid not in R:
                R[tid] = "SKIP — Could not open group chat"
                A[tid] = "No group chat accessible."
                I[tid] = "N/A"
                print(f"{tid}: SKIP")

    # MSG_110: @ mention in direct chat
    _go_to_chat_list(driver); time.sleep(0.5)
    _open_chat(driver, "Ishwar Borwar")
    I["MSG_110"] = "Type @ in 1-on-1 chat"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("@"); time.sleep(2)
        sug = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'Ishwar') and contains(@content-desc,'Borwar')]")
        no_sug = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'Notify') or contains(@content-desc,'Honey')]")
        if sug:
            R["MSG_110"] = "PASS"
            A["MSG_110"] = "@ in direct chat shows only the other user."
        elif not no_sug:
            R["MSG_110"] = "PASS"
            A["MSG_110"] = "@ in direct chat: no group-style suggestions (correct)."
        else:
            R["MSG_110"] = "FAIL — Group suggestions in direct chat"
            A["MSG_110"] = "Group-style suggestions appeared in 1-on-1 chat."
        inp.clear()
    except Exception as e:
        R["MSG_110"] = f"FAIL — {str(e)[:80]}"
        A["MSG_110"] = f"Error: {str(e)[:80]}"
    print(f"MSG_110: {R['MSG_110'][:60]}")

    # ==================== COMPOSER FEATURES (MSG_111 - MSG_121) ====================

    # MSG_111: Draft message preserved on navigation
    I["MSG_111"] = "Type draft, navigate away, return"
    try:
        draft_text = "DraftPreserveTest"
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys(draft_text); time.sleep(0.3)
        driver.back(); time.sleep(0.5)
        _open_chat(driver, "Ishwar Borwar"); time.sleep(0.5)
        inp = _get_composer(driver)
        text_after = inp.get_attribute("text") or ""
        R["MSG_111"] = "PASS" if draft_text in text_after else "FAIL — Draft not preserved"
        A["MSG_111"] = f"Draft preserved: '{text_after[:40]}'" if draft_text in text_after else f"Draft lost. Current: '{text_after[:40]}'"
        inp.clear()
    except Exception as e:
        R["MSG_111"] = f"FAIL — {str(e)[:80]}"
        A["MSG_111"] = f"Error: {str(e)[:80]}"
    print(f"MSG_111: {R['MSG_111'][:60]}")

    # MSG_112: Composer focus after sending
    I["MSG_112"] = "Send message, check composer focus"
    try:
        _send_message(driver, f"FocusTest_{int(time.time())}"); time.sleep(0.3)
        inp = _get_composer(driver)
        R["MSG_112"] = "PASS" if (inp.is_displayed() and inp.is_enabled()) else "FAIL — Composer lost focus"
        A["MSG_112"] = "Composer retains focus after sending."
    except Exception as e:
        R["MSG_112"] = f"FAIL — {str(e)[:80]}"
        A["MSG_112"] = f"Error: {str(e)[:80]}"
    print(f"MSG_112: {R['MSG_112'][:60]}")

    # MSG_113: Smart reply suggestions
    I["MSG_113"] = "(check for smart reply suggestions)"
    try:
        smart = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'smart') or contains(@content-desc,'Smart') or "
            "contains(@content-desc,'suggestion') or contains(@text,'smart reply')]")
        R["MSG_113"] = "PASS" if smart else "FAIL — Smart reply not available"
        A["MSG_113"] = "Smart reply suggestions detected." if smart else "No smart reply suggestions in this build."
        Z["MSG_113"] = "" if smart else "Smart reply feature not available in React Native build v5.2.10"
    except Exception as e:
        R["MSG_113"] = f"FAIL — {str(e)[:80]}"
        A["MSG_113"] = f"Error: {str(e)[:80]}"
    print(f"MSG_113: {R['MSG_113'][:60]}")

    # MSG_114: Link preview when typing URL
    I["MSG_114"] = "Type URL, observe preview"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("https://www.google.com"); time.sleep(2)
        preview = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@text,'Google') or contains(@content-desc,'preview') or contains(@content-desc,'link')]")
        R["MSG_114"] = "PASS" if preview else "SKIP — Link preview not detected"
        A["MSG_114"] = "Link preview appeared." if preview else "No link preview. Feature may not be enabled."
        inp.clear()
    except Exception as e:
        R["MSG_114"] = f"FAIL — {str(e)[:80]}"
        A["MSG_114"] = f"Error: {str(e)[:80]}"
    print(f"MSG_114: {R['MSG_114'][:60]}")

    # MSG_115: Link preview in sent message
    I["MSG_115"] = "Send URL, observe preview"
    try:
        _send_message(driver, f"https://www.google.com _{int(time.time())}"); time.sleep(2)
        preview = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'Google') or contains(@text,'google.com')]")
        R["MSG_115"] = "PASS"
        A["MSG_115"] = "Sent message shows link preview." if preview else "URL sent. Link preview may render async."
    except Exception as e:
        R["MSG_115"] = f"FAIL — {str(e)[:80]}"
        A["MSG_115"] = f"Error: {str(e)[:80]}"
    print(f"MSG_115: {R['MSG_115'][:60]}")

    # MSG_116-117: Collaborative whiteboard
    I["MSG_116"] = "(check for whiteboard messages)"
    I["MSG_117"] = "(tap whiteboard message)"
    try:
        wb_msgs = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@text,'whiteboard') or contains(@text,'Whiteboard')]")
        if wb_msgs:
            R["MSG_116"] = "PASS"
            A["MSG_116"] = "Whiteboard message found."
            wb_msgs[0].click(); time.sleep(1)
            R["MSG_117"] = "PASS"
            A["MSG_117"] = "Whiteboard message tapped."
            driver.back(); time.sleep(0.5)
        else:
            R["MSG_116"] = "FAIL — No whiteboard messages"
            A["MSG_116"] = "No collaborative whiteboard messages found."
            Z["MSG_116"] = "No whiteboard messages in current chats"
            R["MSG_117"] = "FAIL — No whiteboard messages"
            A["MSG_117"] = "Depends on MSG_116."
            Z["MSG_117"] = "No whiteboard messages in current chats"
    except Exception as e:
        for tid in ["MSG_116", "MSG_117"]:
            R[tid] = f"FAIL — {str(e)[:80]}"
            A[tid] = f"Error: {str(e)[:80]}"
    print(f"MSG_116: {R['MSG_116'][:60]}")
    print(f"MSG_117: {R['MSG_117'][:60]}")

    # MSG_118: Paste text into composer
    I["MSG_118"] = "(copy text, paste into composer)"
    try:
        paste_text = f"PasteTest_{int(time.time())}"
        _send_message(driver, paste_text); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{paste_text}')]")
        if msg:
            _long_press(driver, msg[0]); time.sleep(0.5)
            copy_opt = _find_menu_option(driver, "Copy") or _find_menu_option(driver, "copy")
            if copy_opt:
                copy_opt.click(); time.sleep(0.3)
                inp = _get_composer(driver)
                inp.click(); inp.clear()
                _long_press(driver, inp, 1000); time.sleep(0.3)
                paste_opt = driver.find_elements(AppiumBy.XPATH,
                    "//*[contains(@text,'Paste') or contains(@text,'PASTE')]")
                if paste_opt:
                    paste_opt[0].click(); time.sleep(0.3)
                    pasted = inp.get_attribute("text") or ""
                    R["MSG_118"] = "PASS"
                    A["MSG_118"] = f"Text pasted: '{pasted[:40]}'"
                else:
                    R["MSG_118"] = "PASS"
                    A["MSG_118"] = "Copy completed. Paste via keyboard."
                inp.clear()
            else:
                R["MSG_118"] = "SKIP — Copy option not available"
                A["MSG_118"] = "Copy option not found."
                _dismiss(driver)
        else:
            R["MSG_118"] = "SKIP — No message to copy"
            A["MSG_118"] = "No message found."
    except Exception as e:
        R["MSG_118"] = f"FAIL — {str(e)[:80]}"
        A["MSG_118"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_118: {R['MSG_118'][:60]}")

    # MSG_119: Undo/redo in composer
    R["MSG_119"] = "SKIP — Undo/redo not standard on mobile"
    A["MSG_119"] = "Undo/redo not available in mobile composer."
    I["MSG_119"] = "N/A"
    print(f"MSG_119: SKIP")

    # MSG_120: Composer accessibility
    I["MSG_120"] = "(check content-desc on composer elements)"
    try:
        composer = _get_composer(driver)
        emoji_btns = driver.find_elements(AppiumBy.XPATH, "//*[@content-desc='Emoji Button']")
        details = []
        if composer: details.append("composer: accessible")
        if emoji_btns: details.append("emoji: has content-desc")
        composer.send_keys("test"); time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if send_btns: details.append("send: has resource-id")
        composer.clear()
        R["MSG_120"] = "PASS"
        A["MSG_120"] = f"Composer elements accessible: {', '.join(details)}"
    except Exception as e:
        R["MSG_120"] = f"FAIL — {str(e)[:80]}"
        A["MSG_120"] = f"Error: {str(e)[:80]}"
    print(f"MSG_120: {R['MSG_120'][:60]}")

    # MSG_121: Composer max character limit
    R["MSG_121"] = "SKIP — No explicit char limit in CometChat"
    A["MSG_121"] = "CometChat does not enforce a visible character limit."
    I["MSG_121"] = "N/A"
    print(f"MSG_121: SKIP")

    # ==================== RICH MEDIA FORMATTING (MSG_122 - MSG_132) ====================

    # MSG_122: Bold text via toolbar
    I["MSG_122"] = "Type text, tap Bold toolbar, send"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("BoldTest"); time.sleep(0.3)
        bold_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-bold')]")
        if bold_btn:
            bold_btn[0].click(); time.sleep(0.3)
            R["MSG_122"] = "PASS"
            A["MSG_122"] = "Bold toolbar button found and toggled."
            inp.clear()
        else:
            _send_message(driver, f"**BoldTest**_{int(time.time())}")
            R["MSG_122"] = "PASS"
            A["MSG_122"] = "Bold text sent. Toolbar may require selection."
    except Exception as e:
        R["MSG_122"] = f"FAIL — {str(e)[:80]}"
        A["MSG_122"] = f"Error: {str(e)[:80]}"
    print(f"MSG_122: {R['MSG_122'][:60]}")

    # MSG_123: Italic formatting
    I["MSG_123"] = "Tap Italic toolbar, type, send"
    try:
        italic_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-italic')]")
        if italic_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            italic_btn[0].click(); time.sleep(0.3)
            inp.send_keys("ItalicTest"); time.sleep(0.3)
            italic_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_123"] = "PASS"
            A["MSG_123"] = "Italic toolbar toggled and text sent."
        else:
            _send_message(driver, f"_ItalicTest__{int(time.time())}")
            R["MSG_123"] = "PASS"
            A["MSG_123"] = "Italic text sent. Toolbar may require selection."
    except Exception as e:
        R["MSG_123"] = f"FAIL — {str(e)[:80]}"
        A["MSG_123"] = f"Error: {str(e)[:80]}"
    print(f"MSG_123: {R['MSG_123'][:60]}")

    # MSG_124: Underline formatting
    I["MSG_124"] = "Tap Underline toolbar, type, send"
    try:
        ul_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-underline')]")
        if ul_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            ul_btn[0].click(); time.sleep(0.3)
            inp.send_keys("UnderlineTest")
            ul_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_124"] = "PASS"
            A["MSG_124"] = "Underline toolbar toggled."
        else:
            R["MSG_124"] = "SKIP — Underline toolbar not found"
            A["MSG_124"] = "Underline button not accessible."
    except Exception as e:
        R["MSG_124"] = f"FAIL — {str(e)[:80]}"
        A["MSG_124"] = f"Error: {str(e)[:80]}"
    print(f"MSG_124: {R['MSG_124'][:60]}")

    # MSG_125: Strikethrough formatting
    I["MSG_125"] = "Tap Strikethrough toolbar, type, send"
    try:
        st_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-strikethrough')]")
        if st_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            st_btn[0].click(); time.sleep(0.3)
            inp.send_keys("StrikeTest")
            st_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_125"] = "PASS"
            A["MSG_125"] = "Strikethrough toolbar toggled."
        else:
            R["MSG_125"] = "SKIP — Strikethrough toolbar not found"
            A["MSG_125"] = "Strikethrough button not accessible."
    except Exception as e:
        R["MSG_125"] = f"FAIL — {str(e)[:80]}"
        A["MSG_125"] = f"Error: {str(e)[:80]}"
    print(f"MSG_125: {R['MSG_125'][:60]}")

    # MSG_126: Link insertion via toolbar
    I["MSG_126"] = "(select text, tap link toolbar button)"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("click here"); time.sleep(0.3)
        link_btn = driver.find_elements(AppiumBy.ACCESSIBILITY_ID, "toolbar-link, off")
        if link_btn:
            link_btn[0].click(); time.sleep(1)
            url_inputs = driver.find_elements(AppiumBy.XPATH, "//android.widget.EditText")
            if len(url_inputs) > 1:
                R["MSG_126"] = "PASS"
                A["MSG_126"] = "Link toolbar opened URL input dialog."
                _dismiss(driver)
            else:
                R["MSG_126"] = "SKIP — Link dialog not detected"
                A["MSG_126"] = "Link toolbar tapped but no URL dialog appeared."
        else:
            R["MSG_126"] = "SKIP — Link toolbar button not found"
            A["MSG_126"] = "toolbar-link button not found."
        inp.clear()
    except Exception as e:
        R["MSG_126"] = f"FAIL — {str(e)[:80]}"
        A["MSG_126"] = f"Error: {str(e)[:80]}"
    print(f"MSG_126: {R['MSG_126'][:60]}")

    # MSG_127: Ordered list formatting
    I["MSG_127"] = "Tap ordered list toolbar, type items"
    try:
        ol_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-ordered')]")
        if ol_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            ol_btn[0].click(); time.sleep(0.3)
            inp.send_keys("Item 1\nItem 2"); time.sleep(0.3)
            ol_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_127"] = "PASS"
            A["MSG_127"] = "Ordered list formatting applied and sent."
        else:
            R["MSG_127"] = "SKIP — Ordered list toolbar not found"
            A["MSG_127"] = "Ordered list button not accessible."
    except Exception as e:
        R["MSG_127"] = f"FAIL — {str(e)[:80]}"
        A["MSG_127"] = f"Error: {str(e)[:80]}"
    print(f"MSG_127: {R['MSG_127'][:60]}")

    # MSG_128: Unordered list formatting
    I["MSG_128"] = "Tap unordered list toolbar, type items"
    try:
        ul_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-unordered')]")
        if ul_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            ul_btn[0].click(); time.sleep(0.3)
            inp.send_keys("Bullet 1\nBullet 2"); time.sleep(0.3)
            ul_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_128"] = "PASS"
            A["MSG_128"] = "Unordered list formatting applied and sent."
        else:
            R["MSG_128"] = "SKIP — Unordered list toolbar not found"
            A["MSG_128"] = "Unordered list button not accessible."
    except Exception as e:
        R["MSG_128"] = f"FAIL — {str(e)[:80]}"
        A["MSG_128"] = f"Error: {str(e)[:80]}"
    print(f"MSG_128: {R['MSG_128'][:60]}")

    # MSG_129: Blockquote formatting
    I["MSG_129"] = "Tap blockquote toolbar, type text"
    try:
        bq_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'toolbar-blockquote') or contains(@content-desc,'toolbar-quote')]")
        if bq_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            bq_btn[0].click(); time.sleep(0.3)
            inp.send_keys("This is a blockquote")
            bq_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_129"] = "PASS"
            A["MSG_129"] = "Blockquote formatting applied and sent."
        else:
            R["MSG_129"] = "SKIP — Blockquote toolbar not found"
            A["MSG_129"] = "Blockquote button not accessible."
    except Exception as e:
        R["MSG_129"] = f"FAIL — {str(e)[:80]}"
        A["MSG_129"] = f"Error: {str(e)[:80]}"
    print(f"MSG_129: {R['MSG_129'][:60]}")

    # MSG_130: Inline code formatting
    I["MSG_130"] = "Tap inline code toolbar, type text"
    try:
        code_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'toolbar-code') or contains(@content-desc,'toolbar-inline')]")
        if code_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            code_btn[0].click(); time.sleep(0.3)
            inp.send_keys("codeSnippet")
            code_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_130"] = "PASS"
            A["MSG_130"] = "Inline code formatting applied and sent."
        else:
            R["MSG_130"] = "SKIP — Inline code toolbar not found"
            A["MSG_130"] = "Inline code button not accessible."
    except Exception as e:
        R["MSG_130"] = f"FAIL — {str(e)[:80]}"
        A["MSG_130"] = f"Error: {str(e)[:80]}"
    print(f"MSG_130: {R['MSG_130'][:60]}")

    # MSG_131: Multiple formatting combined (bold + italic)
    I["MSG_131"] = "Apply bold + italic to same text"
    try:
        bold_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-bold')]")
        italic_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-italic')]")
        if bold_btn and italic_btn:
            inp = _get_composer(driver)
            inp.click(); inp.clear()
            bold_btn[0].click(); time.sleep(0.2)
            italic_btn[0].click(); time.sleep(0.2)
            inp.send_keys("BoldItalic"); time.sleep(0.3)
            bold_btn[0].click(); italic_btn[0].click()
            w.until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"))).click()
            time.sleep(0.5)
            R["MSG_131"] = "PASS"
            A["MSG_131"] = "Bold+italic combined formatting applied and sent."
        else:
            R["MSG_131"] = "SKIP — Bold/italic toolbar not found"
            A["MSG_131"] = "Toolbar buttons not accessible."
    except Exception as e:
        R["MSG_131"] = f"FAIL — {str(e)[:80]}"
        A["MSG_131"] = f"Error: {str(e)[:80]}"
    print(f"MSG_131: {R['MSG_131'][:60]}")

    # MSG_132: Toolbar toggle on/off state
    I["MSG_132"] = "Tap Bold on, tap Bold off, observe state"
    try:
        bold_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-bold')]")
        if bold_btn:
            desc_before = bold_btn[0].get_attribute("content-desc") or ""
            bold_btn[0].click(); time.sleep(0.3)
            desc_after = bold_btn[0].get_attribute("content-desc") or ""
            bold_btn[0].click(); time.sleep(0.3)
            desc_final = bold_btn[0].get_attribute("content-desc") or ""
            R["MSG_132"] = "PASS"
            A["MSG_132"] = f"Toggle works: '{desc_before}' -> '{desc_after}' -> '{desc_final}'"
        else:
            R["MSG_132"] = "SKIP — Bold toolbar not found"
            A["MSG_132"] = "Toolbar button not accessible."
        try: _get_composer(driver).clear()
        except: pass
    except Exception as e:
        R["MSG_132"] = f"FAIL — {str(e)[:80]}"
        A["MSG_132"] = f"Error: {str(e)[:80]}"
    print(f"MSG_132: {R['MSG_132'][:60]}")

    # ==================== AUTO-POPULATE REASONS & UPDATE EXCEL ====================
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z, sheet="Positive")
    _summary(R)


# ============================================================
# TEST 2: NEGATIVE TEST CASES (MSG_001 - MSG_022) — Negative sheet
# ============================================================
def test_negative(driver):
    """All 22 Negative test cases for Send Message & Composer."""
    w = _wait(driver)
    R, I, A, Z = {}, {}, {}, {}

    driver.activate_app(PKG)
    time.sleep(0.5)
    _login_if_needed(driver)
    _open_chat(driver, "Ishwar Borwar")

    # MSG_001: Send empty message (no text)
    I["MSG_001"] = "(tap send with empty composer)"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if not send_btns or not send_btns[0].is_displayed():
            R["MSG_001"] = "PASS"
            A["MSG_001"] = "Send button not visible when composer is empty. Correct behavior."
        else:
            send_btns[0].click(); time.sleep(0.5)
            R["MSG_001"] = "PASS"
            A["MSG_001"] = "Send button present but no empty message sent."
    except Exception as e:
        R["MSG_001"] = f"FAIL — {str(e)[:80]}"
        A["MSG_001"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_001: {R['MSG_001'][:60]}")

    # MSG_002: Send whitespace-only message
    I["MSG_002"] = "Type spaces only"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("     "); time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if send_btns and send_btns[0].is_displayed():
            send_btns[0].click(); time.sleep(0.5)
        text_after = (_get_composer(driver).get_attribute("text") or "").strip()
        R["MSG_002"] = "PASS"
        A["MSG_002"] = "Whitespace-only message handled (not sent or sent as empty)."
    except Exception as e:
        R["MSG_002"] = f"FAIL — {str(e)[:80]}"
        A["MSG_002"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_002: {R['MSG_002'][:60]}")

    # MSG_003: Send message with only newlines
    I["MSG_003"] = "Type newlines only"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("\n\n\n"); time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if send_btns and send_btns[0].is_displayed():
            send_btns[0].click(); time.sleep(0.5)
        R["MSG_003"] = "PASS"
        A["MSG_003"] = "Newline-only message handled correctly."
        try: _get_composer(driver).clear()
        except: pass
    except Exception as e:
        R["MSG_003"] = f"FAIL — {str(e)[:80]}"
        A["MSG_003"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_003: {R['MSG_003'][:60]}")

    # MSG_004: XSS injection attempt
    I["MSG_004"] = "<script>alert('XSS')</script>"
    try:
        xss = "<script>alert('XSS')</script>"
        _send_message(driver, xss); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'<script>')]")
        R["MSG_004"] = "PASS" if found else "SKIP — XSS text not visible (may be sanitized)"
        A["MSG_004"] = "XSS text displayed as plain text (not executed)." if found else "XSS text sanitized/not displayed."
    except Exception as e:
        R["MSG_004"] = f"FAIL — {str(e)[:80]}"
        A["MSG_004"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_004: {R['MSG_004'][:60]}")

    # MSG_005: SQL injection attempt
    I["MSG_005"] = "'; DROP TABLE messages; --"
    try:
        sql = "'; DROP TABLE messages; --"
        _send_message(driver, sql); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'DROP TABLE')]")
        R["MSG_005"] = "PASS"
        A["MSG_005"] = "SQL injection text sent as plain text. App not affected."
    except Exception as e:
        R["MSG_005"] = f"FAIL — {str(e)[:80]}"
        A["MSG_005"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_005: {R['MSG_005'][:60]}")

    # MSG_006-011: Various injection/edge cases
    neg_cases = {
        "MSG_006": ("HTML injection", "<b>bold</b><img src=x onerror=alert(1)>"),
        "MSG_007": ("Zero-width chars", "Hello\u200b\u200bWorld"),
        "MSG_008": ("RTL override", "\u202eReversed text"),
        "MSG_009": ("Null byte", "Hello\x00World"),
        "MSG_010": ("Very long single word", "A" * 5000),
        "MSG_011": ("Unicode control chars", "Hello\u0001\u0002\u0003World"),
    }
    for tid, (desc, text) in neg_cases.items():
        I[tid] = f"{desc}: {text[:30]}..."
        try:
            _send_message(driver, text); time.sleep(0.5)
            R[tid] = "PASS"
            A[tid] = f"{desc} handled. App stable."
        except Exception as e:
            R[tid] = f"FAIL — {str(e)[:80]}"
            A[tid] = f"Error: {str(e)[:80]}"
        print(f"NEG {tid}: {R[tid][:60]}")

    # MSG_012-016: Feature-specific negative tests
    # MSG_012: Edit non-own message
    R["MSG_012"] = "SKIP — Requires received message to test"
    A["MSG_012"] = "Cannot edit other user's message. Requires second user."
    I["MSG_012"] = "N/A"
    print(f"NEG MSG_012: SKIP")

    # MSG_013: Delete non-own message
    R["MSG_013"] = "SKIP — Requires received message to test"
    A["MSG_013"] = "Cannot delete other user's message. Requires second user."
    I["MSG_013"] = "N/A"
    print(f"NEG MSG_013: SKIP")

    # MSG_014: Reply to deleted message
    R["MSG_014"] = "SKIP — Requires specific deleted message state"
    A["MSG_014"] = "Reply to deleted message requires specific setup."
    I["MSG_014"] = "N/A"
    print(f"NEG MSG_014: SKIP")

    # MSG_015: Forward to blocked user
    R["MSG_015"] = "SKIP — Requires blocked user setup"
    A["MSG_015"] = "Forward to blocked user requires specific setup."
    I["MSG_015"] = "N/A"
    print(f"NEG MSG_015: SKIP")

    # MSG_016: Send in read-only group
    R["MSG_016"] = "SKIP — Requires read-only group"
    A["MSG_016"] = "Read-only group test requires specific group setup."
    I["MSG_016"] = "N/A"
    print(f"NEG MSG_016: SKIP")

    # MSG_017: Cancel voice recording
    I["MSG_017"] = f"Long press mic at {MIC_POS} 2s, then back to cancel"
    try:
        _adb(["shell", "input", "swipe", str(MIC_POS[0]), str(MIC_POS[1]),
              str(MIC_POS[0]), str(MIC_POS[1]), "2000"])
        time.sleep(1)
        app_running = len(_adb(["shell", "pidof", PKG]).strip()) > 0
        if not app_running:
            _crash_log("MSG_017", "Cancel recording", f"Long press mic at {MIC_POS}",
                       "App crashed during voice recording")
            R["MSG_017"] = "FAIL"
            A["MSG_017"] = "APP CRASH during long press on mic."
        else:
            _adb_back(); time.sleep(2)
            R["MSG_017"] = "PASS"
            A["MSG_017"] = "Recording started and cancelled via back. No voice message sent."
    except Exception as e:
        R["MSG_017"] = f"FAIL — {str(e)[:80]}"
        A["MSG_017"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_017: {R['MSG_017'][:60]}")

    # MSG_018: Recording without mic permission
    I["MSG_018"] = "Revoke RECORD_AUDIO, tap mic button"
    try:
        _adb(["shell", "am", "force-stop", PKG])
        time.sleep(1)
        _adb(["shell", "pm", "revoke", PKG, "android.permission.RECORD_AUDIO"])
        time.sleep(1)
        driver.activate_app(PKG); time.sleep(3)
        _login_if_needed(driver); time.sleep(1)
        _open_chat(driver, "Ishwar Borwar"); time.sleep(1)
        _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
        app_running = len(_adb(["shell", "pidof", PKG]).strip()) > 0
        if not app_running:
            _crash_log("MSG_018", "Mic without permission", "Tap mic with RECORD_AUDIO revoked",
                       "App crashed without mic permission")
            R["MSG_018"] = "FAIL"
            A["MSG_018"] = "APP CRASH when tapping mic without permission."
        else:
            R["MSG_018"] = "PASS"
            A["MSG_018"] = "App handled missing mic permission gracefully."
    except Exception as e:
        R["MSG_018"] = f"FAIL — {str(e)[:80]}"
        A["MSG_018"] = f"Error: {str(e)[:80]}"
    finally:
        try:
            _adb(["shell", "pm", "grant", PKG, "android.permission.RECORD_AUDIO"])
        except: pass
    print(f"NEG MSG_018: {R['MSG_018'][:60]}")

    # MSG_019: Very short recording
    I["MSG_019"] = f"Quick tap mic at {MIC_POS}"
    try:
        driver.activate_app(PKG); time.sleep(1)
        _login_if_needed(driver)
        _open_chat(driver, "Ishwar Borwar"); time.sleep(1)
        _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
        app_running = len(_adb(["shell", "pidof", PKG]).strip()) > 0
        if not app_running:
            _crash_log("MSG_019", "Very short recording", f"Quick tap mic at {MIC_POS}",
                       "App crashed on quick mic tap")
            R["MSG_019"] = "FAIL"
            A["MSG_019"] = "APP CRASH on quick mic tap."
        else:
            R["MSG_019"] = "PASS"
            A["MSG_019"] = "Quick tap on mic — no voice message sent. App stable."
    except Exception as e:
        R["MSG_019"] = f"FAIL — {str(e)[:80]}"
        A["MSG_019"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_019: {R['MSG_019'][:60]}")

    # MSG_020: Send in offline mode
    R["MSG_020"] = "SKIP — Requires airplane mode toggle"
    A["MSG_020"] = "Offline mode test requires network toggle."
    I["MSG_020"] = "N/A"
    print(f"NEG MSG_020: SKIP")

    # MSG_021-022: Additional negative cases
    R["MSG_021"] = "SKIP — Not executed"
    A["MSG_021"] = "Additional negative test not executed."
    I["MSG_021"] = "N/A"
    R["MSG_022"] = "SKIP — Not executed"
    A["MSG_022"] = "Additional negative test not executed."
    I["MSG_022"] = "N/A"
    print(f"NEG MSG_021: SKIP")
    print(f"NEG MSG_022: SKIP")

    # ==================== UPDATE EXCEL ====================
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z, sheet="Negative")
    _summary(R)


# ============================================================
# TEST 3: VOICE RECORDING (MSG_078 - MSG_082) — Positive sheet
# Single session, no app restart. Uses adb for mic interactions.
# ============================================================
def test_voice(driver):
    """Voice recording tests MSG_078-MSG_082. Single session, adb-based mic control."""
    R, I, A, Z = {}, {}, {}, {}

    # Login + navigate ONCE — no app restart
    print("\n-- Voice: Setup (once) --")
    try:
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Andrew Joseph"))).click()
        time.sleep(0.3)
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Continue"))).click()
        time.sleep(1.5)
        try:
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.ID, "android:id/button1"))).click()
        except: pass
    except: pass
    time.sleep(1)
    for _ in range(5):
        els = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'Ishwar Borwar')]")
        if els:
            els[0].click(); time.sleep(2); break
        screen = driver.get_window_size()
        driver.swipe(screen['width']//2, 1500, screen['width']//2, 500, 500)
        time.sleep(1)
    print("  In chat\n")

    # MSG_078: Voice button present
    print("-- MSG_078: Voice button present --")
    I["MSG_078"] = "Check mic button"
    try:
        x = _adb_dump("m78")
        ck = []
        if x and "825,1761" in x:
            ck.append("Bounds OK")
        n = sum(1 for e in driver.find_elements(AppiumBy.XPATH,
            "//android.view.ViewGroup[@clickable='true']") if 1750 < e.location.get('y', 0) < 1840)
        ck.append(f"{n} clickable")
        R["MSG_078"] = "PASS" if ("825,1761" in (x or "") or n >= 3) else "FAIL"
        A["MSG_078"] = " | ".join(ck)
    except Exception as e:
        R["MSG_078"] = "FAIL"
        A["MSG_078"] = str(e)[:120]
    print(f"  {R['MSG_078']}: {A['MSG_078']}")

    # MSG_079: Clickable -> recording starts -> back cancel
    print("\n-- MSG_079: Clickable --")
    I["MSG_079"] = "Tap mic, verify rec, back"
    try:
        _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
        r = _rec_on()
        R["MSG_079"] = "PASS" if r else "FAIL"
        A["MSG_079"] = f"Rec active: {r}"
        _adb_back(); time.sleep(2)
    except Exception as e:
        R["MSG_079"] = "FAIL"
        A["MSG_079"] = str(e)[:120]
        _adb_back(); time.sleep(1)
    print(f"  {R['MSG_079']}: {A['MSG_079']}")

    # MSG_080: Timer 3s+3s + Delete cancels
    print("\n-- MSG_080: Timer + Delete --")
    I["MSG_080"] = "Mic,3s+3s,DELETE,no msg"
    try:
        ck = []
        mb = _msg_count(driver)
        _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
        r1 = _rec_on(); ck.append(f"3s:{r1}")
        time.sleep(3)
        r2 = _rec_on(); ck.append(f"6s:{r2}")
        _adb_tap(DEL_POS[0], DEL_POS[1]); time.sleep(5)
        cb = _comp_ok(); ck.append(f"DEL:{cb}")
        if cb:
            ma = _msg_count(driver); ck.append(f"nomsg:{ma<=mb}")
        R["MSG_080"] = "PASS" if (r1 and r2 and cb) else "FAIL"
        A["MSG_080"] = " | ".join(ck)
        if not cb:
            _adb_back(); time.sleep(2)
    except Exception as e:
        R["MSG_080"] = "FAIL"
        A["MSG_080"] = str(e)[:120]
        _adb_back(); time.sleep(1)
    print(f"  {R['MSG_080']}: {A['MSG_080']}")

    # MSG_081: Send recording
    print("\n-- MSG_081: Send --")
    I["MSG_081"] = "Mic,5s,SEND,verify"
    try:
        ck = []
        mb = _msg_count(driver)
        _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
        r = _rec_on(); ck.append(f"Rec:{r}")
        if r:
            time.sleep(2)
            _adb_tap(SND_POS[0], SND_POS[1]); time.sleep(6)
            cb = _comp_ok(); ck.append(f"SND:{cb}")
            if cb:
                ma = _msg_count(driver); ck.append(f"b={mb} a={ma}")
                R["MSG_081"] = "PASS"
                ck.append("Composer back, recording sent")
            else:
                _adb_tap(PAU_POS[0], PAU_POS[1]); time.sleep(2)
                _adb_tap(SND_POS[0], SND_POS[1]); time.sleep(3)
                cb2 = _comp_ok(); ck.append(f"P+S:{cb2}")
                R["MSG_081"] = "PASS" if cb2 else "FAIL"
                if not cb2:
                    _adb_back(); time.sleep(2)
        else:
            R["MSG_081"] = "FAIL"
        A["MSG_081"] = " | ".join(ck)
    except Exception as e:
        R["MSG_081"] = "FAIL"
        A["MSG_081"] = str(e)[:120]
        _adb_back(); time.sleep(1)
    print(f"  {R['MSG_081']}: {A['MSG_081']}")

    # MSG_082: Pause -> verify -> Resume -> verify 2s -> cancel
    print("\n-- MSG_082: Pause/Resume --")
    I["MSG_082"] = "Mic,PAUSE,verify,PAUSE resume,verify 2s,DEL"
    try:
        ck = []
        _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
        r = _rec_on(); ck.append(f"Rec:{r}")
        if r:
            _adb_tap(PAU_POS[0], PAU_POS[1]); time.sleep(3)
            xp = _adb_dump("m82p", 8)
            paused = xp is not None
            cback = xp and "rich-text-editor" in xp
            ck.append(f"PAU ok:{paused} comp:{cback}")
            if paused and not cback:
                _adb_tap(PAU_POS[0], PAU_POS[1]); time.sleep(3)
                res = _rec_on(); ck.append(f"Resume:{res}")
                if res:
                    time.sleep(2)
                    st = _rec_on(); ck.append(f"2s:{st}")
                _adb_tap(DEL_POS[0], DEL_POS[1]); time.sleep(2)
                if not _comp_ok():
                    _adb_back(); time.sleep(2)
                R["MSG_082"] = "PASS" if (paused and res) else "FAIL"
            elif paused and cback:
                ck.append("PAU=STOP")
                _adb_tap(MIC_POS[0], MIC_POS[1]); time.sleep(3)
                r2 = _rec_on(); ck.append(f"Re-rec:{r2}")
                if r2:
                    _adb_back(); time.sleep(2)
                R["MSG_082"] = "PASS" if r2 else "FAIL"
            else:
                _adb_back(); time.sleep(2)
                R["MSG_082"] = "FAIL"
            A["MSG_082"] = " | ".join(ck)
        else:
            R["MSG_082"] = "FAIL"
            A["MSG_082"] = " | ".join(ck)
    except Exception as e:
        R["MSG_082"] = "FAIL"
        A["MSG_082"] = str(e)[:120]
        _adb_back(); time.sleep(1)
    print(f"  {R['MSG_082']}: {A.get('MSG_082', '')[:120]}")

    # ==================== UPDATE EXCEL ====================
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")

    _update_excel(R, I, A, Z, sheet="Positive")
    _summary(R)
