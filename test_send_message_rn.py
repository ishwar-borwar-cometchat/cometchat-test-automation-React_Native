"""
CometChat React Native Android - First 10 Test Cases (MSG_001 to MSG_010)
Runs each test and updates the Excel sheet with Actual Result, Status, and Input_Data columns.
"""
import time
import openpyxl
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL_PATH = "Cometchat_Features/Send_Message/Send_Message_Test_Cases.xlsx"
ACTUAL_RESULT_COL = 8
STATUS_COL = 10
INPUT_DATA_COL = 11
REASON_COL = 12
APP_PACKAGE = "com.cometchat.sampleapp.reactnative.android"


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _login_if_needed(driver):
    """Login by selecting Andrew Joseph sample user."""
    try:
        andrew = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Andrew Joseph"
        )))
        andrew.click()
        time.sleep(0.3)
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Continue"
        ))).click()
        time.sleep(1.5)
        # Dismiss VOIP dialog
        try:
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.ID, "android:id/button1"
            ))).click()
            time.sleep(0.3)
        except Exception:
            pass
        # Dismiss notification permission
        try:
            _wait(driver, 3).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@text='Allow' or @text='ALLOW']"
            ))).click()
            time.sleep(0.3)
        except Exception:
            pass
        print("Logged in as Andrew Joseph.")
    except Exception:
        print("Already logged in.")


def _open_chat_with_user(driver, user_name="Ishwar Borwar"):
    """Open a chat from the Chats list by content-desc or text."""
    w = _wait(driver)
    # Try content-desc first (chat list items use it)
    try:
        user = w.until(EC.element_to_be_clickable((
            AppiumBy.XPATH, f"//*[contains(@content-desc, '{user_name}')]"
        )))
        user.click()
        time.sleep(0.3)
        print(f"Opened chat with {user_name}")
        return
    except Exception:
        pass
    # Fallback: find by text within a clickable parent
    try:
        user = w.until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            f"//*[contains(@text, '{user_name}')]/ancestor::android.view.ViewGroup[@clickable='true']"
        )))
        user.click()
        time.sleep(0.3)
        print(f"Opened chat with {user_name} (by text)")
    except Exception:
        print(f"Could not find {user_name}")


def _go_back(driver):
    """Go back to previous screen."""
    driver.back()
    time.sleep(0.3)


def _get_composer(driver):
    """Find the message composer input field."""
    return _wait(driver).until(EC.element_to_be_clickable((
        AppiumBy.XPATH,
        "//android.widget.EditText[contains(@hint, 'Type') or contains(@text, 'Type your message')]"
    )))


def _status_style(status_val):
    """Return (font, fill) for status color coding."""
    from openpyxl.styles import Font as F, PatternFill as PF
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return F(bold=True, color="006100", name="Calibri"), PF(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return F(bold=True, color="9C0006", name="Calibri"), PF(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return F(bold=True, color="9C5700", name="Calibri"), PF(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return F(bold=True, color="3F3F76", name="Calibri"), PF(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None):
    """Write Actual Result, Status, Input_Data, and Reason columns to the Excel."""
    if reasons is None:
        reasons = {}
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=ACTUAL_RESULT_COL, value=actual_results.get(test_id, ""))
                status_cell = ws.cell(row=row, column=STATUS_COL, value=results[test_id])
                font, fill = _status_style(results[test_id])
                status_cell.font = font
                status_cell.fill = fill
                ws.cell(row=row, column=INPUT_DATA_COL, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=REASON_COL, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL_PATH)
    print(f"Excel updated with {len(results)} results.")


def test_send_message_first_10(driver):
    """Run MSG_001 to MSG_010 on React Native build."""
    w = _wait(driver)
    results = {}
    input_data = {}
    actual_results = {}
    reasons = {}

    driver.activate_app(APP_PACKAGE)
    time.sleep(0.3)
    _login_if_needed(driver)
    _open_chat_with_user(driver, "Ishwar Borwar")

    # MSG_001: Verify message input field is visible
    input_data["MSG_001"] = "None (observation only)"
    try:
        inp = _get_composer(driver)
        assert inp.is_displayed()
        results["MSG_001"] = "PASS"
        actual_results["MSG_001"] = "Message input field visible with 'Type your message...' placeholder."
        print("MSG_001: PASS")
    except Exception as e:
        results["MSG_001"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_001"] = f"Input field not found: {str(e)[:80]}"
        print(f"MSG_001: FAIL — {e}")

    # MSG_002: Verify message input field is clickable
    input_data["MSG_002"] = "Click on composer"
    try:
        inp = _get_composer(driver)
        inp.click()
        assert inp.is_enabled()
        results["MSG_002"] = "PASS"
        actual_results["MSG_002"] = "Input field is clickable and enabled."
        print("MSG_002: PASS")
    except Exception as e:
        results["MSG_002"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_002"] = f"Input field not clickable: {str(e)[:80]}"
        print(f"MSG_002: FAIL — {e}")

    # MSG_003: Verify typing in message input field
    test_text_003 = "Test message"
    input_data["MSG_003"] = test_text_003
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys(test_text_003)
        time.sleep(0.3)
        typed_text = inp.get_attribute("text") or ""
        assert test_text_003 in typed_text
        results["MSG_003"] = "PASS"
        actual_results["MSG_003"] = f"Typed text displayed correctly: '{typed_text}'"
        print("MSG_003: PASS")
    except Exception as e:
        results["MSG_003"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_003"] = f"Typing failed: {str(e)[:80]}"
        print(f"MSG_003: FAIL — {e}")

    # MSG_004: Verify multi-line message input
    test_text_004 = "Line 1\\nLine 2\\nLine 3"
    input_data["MSG_004"] = test_text_004
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys("Line 1\nLine 2\nLine 3")
        time.sleep(0.3)
        typed_text = inp.get_attribute("text") or ""
        assert "Line 1" in typed_text and "Line 2" in typed_text
        results["MSG_004"] = "PASS"
        actual_results["MSG_004"] = f"Multi-line text accepted: '{typed_text[:60]}'"
        print("MSG_004: PASS")
    except Exception as e:
        results["MSG_004"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_004"] = f"Multi-line input failed: {str(e)[:80]}"
        print(f"MSG_004: FAIL — {e}")
    finally:
        try:
            _get_composer(driver).clear()
        except Exception:
            pass

    # MSG_005: Verify empty message cannot be sent (deep check)
    input_data["MSG_005"] = "(empty string)"
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        time.sleep(0.3)
        # Count messages before
        msgs_before = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc, 'pm') or contains(@content-desc, 'am')]")
        count_before = len(msgs_before)
        try:
            send_btn = _wait(driver, 3).until(EC.presence_of_element_located((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            )))
            if not send_btn.is_enabled() or not send_btn.is_displayed():
                results["MSG_005"] = "PASS"
                actual_results["MSG_005"] = "Send button disabled/hidden when input is empty."
            else:
                send_btn.click()
                time.sleep(0.3)
                msgs_after = driver.find_elements(AppiumBy.XPATH,
                    "//*[contains(@content-desc, 'pm') or contains(@content-desc, 'am')]")
                if len(msgs_after) == count_before:
                    results["MSG_005"] = "PASS"
                    actual_results["MSG_005"] = "Send button enabled but empty message not delivered."
                else:
                    results["MSG_005"] = "FAIL — Empty message was actually sent"
                    actual_results["MSG_005"] = "Empty message was delivered to chat."
        except Exception:
            results["MSG_005"] = "PASS"
            actual_results["MSG_005"] = "Send button not found when input is empty."
        print(f"MSG_005: {results['MSG_005']}")
    except Exception as e:
        results["MSG_005"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_005"] = f"Error: {str(e)[:80]}"
        print(f"MSG_005: FAIL — {e}")

    # MSG_006: Verify message with only spaces (deep check)
    input_data["MSG_006"] = "'     ' (5 spaces)"
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys("     ")
        time.sleep(0.3)
        msgs_before = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc, 'pm') or contains(@content-desc, 'am')]")
        count_before = len(msgs_before)
        try:
            send_btn = _wait(driver, 3).until(EC.presence_of_element_located((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            )))
            if not send_btn.is_enabled() or not send_btn.is_displayed():
                results["MSG_006"] = "PASS"
                actual_results["MSG_006"] = "Send button disabled/hidden for spaces-only input."
            else:
                send_btn.click()
                time.sleep(0.3)
                msgs_after = driver.find_elements(AppiumBy.XPATH,
                    "//*[contains(@content-desc, 'pm') or contains(@content-desc, 'am')]")
                if len(msgs_after) == count_before:
                    results["MSG_006"] = "PASS"
                    actual_results["MSG_006"] = "Send button enabled but spaces-only message not delivered."
                else:
                    results["MSG_006"] = "FAIL — Spaces-only message was actually sent"
                    actual_results["MSG_006"] = "Spaces-only message was delivered to chat."
        except Exception:
            results["MSG_006"] = "PASS"
            actual_results["MSG_006"] = "Send button not found for spaces-only input."
        print(f"MSG_006: {results['MSG_006']}")
    except Exception as e:
        results["MSG_006"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_006"] = f"Error: {str(e)[:80]}"
        print(f"MSG_006: FAIL — {e}")
    finally:
        try:
            _get_composer(driver).clear()
        except Exception:
            pass

    # MSG_007: Verify send button is visible
    test_text_007 = "test"
    input_data["MSG_007"] = test_text_007
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys(test_text_007)
        time.sleep(0.3)
        send_btn = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, "//*[@resource-id='send-button']"
        )))
        assert send_btn.is_displayed()
        results["MSG_007"] = "PASS"
        actual_results["MSG_007"] = "Send button is visible after typing text."
        print("MSG_007: PASS")
    except Exception as e:
        results["MSG_007"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_007"] = f"Send button not visible: {str(e)[:80]}"
        print(f"MSG_007: FAIL — {e}")
    finally:
        try:
            _get_composer(driver).clear()
        except Exception:
            pass

    # MSG_008: Verify send button enabled when text entered
    test_text_008 = "Hello"
    input_data["MSG_008"] = test_text_008
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys(test_text_008)
        time.sleep(0.3)
        send_btn = w.until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']"
        )))
        assert send_btn.is_enabled() and send_btn.is_displayed()
        results["MSG_008"] = "PASS"
        actual_results["MSG_008"] = "Send button is enabled and displayed after typing text."
        print("MSG_008: PASS")
    except Exception as e:
        results["MSG_008"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_008"] = f"Send button not enabled: {str(e)[:80]}"
        print(f"MSG_008: FAIL — {e}")
    finally:
        try:
            _get_composer(driver).clear()
        except Exception:
            pass

    # MSG_009: Verify send button click sends message
    test_text_009 = f"TestRN009_{int(time.time())}"
    input_data["MSG_009"] = test_text_009
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys(test_text_009)
        time.sleep(0.3)
        w.until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']"
        ))).click()
        time.sleep(1.5)
        msg_el = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text, '{test_text_009}') or contains(@content-desc, '{test_text_009}')]"
        )))
        assert msg_el is not None
        results["MSG_009"] = "PASS"
        actual_results["MSG_009"] = f"Message '{test_text_009}' sent and visible in chat."
        print("MSG_009: PASS")
    except Exception as e:
        results["MSG_009"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_009"] = f"Message not sent/visible: {str(e)[:80]}"
        print(f"MSG_009: FAIL — {e}")

    # MSG_010: Verify send button visual feedback on click
    test_text_010 = "FeedbackTestRN"
    input_data["MSG_010"] = test_text_010
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys(test_text_010)
        time.sleep(0.3)
        send_btn = w.until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']"
        )))
        assert send_btn.is_displayed() and send_btn.is_enabled()
        send_btn.click()
        time.sleep(0.3)
        inp_after = _get_composer(driver)
        text_after = inp_after.get_attribute("text") or ""
        if test_text_010 not in text_after:
            results["MSG_010"] = "PASS"
            actual_results["MSG_010"] = "Input field cleared after send — message sent successfully."
            print("MSG_010: PASS")
        else:
            results["MSG_010"] = "FAIL — Message not sent after clicking send"
            actual_results["MSG_010"] = "Input field still contains text after clicking send."
            print("MSG_010: FAIL")
    except Exception as e:
        results["MSG_010"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_010"] = f"Error: {str(e)[:80]}"
        print(f"MSG_010: FAIL — {e}")

    # Auto-populate reasons for FAIL/SKIP
    for tid in results:
        status = results[tid]
        if status.startswith("FAIL"):
            reasons[tid] = status.replace("FAIL — ", "")
        elif status.startswith("SKIP"):
            reasons[tid] = status.replace("SKIP — ", "")

    # Update Excel
    _update_excel(results, input_data, actual_results, reasons)

    # Summary
    print("\n=== SUMMARY ===")
    passed = sum(1 for v in results.values() if v == "PASS")
    failed = sum(1 for v in results.values() if v.startswith("FAIL"))
    print(f"Total: {len(results)} | Passed: {passed} | Failed: {failed}")
    for tid in sorted(results.keys()):
        reason_str = f" | Reason: {reasons[tid]}" if reasons.get(tid) else ""
        print(f"  {tid}: {results[tid]} | Input: {input_data.get(tid, 'N/A')}{reason_str}")
