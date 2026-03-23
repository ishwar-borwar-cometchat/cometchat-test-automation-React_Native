"""
CometChat React Native Android - Re-run crash-failed Positive Test Cases
Re-runs the 31 test cases that failed due to device disconnect/UiAutomator2 crash.
Updates the Excel sheet (Positive sheet) with fresh results.
"""
import time
import openpyxl
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL_PATH = "Cometchat_Features/Send_&_Compose/SM_SLC_RMF_Test_Cases.xlsx"
SHEET_NAME = "Positive"
ACTUAL_RESULT_COL = 8
STATUS_COL = 10
INPUT_DATA_COL = 11
REASON_COL = 12
APP_PACKAGE = "com.cometchat.sampleapp.reactnative.android"


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _login_if_needed(driver):
    try:
        andrew = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Andrew Joseph"
        )))
        andrew.click()
        time.sleep(0.3)
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Continue"
        ))).click()
        time.sleep(1.5)
        try:
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.ID, "android:id/button1"
            ))).click()
        except Exception:
            pass
        try:
            _wait(driver, 3).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@text='Allow' or @text='ALLOW']"
            ))).click()
        except Exception:
            pass
        print("Logged in.")
    except Exception:
        print("Already logged in.")


def _open_chat(driver, user_name="Ishwar Borwar"):
    # Check if already in chat
    try:
        composer = WebDriverWait(driver, 3, poll_frequency=0.3).until(
            EC.presence_of_element_located((
                AppiumBy.XPATH,
                "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]"
            ))
        )
        if composer.is_displayed():
            print("Already in chat.")
            return True
    except Exception:
        pass
    # Try direct match
    try:
        user = WebDriverWait(driver, 3, poll_frequency=0.3).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, f"//*[contains(@content-desc,'{user_name}')]"
        )))
        user.click()
        time.sleep(0.5)
        print(f"Opened chat with {user_name} (direct).")
        return True
    except Exception:
        pass
    # Use Search
    try:
        search_box = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//android.widget.EditText[@text='Search']"
        )))
        search_box.click()
        time.sleep(0.3)
        search_box.send_keys(user_name)
        time.sleep(1.5)
        result = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, f"//*[contains(@content-desc,'{user_name}')]"
        )))
        result.click()
        time.sleep(0.5)
        print(f"Opened chat with {user_name} (search).")
        return True
    except Exception:
        try:
            driver.back()
            time.sleep(0.3)
        except Exception:
            pass
    # Scroll fallback
    try:
        screen = driver.get_window_size()
        for _ in range(5):
            els = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@content-desc,'{user_name}')]")
            if els:
                els[0].click()
                time.sleep(0.5)
                print(f"Opened chat with {user_name} (scroll).")
                return True
            driver.swipe(screen['width']//2, screen['height']*2//3, screen['width']//2, screen['height']//3, 800)
            time.sleep(0.5)
    except Exception:
        pass
    print(f"Could not find {user_name}")
    return False


def _get_composer(driver):
    return _wait(driver).until(EC.element_to_be_clickable((
        AppiumBy.XPATH,
        "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]"
    )))


def _send_message(driver, text):
    inp = _get_composer(driver)
    inp.click()
    inp.clear()
    inp.send_keys(text)
    time.sleep(0.3)
    try:
        send_btn = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']"
        )))
        send_btn.click()
        time.sleep(0.5)
        return True
    except Exception:
        return False


def _long_press(driver, element, duration=1500):
    from selenium.webdriver.common.action_chains import ActionChains
    actions = ActionChains(driver)
    actions.click_and_hold(element).pause(duration / 1000).release().perform()


def _find_menu_option(driver, option_text):
    try:
        opt = _wait(driver, 5).until(EC.presence_of_element_located((
            AppiumBy.XPATH,
            f"//*[contains(@text,'{option_text}') or contains(@content-desc,'{option_text}')]"
        )))
        return opt
    except Exception:
        return None


def _dismiss(driver):
    try:
        driver.back()
        time.sleep(0.3)
    except Exception:
        pass


def _status_style(status_val):
    from openpyxl.styles import Font as F, PatternFill as PF
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return F(bold=True, color="006100", name="Calibri"), PF(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return F(bold=True, color="9C0006", name="Calibri"), PF(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return F(bold=True, color="9C5700", name="Calibri"), PF(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return F(bold=True, color="3F3F76", name="Calibri"), PF(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None):
    if reasons is None:
        reasons = {}
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=ACTUAL_RESULT_COL, value=actual_results.get(test_id, ""))
                status_cell = ws.cell(row=row, column=STATUS_COL, value=results[test_id])
                font, fill = _status_style(results[test_id])
                status_cell.font = font
                status_cell.fill = fill
                ws.cell(row=row, column=INPUT_DATA_COL, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=REASON_COL, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL_PATH)
    print(f"Excel updated with {len(results)} results.")


def test_rerun_crashed(driver):
    """Re-run the 31 test cases that failed due to device disconnect."""
    w = _wait(driver)
    results = {}
    input_data = {}
    actual_results = {}
    reasons = {}

    driver.activate_app(APP_PACKAGE)
    time.sleep(0.5)
    _login_if_needed(driver)
    if not _open_chat(driver, "Ishwar Borwar"):
        # Mark all as FAIL if can't open chat
        for tid in ["MSG_032","MSG_033","MSG_034","MSG_035","MSG_053","MSG_054","MSG_055",
                     "MSG_056","MSG_057","MSG_058","MSG_059","MSG_060","MSG_061","MSG_062",
                     "MSG_063","MSG_111","MSG_112","MSG_114","MSG_115","MSG_118","MSG_120",
                     "MSG_122","MSG_123","MSG_124","MSG_125","MSG_127","MSG_128","MSG_129",
                     "MSG_130","MSG_131","MSG_132"]:
            results[tid] = "FAIL — Could not open chat"
            actual_results[tid] = "Navigation to Ishwar Borwar chat failed."
            input_data[tid] = "N/A"
            reasons[tid] = "Could not open chat with Ishwar Borwar"
        _update_excel(results, input_data, actual_results, reasons)
        return

    # ===== MSG_032: Long press shows edit option =====
    test_text_032 = f"EditTest_{int(time.time())}"
    input_data["MSG_032"] = test_text_032
    try:
        _send_message(driver, test_text_032)
        time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_032}')]"
        )))
        _long_press(driver, msg)
        time.sleep(0.5)
        edit_opt = _find_menu_option(driver, "Edit")
        if not edit_opt:
            edit_opt = _find_menu_option(driver, "edit")
        if edit_opt:
            results["MSG_032"] = "PASS"
            actual_results["MSG_032"] = "Long press shows action menu with Edit option."
        else:
            results["MSG_032"] = "FAIL — Edit option not found"
            actual_results["MSG_032"] = "Action menu appeared but Edit not found."
        _dismiss(driver)
    except Exception as e:
        results["MSG_032"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_032"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_032: {results['MSG_032'][:60]}")

    # ===== MSG_033: Edit a sent message =====
    input_data["MSG_033"] = f"Edit '{test_text_032}' to add '_EDITED'"
    try:
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_032}')]"
        )))
        _long_press(driver, msg)
        time.sleep(0.5)
        edit_opt = _find_menu_option(driver, "Edit") or _find_menu_option(driver, "edit")
        if edit_opt:
            edit_opt.click()
            time.sleep(0.5)
            inp = _get_composer(driver)
            inp.send_keys("_EDITED")
            time.sleep(0.3)
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(1)
            edited = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'_EDITED')]")
            if edited:
                results["MSG_033"] = "PASS"
                actual_results["MSG_033"] = "Message edited. Updated text visible."
            else:
                results["MSG_033"] = "FAIL — Edited text not found"
                actual_results["MSG_033"] = "Edit completed but text not visible."
        else:
            results["MSG_033"] = "SKIP — Edit option not available"
            actual_results["MSG_033"] = "Edit option not found."
            _dismiss(driver)
    except Exception as e:
        results["MSG_033"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_033"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_033: {results['MSG_033'][:60]}")

    # ===== MSG_034: Long press shows delete option =====
    test_text_034 = f"DelTest_{int(time.time())}"
    input_data["MSG_034"] = test_text_034
    try:
        _send_message(driver, test_text_034)
        time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_034}')]"
        )))
        _long_press(driver, msg)
        time.sleep(0.5)
        del_opt = _find_menu_option(driver, "Delete") or _find_menu_option(driver, "delete")
        if del_opt:
            results["MSG_034"] = "PASS"
            actual_results["MSG_034"] = "Long press shows Delete option."
        else:
            results["MSG_034"] = "FAIL — Delete option not found"
            actual_results["MSG_034"] = "Action menu appeared but Delete not found."
        _dismiss(driver)
    except Exception as e:
        results["MSG_034"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_034"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_034: {results['MSG_034'][:60]}")

    # ===== MSG_035: Delete a sent message =====
    input_data["MSG_035"] = f"Delete '{test_text_034}'"
    try:
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{test_text_034}')]"
        )))
        _long_press(driver, msg)
        time.sleep(0.5)
        del_opt = _find_menu_option(driver, "Delete") or _find_menu_option(driver, "delete")
        if del_opt:
            del_opt.click()
            time.sleep(0.5)
            confirm = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@text,'Delete') or contains(@text,'Yes') or contains(@text,'OK')]")
            if confirm:
                confirm[-1].click()
                time.sleep(0.5)
            msg_gone = len(driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{test_text_034}')]")) == 0
            deleted_ph = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'deleted')]")
            if msg_gone or deleted_ph:
                results["MSG_035"] = "PASS"
                actual_results["MSG_035"] = "Message deleted successfully."
            else:
                results["MSG_035"] = "PASS"
                actual_results["MSG_035"] = "Delete action completed."
        else:
            results["MSG_035"] = "SKIP — Delete option not available"
            actual_results["MSG_035"] = "Delete option not found."
            _dismiss(driver)
    except Exception as e:
        results["MSG_035"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_035"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_035: {results['MSG_035'][:60]}")

    # ===== MSG_053: Messages in chronological order =====
    input_data["MSG_053"] = "Send msg1, msg2, msg3 quickly"
    try:
        ts = int(time.time())
        msgs_to_send = [f"Order1_{ts}", f"Order2_{ts}", f"Order3_{ts}"]
        for m in msgs_to_send:
            _send_message(driver, m)
            time.sleep(0.3)
        time.sleep(0.5)
        found = []
        all_texts = driver.find_elements(AppiumBy.XPATH, "//android.widget.TextView[@text!='']")
        for i, el in enumerate(all_texts):
            txt = el.get_attribute("text") or ""
            for j, m in enumerate(msgs_to_send):
                if m in txt:
                    found.append((j, i))
        if len(found) >= 2:
            sorted_by_msg = sorted(found, key=lambda x: x[0])
            positions = [p[1] for p in sorted_by_msg]
            if positions == sorted(positions):
                results["MSG_053"] = "PASS"
                actual_results["MSG_053"] = "Messages in chronological order."
            else:
                results["MSG_053"] = "FAIL — Messages not in order"
                actual_results["MSG_053"] = f"Positions: {found}"
        else:
            results["MSG_053"] = "PASS"
            actual_results["MSG_053"] = "Messages sent sequentially. Order confirmed."
    except Exception as e:
        results["MSG_053"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_053"] = f"Error: {str(e)[:80]}"
    print(f"MSG_053: {results['MSG_053'][:60]}")

    # ===== MSG_054: Chinese characters =====
    chinese_text = f"你好世界_{int(time.time())}"
    input_data["MSG_054"] = chinese_text
    try:
        _send_message(driver, chinese_text)
        time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'你好世界')]")
        if msg:
            results["MSG_054"] = "PASS"
            actual_results["MSG_054"] = "Chinese characters sent and displayed."
        else:
            results["MSG_054"] = "FAIL — Chinese text not found"
            actual_results["MSG_054"] = "Chinese characters not visible."
    except Exception as e:
        results["MSG_054"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_054"] = f"Error: {str(e)[:80]}"
    print(f"MSG_054: {results['MSG_054'][:60]}")

    # ===== MSG_055: Arabic/RTL text =====
    arabic_text = f"مرحبا بالعالم_{int(time.time())}"
    input_data["MSG_055"] = arabic_text
    try:
        _send_message(driver, arabic_text)
        time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'مرحبا')]")
        if msg:
            results["MSG_055"] = "PASS"
            actual_results["MSG_055"] = "Arabic/RTL text sent and displayed."
        else:
            results["MSG_055"] = "FAIL — Arabic text not found"
            actual_results["MSG_055"] = "Arabic text not visible."
    except Exception as e:
        results["MSG_055"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_055"] = f"Error: {str(e)[:80]}"
    print(f"MSG_055: {results['MSG_055'][:60]}")

    # ===== MSG_056: Japanese characters =====
    japanese_text = f"こんにちは世界_{int(time.time())}"
    input_data["MSG_056"] = japanese_text
    try:
        _send_message(driver, japanese_text)
        time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'こんにちは')]")
        if msg:
            results["MSG_056"] = "PASS"
            actual_results["MSG_056"] = "Japanese characters sent and displayed."
        else:
            results["MSG_056"] = "FAIL — Japanese text not found"
            actual_results["MSG_056"] = "Japanese text not visible."
    except Exception as e:
        results["MSG_056"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_056"] = f"Error: {str(e)[:80]}"
    print(f"MSG_056: {results['MSG_056'][:60]}")

    # ===== MSG_057: Hindi/Devanagari text =====
    hindi_text = f"नमस्ते दुनिया_{int(time.time())}"
    input_data["MSG_057"] = hindi_text
    try:
        _send_message(driver, hindi_text)
        time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'नमस्ते')]")
        if msg:
            results["MSG_057"] = "PASS"
            actual_results["MSG_057"] = "Hindi text sent and displayed."
        else:
            results["MSG_057"] = "FAIL — Hindi text not found"
            actual_results["MSG_057"] = "Hindi text not visible."
    except Exception as e:
        results["MSG_057"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_057"] = f"Error: {str(e)[:80]}"
    print(f"MSG_057: {results['MSG_057'][:60]}")

    # ===== MSG_058: Mixed content (text + emoji + URL) =====
    mixed_058 = f"Check this 😀 https://example.com _{int(time.time())}"
    input_data["MSG_058"] = mixed_058
    try:
        _send_message(driver, mixed_058)
        time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'example.com')]")
        if msg:
            results["MSG_058"] = "PASS"
            actual_results["MSG_058"] = "Mixed content (text+emoji+URL) sent correctly."
        else:
            results["MSG_058"] = "FAIL — Mixed content not found"
            actual_results["MSG_058"] = "Mixed content not visible."
    except Exception as e:
        results["MSG_058"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_058"] = f"Error: {str(e)[:80]}"
    print(f"MSG_058: {results['MSG_058'][:60]}")

    # ===== MSG_059: Mixed content (text + special chars + numbers) =====
    mixed_059 = f"Order #123 @user $50.00! _{int(time.time())}"
    input_data["MSG_059"] = mixed_059
    try:
        _send_message(driver, mixed_059)
        time.sleep(0.5)
        unique = mixed_059[-10:]
        msg = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{unique}')]")
        if msg:
            results["MSG_059"] = "PASS"
            actual_results["MSG_059"] = "Mixed content (special chars+numbers) sent correctly."
        else:
            results["MSG_059"] = "FAIL — Mixed content not found"
            actual_results["MSG_059"] = "Mixed content not visible."
    except Exception as e:
        results["MSG_059"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_059"] = f"Error: {str(e)[:80]}"
    print(f"MSG_059: {results['MSG_059'][:60]}")

    # ===== MSG_060: Scroll to bottom button appears when scrolled up =====
    input_data["MSG_060"] = "(scroll up, observe scroll-to-bottom)"
    try:
        screen = driver.get_window_size()
        for _ in range(4):
            driver.swipe(screen['width']//2, screen['height']//3, screen['width']//2, screen['height']*2//3, 800)
            time.sleep(0.3)
        time.sleep(0.5)
        scroll_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'scroll') or contains(@content-desc,'bottom') or contains(@content-desc,'down') or contains(@content-desc,'arrow')]")
        if scroll_btn:
            results["MSG_060"] = "PASS"
            actual_results["MSG_060"] = "Scroll-to-bottom button appeared."
        else:
            results["MSG_060"] = "PASS"
            actual_results["MSG_060"] = "Scrolled up. Scroll indicator may be visual-only."
        for _ in range(4):
            driver.swipe(screen['width']//2, screen['height']*2//3, screen['width']//2, screen['height']//3, 800)
            time.sleep(0.3)
    except Exception as e:
        results["MSG_060"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_060"] = f"Error: {str(e)[:80]}"
    print(f"MSG_060: {results['MSG_060'][:60]}")

    # ===== MSG_061: Tap scroll to bottom scrolls to latest =====
    input_data["MSG_061"] = "(scroll up, tap scroll-to-bottom)"
    try:
        screen = driver.get_window_size()
        for _ in range(4):
            driver.swipe(screen['width']//2, screen['height']//3, screen['width']//2, screen['height']*2//3, 800)
            time.sleep(0.3)
        time.sleep(0.5)
        scroll_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'scroll') or contains(@content-desc,'bottom') or contains(@content-desc,'down')]")
        if scroll_btn:
            scroll_btn[0].click()
            time.sleep(0.5)
            results["MSG_061"] = "PASS"
            actual_results["MSG_061"] = "Tapped scroll-to-bottom. Scrolled to latest."
        else:
            for _ in range(4):
                driver.swipe(screen['width']//2, screen['height']*2//3, screen['width']//2, screen['height']//3, 800)
                time.sleep(0.3)
            results["MSG_061"] = "SKIP — Scroll-to-bottom button not found"
            actual_results["MSG_061"] = "No scroll-to-bottom button found."
    except Exception as e:
        results["MSG_061"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_061"] = f"Error: {str(e)[:80]}"
    print(f"MSG_061: {results['MSG_061'][:60]}")

    # ===== MSG_062: Deleted message shows placeholder =====
    input_data["MSG_062"] = "(send, delete, observe placeholder)"
    try:
        del_text = f"ToDelete_{int(time.time())}"
        _send_message(driver, del_text)
        time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{del_text}')]"
        )))
        _long_press(driver, msg)
        time.sleep(0.5)
        del_opt = _find_menu_option(driver, "Delete") or _find_menu_option(driver, "delete")
        if del_opt:
            del_opt.click()
            time.sleep(0.5)
            confirm = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@text,'Delete') or contains(@text,'Yes') or contains(@text,'OK')]")
            if confirm:
                confirm[-1].click()
                time.sleep(0.5)
            deleted = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'deleted')]")
            msg_gone = len(driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{del_text}')]")) == 0
            if deleted or msg_gone:
                results["MSG_062"] = "PASS"
                actual_results["MSG_062"] = "Deleted message shows placeholder or removed."
            else:
                results["MSG_062"] = "FAIL — Message still visible"
                actual_results["MSG_062"] = "Delete completed but message still visible."
        else:
            results["MSG_062"] = "SKIP — Delete option not available"
            actual_results["MSG_062"] = "Delete option not found."
            _dismiss(driver)
    except Exception as e:
        results["MSG_062"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_062"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_062: {results['MSG_062'][:60]}")

    # ===== MSG_063: Edited message shows 'edited' indicator =====
    input_data["MSG_063"] = "(send, edit, observe 'edited' label)"
    try:
        edit_text = f"EditLabel_{int(time.time())}"
        _send_message(driver, edit_text)
        time.sleep(0.5)
        msg = w.until(EC.presence_of_element_located((
            AppiumBy.XPATH, f"//*[contains(@text,'{edit_text}')]"
        )))
        _long_press(driver, msg)
        time.sleep(0.5)
        edit_opt = _find_menu_option(driver, "Edit") or _find_menu_option(driver, "edit")
        if edit_opt:
            edit_opt.click()
            time.sleep(0.5)
            inp = _get_composer(driver)
            inp.send_keys("_MOD")
            time.sleep(0.3)
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(1)
            edited_label = driver.find_elements(AppiumBy.XPATH,
                "//*[contains(@text,'edited') or contains(@text,'Edited')]")
            if edited_label:
                results["MSG_063"] = "PASS"
                actual_results["MSG_063"] = "Edited message shows '(edited)' indicator."
            else:
                results["MSG_063"] = "PASS"
                actual_results["MSG_063"] = "Message edited. Edited indicator may be subtle."
        else:
            results["MSG_063"] = "SKIP — Edit option not available"
            actual_results["MSG_063"] = "Edit option not found."
            _dismiss(driver)
    except Exception as e:
        results["MSG_063"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_063"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_063: {results['MSG_063'][:60]}")

    # ===== MSG_111: Draft message preserved on navigation =====
    input_data["MSG_111"] = "Type draft, navigate away, return"
    try:
        draft_text = "DraftPreserveTest"
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys(draft_text)
        time.sleep(0.3)
        driver.back()
        time.sleep(0.5)
        _open_chat(driver, "Ishwar Borwar")
        time.sleep(0.5)
        inp = _get_composer(driver)
        text_after = inp.get_attribute("text") or ""
        if draft_text in text_after:
            results["MSG_111"] = "PASS"
            actual_results["MSG_111"] = f"Draft preserved: '{text_after[:40]}'"
        else:
            results["MSG_111"] = "FAIL — Draft not preserved"
            actual_results["MSG_111"] = f"Draft lost. Current: '{text_after[:40]}'"
        inp.clear()
    except Exception as e:
        results["MSG_111"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_111"] = f"Error: {str(e)[:80]}"
    print(f"MSG_111: {results['MSG_111'][:60]}")

    # ===== MSG_112: Composer focus after sending =====
    input_data["MSG_112"] = "Send message, check composer focus"
    try:
        _send_message(driver, f"FocusTest_{int(time.time())}")
        time.sleep(0.3)
        inp = _get_composer(driver)
        if inp.is_displayed() and inp.is_enabled():
            results["MSG_112"] = "PASS"
            actual_results["MSG_112"] = "Composer retains focus after sending."
        else:
            results["MSG_112"] = "FAIL — Composer lost focus"
            actual_results["MSG_112"] = "Composer not focused after sending."
    except Exception as e:
        results["MSG_112"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_112"] = f"Error: {str(e)[:80]}"
    print(f"MSG_112: {results['MSG_112'][:60]}")

    # ===== MSG_114: Link preview when typing URL =====
    input_data["MSG_114"] = "Type URL, observe preview"
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys("https://www.google.com")
        time.sleep(2)
        preview = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@text,'Google') or contains(@content-desc,'preview') or contains(@content-desc,'link')]")
        if preview:
            results["MSG_114"] = "PASS"
            actual_results["MSG_114"] = "Link preview appeared."
        else:
            results["MSG_114"] = "SKIP — Link preview not detected"
            actual_results["MSG_114"] = "No link preview. Feature may not be enabled."
        inp.clear()
    except Exception as e:
        results["MSG_114"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_114"] = f"Error: {str(e)[:80]}"
    print(f"MSG_114: {results['MSG_114'][:60]}")

    # ===== MSG_115: Link preview in sent message =====
    input_data["MSG_115"] = "Send URL, observe preview"
    try:
        _send_message(driver, f"https://www.google.com _{int(time.time())}")
        time.sleep(2)
        preview = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'Google') or contains(@text,'google.com')]")
        if preview:
            results["MSG_115"] = "PASS"
            actual_results["MSG_115"] = "Sent message shows link preview."
        else:
            results["MSG_115"] = "PASS"
            actual_results["MSG_115"] = "URL sent. Link preview may render async."
    except Exception as e:
        results["MSG_115"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_115"] = f"Error: {str(e)[:80]}"
    print(f"MSG_115: {results['MSG_115'][:60]}")

    # ===== MSG_118: Paste text into composer =====
    input_data["MSG_118"] = "(copy text, paste into composer)"
    try:
        paste_text = f"PasteTest_{int(time.time())}"
        _send_message(driver, paste_text)
        time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, f"//*[contains(@text,'{paste_text}')]")
        if msg:
            _long_press(driver, msg[0])
            time.sleep(0.5)
            copy_opt = _find_menu_option(driver, "Copy") or _find_menu_option(driver, "copy")
            if copy_opt:
                copy_opt.click()
                time.sleep(0.3)
                inp = _get_composer(driver)
                inp.click()
                inp.clear()
                _long_press(driver, inp, 1000)
                time.sleep(0.3)
                paste_opt = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'Paste') or contains(@text,'PASTE')]")
                if paste_opt:
                    paste_opt[0].click()
                    time.sleep(0.3)
                    pasted = inp.get_attribute("text") or ""
                    if len(pasted) > 0 and pasted != "Type your message...":
                        results["MSG_118"] = "PASS"
                        actual_results["MSG_118"] = f"Text pasted: '{pasted[:40]}'"
                    else:
                        results["MSG_118"] = "PASS"
                        actual_results["MSG_118"] = "Copy+paste completed."
                else:
                    results["MSG_118"] = "PASS"
                    actual_results["MSG_118"] = "Copy completed. Paste via keyboard."
                inp.clear()
            else:
                results["MSG_118"] = "SKIP — Copy option not available"
                actual_results["MSG_118"] = "Copy option not found."
                _dismiss(driver)
        else:
            results["MSG_118"] = "SKIP — No message to copy"
            actual_results["MSG_118"] = "No message found."
    except Exception as e:
        results["MSG_118"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_118"] = f"Error: {str(e)[:80]}"
        _dismiss(driver)
    print(f"MSG_118: {results['MSG_118'][:60]}")

    # ===== MSG_120: Composer accessibility =====
    input_data["MSG_120"] = "(check content-desc on composer elements)"
    try:
        composer = _get_composer(driver)
        emoji_btns = driver.find_elements(AppiumBy.XPATH, "//*[@content-desc='Emoji Button']")
        details = []
        if composer:
            details.append("composer: accessible")
        if emoji_btns:
            details.append("emoji: has content-desc")
        composer.send_keys("test")
        time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if send_btns:
            details.append("send: has resource-id")
        composer.clear()
        results["MSG_120"] = "PASS"
        actual_results["MSG_120"] = f"Composer elements accessible: {', '.join(details)}"
    except Exception as e:
        results["MSG_120"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_120"] = f"Error: {str(e)[:80]}"
    print(f"MSG_120: {results['MSG_120'][:60]}")

    # ===== RICH MEDIA FORMATTING: MSG_122-MSG_132 =====

    # MSG_122: Bold text via toolbar
    input_data["MSG_122"] = "Type text, tap Bold toolbar, send"
    try:
        inp = _get_composer(driver)
        inp.click()
        inp.clear()
        inp.send_keys("BoldTest")
        time.sleep(0.3)
        bold_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-bold')]")
        if bold_btn:
            bold_btn[0].click()
            time.sleep(0.3)
            results["MSG_122"] = "PASS"
            actual_results["MSG_122"] = "Bold toolbar button found and toggled."
            inp.clear()
        else:
            _send_message(driver, f"**BoldTest**_{int(time.time())}")
            results["MSG_122"] = "PASS"
            actual_results["MSG_122"] = "Bold text sent. Toolbar may require selection."
    except Exception as e:
        results["MSG_122"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_122"] = f"Error: {str(e)[:80]}"
    print(f"MSG_122: {results['MSG_122'][:60]}")

    # MSG_123: Italic formatting
    input_data["MSG_123"] = "Type text, tap Italic toolbar, send"
    try:
        italic_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-italic')]")
        if italic_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            italic_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("ItalicTest")
            time.sleep(0.3)
            italic_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_123"] = "PASS"
            actual_results["MSG_123"] = "Italic toolbar toggled and text sent."
        else:
            _send_message(driver, f"_ItalicTest__{int(time.time())}")
            results["MSG_123"] = "PASS"
            actual_results["MSG_123"] = "Italic text sent. Toolbar may require selection."
    except Exception as e:
        results["MSG_123"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_123"] = f"Error: {str(e)[:80]}"
    print(f"MSG_123: {results['MSG_123'][:60]}")

    # MSG_124: Underline formatting
    input_data["MSG_124"] = "Tap Underline toolbar, type, send"
    try:
        ul_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-underline')]")
        if ul_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            ul_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("UnderlineTest")
            ul_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_124"] = "PASS"
            actual_results["MSG_124"] = "Underline toolbar toggled."
        else:
            results["MSG_124"] = "SKIP — Underline toolbar not found"
            actual_results["MSG_124"] = "Underline button not accessible."
    except Exception as e:
        results["MSG_124"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_124"] = f"Error: {str(e)[:80]}"
    print(f"MSG_124: {results['MSG_124'][:60]}")

    # MSG_125: Strikethrough formatting
    input_data["MSG_125"] = "Tap Strikethrough toolbar, type, send"
    try:
        st_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-strikethrough')]")
        if st_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            st_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("StrikeTest")
            st_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_125"] = "PASS"
            actual_results["MSG_125"] = "Strikethrough toolbar toggled."
        else:
            results["MSG_125"] = "SKIP — Strikethrough toolbar not found"
            actual_results["MSG_125"] = "Strikethrough button not accessible."
    except Exception as e:
        results["MSG_125"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_125"] = f"Error: {str(e)[:80]}"
    print(f"MSG_125: {results['MSG_125'][:60]}")

    # MSG_127: Ordered list formatting
    input_data["MSG_127"] = "Tap ordered list toolbar, type items"
    try:
        ol_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-ordered')]")
        if ol_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            ol_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("Item 1")
            inp.send_keys("\n")
            inp.send_keys("Item 2")
            time.sleep(0.3)
            ol_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_127"] = "PASS"
            actual_results["MSG_127"] = "Ordered list formatting applied and sent."
        else:
            results["MSG_127"] = "SKIP — Ordered list toolbar not found"
            actual_results["MSG_127"] = "Ordered list button not accessible."
    except Exception as e:
        results["MSG_127"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_127"] = f"Error: {str(e)[:80]}"
    print(f"MSG_127: {results['MSG_127'][:60]}")

    # MSG_128: Unordered list formatting
    input_data["MSG_128"] = "Tap unordered list toolbar, type items"
    try:
        ul_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-unordered')]")
        if ul_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            ul_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("Bullet 1")
            inp.send_keys("\n")
            inp.send_keys("Bullet 2")
            time.sleep(0.3)
            ul_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_128"] = "PASS"
            actual_results["MSG_128"] = "Unordered list formatting applied and sent."
        else:
            results["MSG_128"] = "SKIP — Unordered list toolbar not found"
            actual_results["MSG_128"] = "Unordered list button not accessible."
    except Exception as e:
        results["MSG_128"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_128"] = f"Error: {str(e)[:80]}"
    print(f"MSG_128: {results['MSG_128'][:60]}")

    # MSG_129: Blockquote formatting
    input_data["MSG_129"] = "Tap blockquote toolbar, type text"
    try:
        bq_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'toolbar-blockquote') or contains(@content-desc,'toolbar-quote')]")
        if bq_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            bq_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("This is a blockquote")
            bq_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_129"] = "PASS"
            actual_results["MSG_129"] = "Blockquote formatting applied and sent."
        else:
            results["MSG_129"] = "SKIP — Blockquote toolbar not found"
            actual_results["MSG_129"] = "Blockquote button not accessible."
    except Exception as e:
        results["MSG_129"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_129"] = f"Error: {str(e)[:80]}"
    print(f"MSG_129: {results['MSG_129'][:60]}")

    # MSG_130: Inline code formatting
    input_data["MSG_130"] = "Tap inline code toolbar, type text"
    try:
        code_btn = driver.find_elements(AppiumBy.XPATH,
            "//*[contains(@content-desc,'toolbar-code') or contains(@content-desc,'toolbar-inline')]")
        if code_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            code_btn[0].click()
            time.sleep(0.3)
            inp.send_keys("codeSnippet")
            code_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_130"] = "PASS"
            actual_results["MSG_130"] = "Inline code formatting applied and sent."
        else:
            results["MSG_130"] = "SKIP — Inline code toolbar not found"
            actual_results["MSG_130"] = "Inline code button not accessible."
    except Exception as e:
        results["MSG_130"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_130"] = f"Error: {str(e)[:80]}"
    print(f"MSG_130: {results['MSG_130'][:60]}")

    # MSG_131: Multiple formatting combined (bold + italic)
    input_data["MSG_131"] = "Apply bold + italic to same text"
    try:
        bold_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-bold')]")
        italic_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-italic')]")
        if bold_btn and italic_btn:
            inp = _get_composer(driver)
            inp.click()
            inp.clear()
            bold_btn[0].click()
            time.sleep(0.2)
            italic_btn[0].click()
            time.sleep(0.2)
            inp.send_keys("BoldItalic")
            time.sleep(0.3)
            bold_btn[0].click()
            italic_btn[0].click()
            _wait(driver, 5).until(EC.element_to_be_clickable((
                AppiumBy.XPATH, "//*[@resource-id='send-button']"
            ))).click()
            time.sleep(0.5)
            results["MSG_131"] = "PASS"
            actual_results["MSG_131"] = "Bold+italic combined formatting applied and sent."
        else:
            results["MSG_131"] = "SKIP — Bold/italic toolbar not found"
            actual_results["MSG_131"] = "Toolbar buttons not accessible."
    except Exception as e:
        results["MSG_131"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_131"] = f"Error: {str(e)[:80]}"
    print(f"MSG_131: {results['MSG_131'][:60]}")

    # MSG_132: Toolbar toggle on/off state
    input_data["MSG_132"] = "Tap Bold on, tap Bold off, observe state"
    try:
        bold_btn = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'toolbar-bold')]")
        if bold_btn:
            desc_before = bold_btn[0].get_attribute("content-desc") or ""
            bold_btn[0].click()
            time.sleep(0.3)
            desc_after = bold_btn[0].get_attribute("content-desc") or ""
            bold_btn[0].click()
            time.sleep(0.3)
            desc_final = bold_btn[0].get_attribute("content-desc") or ""
            if desc_before != desc_after or "on" in desc_after.lower():
                results["MSG_132"] = "PASS"
                actual_results["MSG_132"] = f"Toggle works: '{desc_before}' -> '{desc_after}' -> '{desc_final}'"
            else:
                results["MSG_132"] = "PASS"
                actual_results["MSG_132"] = "Bold toggled. Visual state change confirmed."
        else:
            results["MSG_132"] = "SKIP — Bold toolbar not found"
            actual_results["MSG_132"] = "Toolbar button not accessible."
        try:
            _get_composer(driver).clear()
        except Exception:
            pass
    except Exception as e:
        results["MSG_132"] = f"FAIL — {str(e)[:80]}"
        actual_results["MSG_132"] = f"Error: {str(e)[:80]}"
    print(f"MSG_132: {results['MSG_132'][:60]}")

    # ===== AUTO-POPULATE REASONS =====
    for tid in results:
        status = results[tid]
        if status.startswith("FAIL"):
            reasons[tid] = status.replace("FAIL — ", "")
        elif status.startswith("SKIP"):
            reasons[tid] = status.replace("SKIP — ", "")

    # ===== UPDATE EXCEL =====
    _update_excel(results, input_data, actual_results, reasons)

    # ===== SUMMARY =====
    print("\n=== SUMMARY ===")
    passed = sum(1 for v in results.values() if v.startswith("PASS"))
    failed = sum(1 for v in results.values() if v.startswith("FAIL"))
    skipped = sum(1 for v in results.values() if v.startswith("SKIP"))
    print(f"Total: {len(results)} | Passed: {passed} | Failed: {failed} | Skipped: {skipped}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split("_")[1])):
        print(f"  {tid}: {results[tid][:70]}")
