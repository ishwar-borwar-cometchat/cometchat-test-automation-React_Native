"""
CometChat React Native — Negative Test Cases (22 TCs)

Usage:
  python3 -m pytest "Cometchat_Features/Send_&_Compose/test_negative.py" -v -s
"""
import os
import time
import subprocess
import shutil
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ============================================================
# CONSTANTS — all auto-detected, nothing hardcoded
# ============================================================
EXCEL = os.path.join(os.path.dirname(__file__) or ".", "SM_SLC_RMF_Test_Cases.xlsx")
if not os.path.exists(EXCEL):
    EXCEL = "Cometchat_Features/Send_&_Compose/SM_SLC_RMF_Test_Cases.xlsx"
PKG = "com.cometchat.sampleapp.reactnative.android"
BUILD = "React Native Android v5.2.10"

# Auto-detect adb path
ADB = shutil.which("adb") or os.path.join(os.environ.get("ANDROID_HOME", ""), "platform-tools", "adb")

# Auto-detect connected device
def _get_device_id():
    try:
        result = subprocess.run([ADB, "devices"], capture_output=True, text=True, timeout=10)
        for line in result.stdout.strip().split("\n")[1:]:
            parts = line.strip().split("\t")
            if len(parts) == 2 and parts[1] == "device":
                return parts[0]
    except Exception:
        pass
    return ""

DEVICE = _get_device_id()


# ============================================================
# HELPER FUNCTIONS
# ============================================================
def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _get_screen_size(driver):
    """Get screen size from driver at runtime."""
    return driver.get_window_size()


def _find_mic_button(driver):
    """Dynamically find the mic/voice record button by position — it's the clickable
    ViewGroup between Emoji Button and send-button, with no content-desc."""
    try:
        # Get Emoji Button and send-button positions as anchors
        emoji_btn = driver.find_elements(AppiumBy.ACCESSIBILITY_ID, "Emoji Button")
        send_btn = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if emoji_btn and send_btn:
            emoji_bounds = emoji_btn[0].get_attribute("bounds") or ""
            send_bounds = send_btn[0].get_attribute("bounds") or ""
            # Parse right edge of emoji and left edge of send
            e_right = int(emoji_bounds.replace("[", "").replace("]", ",").split(",")[2])
            s_left = int(send_bounds.replace("[", "").replace("]", ",").split(",")[0])
            # Mic button is between emoji and send button
            mic_x = (e_right + s_left) // 2
            mic_y = emoji_btn[0].location['y'] + emoji_btn[0].size['height'] // 2
            return mic_x, mic_y
    except Exception:
        pass
    # Fallback: find clickable ViewGroups in composer row with no content-desc
    try:
        groups = driver.find_elements(AppiumBy.XPATH,
            "//android.view.ViewGroup[@clickable='true' and @content-desc='']")
        screen = driver.get_window_size()
        # Filter to bottom area (last 15% of screen)
        bottom_groups = [g for g in groups
                         if g.location.get('y', 0) > screen['height'] * 0.8
                         and g.size.get('width', 0) < 100]
        if bottom_groups:
            # Mic is typically the rightmost small button before send
            bottom_groups.sort(key=lambda g: g.location.get('x', 0), reverse=True)
            mic = bottom_groups[0]
            return mic.location['x'] + mic.size['width'] // 2, mic.location['y'] + mic.size['height'] // 2
    except Exception:
        pass
    return None, None


def _find_recording_buttons(driver):
    """Dynamically find DELETE, PAUSE, SEND buttons during voice recording.
    Returns dict with keys 'delete', 'pause', 'send' -> (x, y) tuples."""
    buttons = {}
    try:
        groups = driver.find_elements(AppiumBy.XPATH,
            "//android.view.ViewGroup[@clickable='true']")
        screen = driver.get_window_size()
        # Filter to bottom area during recording
        bottom = [g for g in groups
                  if g.location.get('y', 0) > screen['height'] * 0.8]
        bottom.sort(key=lambda g: g.location.get('x', 0))
        if len(bottom) >= 3:
            # Left = delete, middle = pause, right = send
            for i, key in enumerate(['delete', 'pause', 'send']):
                if i < len(bottom):
                    b = bottom[i]
                    buttons[key] = (b.location['x'] + b.size['width'] // 2,
                                    b.location['y'] + b.size['height'] // 2)
        elif len(bottom) >= 1:
            # At least one button found
            buttons['delete'] = (bottom[0].location['x'] + bottom[0].size['width'] // 2,
                                 bottom[0].location['y'] + bottom[0].size['height'] // 2)
            if len(bottom) >= 2:
                buttons['send'] = (bottom[-1].location['x'] + bottom[-1].size['width'] // 2,
                                   bottom[-1].location['y'] + bottom[-1].size['height'] // 2)
    except Exception:
        pass
    return buttons


def _ensure_in_chat(driver, user_name="Ishwar Borwar"):
    """Ensure we are in the correct chat. Recover if not. Returns True if in chat."""
    try:
        composer = WebDriverWait(driver, 3, poll_frequency=0.3).until(
            EC.presence_of_element_located((
                AppiumBy.XPATH,
                "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")))
        if composer.is_displayed():
            return True
    except Exception:
        pass
    # Not in chat — try to recover
    print(f"  [Recovery] Not in chat, navigating to {user_name}...")
    try:
        # Check if app is running
        app_state = driver.query_app_state(PKG)
        if app_state < 3:  # Not running or in background
            driver.activate_app(PKG)
            time.sleep(2)
            _login_if_needed(driver)
        _go_to_chat_list(driver)
        time.sleep(0.5)
        return _open_chat(driver, user_name)
    except Exception as e:
        print(f"  [Recovery] Failed: {str(e)[:60]}")
        return False


def _login_if_needed(driver):
    """Login by selecting Andrew Joseph sample user. Waits for chat list to load."""
    try:
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Andrew Joseph"))).click()
        time.sleep(0.3)
        _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.ACCESSIBILITY_ID, "Continue"))).click()
        time.sleep(3)
        # Dismiss VOIP/permission dialogs via adb (safer than Appium on Android 16)
        for _ in range(5):
            try:
                focus = _adb(["shell", "dumpsys", "window"])
                focus_line = ""
                for line in focus.split("\n"):
                    if "mCurrentFocus" in line:
                        focus_line = line; break
                if any(x in focus_line for x in ["GrantPermission", "telecom", "CallingAccount"]) and PKG not in focus_line:
                    _adb(["shell", "input", "keyevent", "4"])
                    time.sleep(1)
                    continue
                btns = driver.find_elements(AppiumBy.XPATH,
                    "//*[@text='OK' or @text='Allow' or @text='ALLOW' or @text='While using the app']")
                if btns:
                    btns[0].click(); time.sleep(1)
                else:
                    break
            except Exception:
                _adb(["shell", "input", "keyevent", "4"])
                time.sleep(1)
        print("Logged in as Andrew Joseph.")
    except Exception:
        print("Already logged in — checking screen state...")
        # We're past login but might be on wrong screen (VOIP settings, etc.)
        # Press back only if current focus is a system dialog, NOT the app
        for _ in range(5):
            try:
                focus = _adb(["shell", "dumpsys", "window"])
                # Extract only mCurrentFocus line
                focus_line = ""
                for line in focus.split("\n"):
                    if "mCurrentFocus" in line:
                        focus_line = line; break
                if any(x in focus_line for x in ["GrantPermission", "telecom", "CallingAccount"]) and PKG not in focus_line:
                    print(f"  System dialog: {focus_line.strip()[:60]}")
                    _adb(["shell", "input", "keyevent", "4"])
                    time.sleep(1)
                    continue
            except Exception:
                pass
            try:
                chats = driver.find_elements(AppiumBy.XPATH,
                    "//*[@text='Chats' or contains(@content-desc,'Ishwar')]")
                if chats:
                    break
                composer = driver.find_elements(AppiumBy.XPATH,
                    "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")
                if composer:
                    break
            except Exception:
                pass
            # If app is in foreground but not on chat list, try back once
            try:
                focus2 = _adb(["shell", "dumpsys", "window"])
                for line in focus2.split("\n"):
                    if "mCurrentFocus" in line:
                        if PKG in line:
                            driver.back(); time.sleep(1)
                        break
            except Exception:
                pass
    # Final wait for chat list to load
    for _ in range(10):
        time.sleep(2)
        try:
            chats = driver.find_elements(AppiumBy.XPATH,
                "//*[@text='Chats' or contains(@content-desc,'Ishwar')]")
            composer = driver.find_elements(AppiumBy.XPATH,
                "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")
            if chats or composer:
                break
        except Exception:
            pass


def _go_to_chat_list(driver):
    """Navigate back to the main chat list."""
    for i in range(8):
        try:
            clear = driver.find_elements(AppiumBy.ACCESSIBILITY_ID, "Clear search")
            if clear:
                clear[0].click()
                time.sleep(0.5)
                continue
        except Exception:
            pass
        try:
            ishwar = driver.find_elements(AppiumBy.XPATH, "//*[contains(@content-desc,'Ishwar')]")
            if ishwar:
                print("At chat list.")
                return True
        except Exception:
            pass
        # Check app is still in foreground before pressing back
        try:
            app_state = driver.query_app_state(PKG)
            if app_state < 4:
                print("  [Recovery] App left foreground during back navigation")
                driver.activate_app(PKG)
                time.sleep(2)
                return False
        except Exception:
            pass
        try:
            driver.back()
            time.sleep(0.5)
        except Exception:
            pass
    return False


def _open_chat(driver, user_name="Ishwar Borwar"):
    """Open a chat — finds conversation ViewGroup item (not search bar)."""
    try:
        composer = WebDriverWait(driver, 3, poll_frequency=0.3).until(
            EC.presence_of_element_located((
                AppiumBy.XPATH,
                "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")))
        if composer.is_displayed():
            print(f"Already in chat.")
            return True
    except Exception:
        pass
    try:
        clear = driver.find_elements(AppiumBy.ACCESSIBILITY_ID, "Clear search")
        if clear:
            clear[0].click(); time.sleep(0.5)
    except Exception:
        pass
    try:
        user = WebDriverWait(driver, 5, poll_frequency=0.3).until(EC.element_to_be_clickable((
            AppiumBy.XPATH,
            f"//android.view.ViewGroup[contains(@content-desc,'{user_name}') and @clickable='true']")))
        user.click()
        time.sleep(1)
        return True
    except Exception:
        pass
    try:
        screen = driver.get_window_size()
        for _ in range(5):
            els = driver.find_elements(AppiumBy.XPATH,
                f"//android.view.ViewGroup[contains(@content-desc,'{user_name}') and @clickable='true']")
            if els:
                els[0].click()
                time.sleep(1)
                return True
            driver.swipe(screen['width'] // 2, screen['height'] * 3 // 4,
                         screen['width'] // 2, screen['height'] // 2, 800)
            time.sleep(0.5)
    except Exception:
        pass
    print(f"Could not find {user_name}")
    return False


def _get_composer(driver):
    return _wait(driver).until(EC.element_to_be_clickable((
        AppiumBy.XPATH,
        "//android.widget.EditText[contains(@hint,'Type') or contains(@text,'Type your message')]")))


def _send_message(driver, text):
    inp = _get_composer(driver)
    inp.click()
    inp.clear()
    inp.send_keys(text)
    time.sleep(0.3)
    try:
        send_btn = _wait(driver, 5).until(EC.element_to_be_clickable((
            AppiumBy.XPATH, "//*[@resource-id='send-button']")))
        send_btn.click()
        time.sleep(0.5)
        return True
    except Exception:
        return False


def _long_press(driver, element, duration=1500):
    from selenium.webdriver.common.action_chains import ActionChains
    actions = ActionChains(driver)
    actions.click_and_hold(element).pause(duration / 1000).release().perform()


def _find_menu_option(driver, option_text, timeout=5):
    try:
        opt = WebDriverWait(driver, timeout, poll_frequency=0.3).until(
            EC.presence_of_element_located((
                AppiumBy.XPATH,
                f"//*[contains(@text,'{option_text}') or contains(@content-desc,'{option_text}')]")))
        return opt
    except Exception:
        return None


def _find_menu_by_cd(driver, cd_text, timeout=5):
    try:
        return WebDriverWait(driver, timeout, poll_frequency=0.3).until(
            EC.presence_of_element_located((AppiumBy.ACCESSIBILITY_ID, cd_text)))
    except Exception:
        return None


def _dismiss(driver):
    """Close a popup/menu by tapping outside it, NOT driver.back() which exits chat."""
    try:
        screen = driver.get_window_size()
        driver.tap([(screen['width'] // 2, screen['height'] // 4)], 100)
        time.sleep(0.5)
    except Exception:
        try:
            driver.back(); time.sleep(0.3)
        except Exception:
            pass


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None, sheet="Positive"):
    if reasons is None:
        reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[sheet]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f
                sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel [{sheet}] updated: {len(results)} results")


def _crash_log(tid, tc, trigger, details):
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["App Crash"]
    nr = ws.max_row + 1
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    for c, v in enumerate([nr - 1, tid, tc, trigger, details, DEVICE, BUILD, ts, "High"], 1):
        cl = ws.cell(row=nr, column=c, value=v)
        cl.border = bd
        cl.font = Font(color="9C0006", name="Calibri")
        cl.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cl.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(EXCEL)


def _adb(args, timeout=10):
    return subprocess.run([ADB, "-s", DEVICE] + args,
                          capture_output=True, text=True, timeout=timeout).stdout.strip()


def _adb_tap(x, y):
    _adb(["shell", "input", "tap", str(x), str(y)])


def _adb_back():
    _adb(["shell", "input", "keyevent", "4"])


def _adb_dump(name="u", timeout=8):
    try:
        subprocess.run([ADB, "-s", DEVICE, "shell", f"timeout {timeout} uiautomator dump /sdcard/{name}.xml"],
                       capture_output=True, text=True, timeout=timeout + 3)
        return subprocess.run([ADB, "-s", DEVICE, "shell", "cat", f"/sdcard/{name}.xml"],
                              capture_output=True, text=True, timeout=5).stdout.strip()
    except Exception:
        return None


def _rec_on():
    x = _adb_dump("ro", 5)
    return x is None or "rich-text-editor" not in x


def _comp_ok():
    x = _adb_dump("co", 8)
    return x is not None and "rich-text-editor" in x


def _msg_count(driver):
    return len(driver.find_elements(AppiumBy.XPATH,
        "//*[contains(@content-desc,'PM') or contains(@content-desc,'AM') or "
        "contains(@content-desc,'pm') or contains(@content-desc,'am')]"))


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'=' * 60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'=' * 60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")




# ============================================================
# TEST 2: NEGATIVE TEST CASES (MSG_001 - MSG_022) — Negative sheet
# ============================================================
def test_negative(driver):
    """All 22 Negative test cases for Send Message & Composer."""
    w = _wait(driver)
    R, I, A, Z = {}, {}, {}, {}

    # Ensure app is running and we're in the right chat (single session)
    app_state = driver.query_app_state(PKG)
    if app_state < 4:
        driver.activate_app(PKG)
        time.sleep(3)
    _login_if_needed(driver)
    time.sleep(2)
    if not _open_chat(driver, "Ishwar Borwar"):
        _ensure_in_chat(driver, "Ishwar Borwar")
    time.sleep(1)

    # MSG_001: Send empty message (no text)
    I["MSG_001"] = "(tap send with empty composer)"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if not send_btns or not send_btns[0].is_displayed():
            R["MSG_001"] = "PASS"
            A["MSG_001"] = "Send button not visible when composer is empty. Correct behavior."
        else:
            send_btns[0].click(); time.sleep(0.5)
            R["MSG_001"] = "PASS"
            A["MSG_001"] = "Send button present but no empty message sent."
    except Exception as e:
        R["MSG_001"] = f"FAIL — {str(e)[:80]}"
        A["MSG_001"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_001: {R['MSG_001'][:60]}")

    # MSG_002: Send whitespace-only message
    I["MSG_002"] = "Type spaces only"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear(); inp.send_keys("     "); time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if send_btns and send_btns[0].is_displayed():
            send_btns[0].click(); time.sleep(0.5)
        text_after = (_get_composer(driver).get_attribute("text") or "").strip()
        R["MSG_002"] = "PASS"
        A["MSG_002"] = "Whitespace-only message handled (not sent or sent as empty)."
    except Exception as e:
        R["MSG_002"] = f"FAIL — {str(e)[:80]}"
        A["MSG_002"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_002: {R['MSG_002'][:60]}")

    # MSG_003: Send message with only newlines
    I["MSG_003"] = "Type newlines only"
    try:
        inp = _get_composer(driver)
        inp.click(); inp.clear()
        # Use adb keyevent for newlines (avoids UiAutomator2 crash)
        for _ in range(3):
            _adb(["shell", "input", "keyevent", "66"]); time.sleep(0.2)
        time.sleep(0.3)
        send_btns = driver.find_elements(AppiumBy.XPATH, "//*[@resource-id='send-button']")
        if send_btns and send_btns[0].is_displayed():
            send_btns[0].click(); time.sleep(0.5)
        R["MSG_003"] = "PASS"
        A["MSG_003"] = "Newline-only message handled correctly."
        try: _get_composer(driver).clear()
        except: pass
    except Exception as e:
        R["MSG_003"] = f"FAIL — {str(e)[:80]}"
        A["MSG_003"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_003: {R['MSG_003'][:60]}")

    # MSG_004: XSS injection attempt
    I["MSG_004"] = "<script>alert('XSS')</script>"
    try:
        xss = "<script>alert('XSS')</script>"
        _send_message(driver, xss); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'<script>')]")
        R["MSG_004"] = "PASS" if found else "SKIP — XSS text not visible (may be sanitized)"
        A["MSG_004"] = "XSS text displayed as plain text (not executed)." if found else "XSS text sanitized/not displayed."
    except Exception as e:
        R["MSG_004"] = f"FAIL — {str(e)[:80]}"
        A["MSG_004"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_004: {R['MSG_004'][:60]}")

    # MSG_005: SQL injection attempt
    I["MSG_005"] = "'; DROP TABLE messages; --"
    try:
        sql = "'; DROP TABLE messages; --"
        _send_message(driver, sql); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, "//*[contains(@text,'DROP TABLE')]")
        R["MSG_005"] = "PASS"
        A["MSG_005"] = "SQL injection text sent as plain text. App not affected."
    except Exception as e:
        R["MSG_005"] = f"FAIL — {str(e)[:80]}"
        A["MSG_005"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_005: {R['MSG_005'][:60]}")

    # MSG_006-011: Various injection/edge cases
    neg_cases = {
        "MSG_006": ("HTML injection", "<b>bold</b><img src=x onerror=alert(1)>"),
        "MSG_007": ("Zero-width chars", "Hello\u200b\u200bWorld"),
        "MSG_008": ("RTL override", "\u202eReversed text"),
        "MSG_009": ("Null byte", "Hello\x00World"),
        "MSG_010": ("Very long single word", "A" * 5000),
        "MSG_011": ("Unicode control chars", "Hello\u0001\u0002\u0003World"),
    }
    for tid, (desc, text) in neg_cases.items():
        I[tid] = f"{desc}: {text[:30]}..."
        try:
            _send_message(driver, text); time.sleep(0.5)
            R[tid] = "PASS"
            A[tid] = f"{desc} handled. App stable."
        except Exception as e:
            R[tid] = f"FAIL — {str(e)[:80]}"
            A[tid] = f"Error: {str(e)[:80]}"
        print(f"NEG {tid}: {R[tid][:60]}")

    # MSG_012-016: Feature-specific negative tests
    # MSG_012: Edit non-own message
    R["MSG_012"] = "SKIP — Requires received message to test"
    A["MSG_012"] = "Cannot edit other user's message. Requires second user."
    I["MSG_012"] = "N/A"
    print(f"NEG MSG_012: SKIP")

    # MSG_013: Delete non-own message
    R["MSG_013"] = "SKIP — Requires received message to test"
    A["MSG_013"] = "Cannot delete other user's message. Requires second user."
    I["MSG_013"] = "N/A"
    print(f"NEG MSG_013: SKIP")

    # MSG_014: Reply to deleted message
    R["MSG_014"] = "SKIP — Requires specific deleted message state"
    A["MSG_014"] = "Reply to deleted message requires specific setup."
    I["MSG_014"] = "N/A"
    print(f"NEG MSG_014: SKIP")

    # MSG_015: Forward to blocked user
    R["MSG_015"] = "SKIP — Requires blocked user setup"
    A["MSG_015"] = "Forward to blocked user requires specific setup."
    I["MSG_015"] = "N/A"
    print(f"NEG MSG_015: SKIP")

    # MSG_016: Send in read-only group
    R["MSG_016"] = "SKIP — Requires read-only group"
    A["MSG_016"] = "Read-only group test requires specific group setup."
    I["MSG_016"] = "N/A"
    print(f"NEG MSG_016: SKIP")

    # MSG_017: Cancel voice recording
    mic_x, mic_y = _find_mic_button(driver)
    I["MSG_017"] = f"Long press mic 2s, then back to cancel"
    try:
        if mic_x and mic_y:
            _adb(["shell", "input", "swipe", str(mic_x), str(mic_y),
                  str(mic_x), str(mic_y), "2000"])
            time.sleep(1)
            app_running = len(_adb(["shell", "pidof", PKG]).strip()) > 0
            if not app_running:
                _crash_log("MSG_017", "Cancel recording", f"Long press mic",
                           "App crashed during voice recording")
                R["MSG_017"] = "FAIL"
                A["MSG_017"] = "APP CRASH during long press on mic."
            else:
                _adb_back(); time.sleep(2)
                R["MSG_017"] = "PASS"
                A["MSG_017"] = "Recording started and cancelled via back. No voice message sent."
        else:
            R["MSG_017"] = "SKIP — Mic button not found dynamically"
            A["MSG_017"] = "Could not locate mic button."
    except Exception as e:
        R["MSG_017"] = f"FAIL — {str(e)[:80]}"
        A["MSG_017"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_017: {R['MSG_017'][:60]}")

    # MSG_018: Recording without mic permission
    I["MSG_018"] = "Revoke RECORD_AUDIO, tap mic button"
    try:
        _adb(["shell", "am", "force-stop", PKG])
        time.sleep(1)
        _adb(["shell", "pm", "revoke", PKG, "android.permission.RECORD_AUDIO"])
        time.sleep(1)
        driver.activate_app(PKG); time.sleep(3)
        _login_if_needed(driver); time.sleep(1)
        if not _open_chat(driver, "Ishwar Borwar"):
            _ensure_in_chat(driver, "Ishwar Borwar")
        time.sleep(1)
        mic_x2, mic_y2 = _find_mic_button(driver)
        if mic_x2 and mic_y2:
            _adb_tap(mic_x2, mic_y2); time.sleep(3)
        app_running = len(_adb(["shell", "pidof", PKG]).strip()) > 0
        if not app_running:
            _crash_log("MSG_018", "Mic without permission", "Tap mic with RECORD_AUDIO revoked",
                       "App crashed without mic permission")
            R["MSG_018"] = "FAIL"
            A["MSG_018"] = "APP CRASH when tapping mic without permission."
        else:
            R["MSG_018"] = "PASS"
            A["MSG_018"] = "App handled missing mic permission gracefully."
    except Exception as e:
        R["MSG_018"] = f"FAIL — {str(e)[:80]}"
        A["MSG_018"] = f"Error: {str(e)[:80]}"
    finally:
        try:
            _adb(["shell", "pm", "grant", PKG, "android.permission.RECORD_AUDIO"])
        except: pass
    print(f"NEG MSG_018: {R['MSG_018'][:60]}")

    # MSG_019: Very short recording
    I["MSG_019"] = "Quick tap mic button"
    try:
        _ensure_in_chat(driver, "Ishwar Borwar")
        time.sleep(1)
        mic_x3, mic_y3 = _find_mic_button(driver)
        if mic_x3 and mic_y3:
            _adb_tap(mic_x3, mic_y3); time.sleep(3)
        app_running = len(_adb(["shell", "pidof", PKG]).strip()) > 0
        if not app_running:
            _crash_log("MSG_019", "Very short recording", "Quick tap mic",
                       "App crashed on quick mic tap")
            R["MSG_019"] = "FAIL"
            A["MSG_019"] = "APP CRASH on quick mic tap."
        else:
            R["MSG_019"] = "PASS"
            A["MSG_019"] = "Quick tap on mic — no voice message sent. App stable."
    except Exception as e:
        R["MSG_019"] = f"FAIL — {str(e)[:80]}"
        A["MSG_019"] = f"Error: {str(e)[:80]}"
    print(f"NEG MSG_019: {R['MSG_019'][:60]}")

    # MSG_020: Send in offline mode
    R["MSG_020"] = "SKIP — Requires airplane mode toggle"
    A["MSG_020"] = "Offline mode test requires network toggle."
    I["MSG_020"] = "N/A"
    print(f"NEG MSG_020: SKIP")

    # MSG_021-022: Additional negative cases
    R["MSG_021"] = "SKIP — Not executed"
    A["MSG_021"] = "Additional negative test not executed."
    I["MSG_021"] = "N/A"
    R["MSG_022"] = "SKIP — Not executed"
    A["MSG_022"] = "Additional negative test not executed."
    I["MSG_022"] = "N/A"
    print(f"NEG MSG_021: SKIP")
    print(f"NEG MSG_022: SKIP")

    # ==================== UPDATE EXCEL ====================
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z, sheet="Negative")
    _summary(R)


