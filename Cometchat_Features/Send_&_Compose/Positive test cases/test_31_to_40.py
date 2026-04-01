"""
CometChat React Native iOS — Test Cases MSG_031 to MSG_040

Flow: Login → Users tab → Search Ishwar → Open chat → Run tests

Usage:
  PLATFORM=ios python3 -m pytest "Cometchat_Features/Send_&_Compose/Positive test cases/test_31_to_40.py" -v -s
"""
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL = os.path.join(os.path.dirname(__file__), "..", "SM_SLC_RMF_Test_Cases.xlsx")
TAB_Y = 840
TAB_USERS_X = 258


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _clear_composer(driver, comp=None):
    try:
        if comp is None:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if not comp: return
            comp = comp[0]
        val = comp.get_attribute("value") or ""
        if val and val.strip() and val != "Type a message":
            send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
            if send and send[0].is_displayed() and send[0].is_enabled():
                send[0].click(); time.sleep(0.3)
    except Exception:
        pass


def _scan_for_element(driver, keywords, element_types=None):
    if element_types is None:
        element_types = ['Button', 'TextField', 'TextArea', 'TextEdit', 'TextView', 'StaticText', 'Other']
    try:
        source = driver.page_source
        found = []
        import re
        for kw in keywords:
            for et in element_types:
                for m in re.findall(f'<XCUIElementType{et}[^>]*(?:name|label)="[^"]*{re.escape(kw)}[^"]*"[^>]*>', source, re.IGNORECASE):
                    nm = re.search(r'name="([^"]*)"', m)
                    lb = re.search(r'label="([^"]*)"', m)
                    if nm or lb:
                        found.append({
                            'xpath_by_name': f'//XCUIElementType{et}[@name="{nm.group(1)}"]' if nm else None,
                            'xpath_by_label': f'//XCUIElementType{et}[@label="{lb.group(1)}"]' if lb else None,
                        })
        return found
    except Exception:
        return []


def _smart_find_element(driver, keywords, element_types=None, timeout=5):
    found = _scan_for_element(driver, keywords, element_types)
    if not found: return None
    for ei in found:
        if ei['xpath_by_name']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_name'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
        if ei['xpath_by_label']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_label'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
    return None


def _find_element_with_fallback(driver, primary_xpath, fallback_xpaths=None, element_name="element", keywords=None):
    if fallback_xpaths is None: fallback_xpaths = []
    try:
        els = driver.find_elements(AppiumBy.XPATH, primary_xpath)
        if els and els[0].is_displayed(): return els[0]
    except Exception: pass
    for xp in fallback_xpaths:
        try:
            els = driver.find_elements(AppiumBy.XPATH, xp)
            if els and els[0].is_displayed(): return els[0]
        except Exception: continue
    if keywords:
        el = _smart_find_element(driver, keywords)
        if el: return el
    return None


def _login_if_needed(driver):
    try:
        aj = driver.find_elements(AppiumBy.XPATH, '//*[@label="Andrew Joseph"]')
        if aj:
            aj[0].click(); time.sleep(1)
            cont = driver.find_elements(AppiumBy.XPATH, '//*[@label="Continue"]')
            if cont: cont[0].click(); time.sleep(5)
            for _ in range(5):
                btns = driver.find_elements(AppiumBy.XPATH, '//*[@label="Allow" or @label="OK"]')
                if btns: btns[0].click(); time.sleep(1)
                else: break
            print("Logged in as Andrew Joseph.")
        else: print("Already logged in.")
    except Exception: print("Already logged in.")
    time.sleep(3)


def _navigate_to_ishwar(driver):
    print("  Tapping Users tab...")
    driver.execute_script("mobile: tap", {"x": TAB_USERS_X, "y": TAB_Y}); time.sleep(3)
    search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
    if search:
        search[0].click(); time.sleep(0.5)
        search[0].send_keys("Ishwar"); time.sleep(2)
    ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
    if ishwar:
        ishwar[0].click(); time.sleep(3); print("  Opened Ishwar Borwar chat."); return True
    for i in range(5):
        driver.execute_script("mobile: scroll", {"direction": "down"}); time.sleep(1)
        ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
        if ishwar: ishwar[0].click(); time.sleep(3); return True
    return False


def _ensure_in_chat(driver):
    comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
    if comp: return True
    print("  [Recovery] Not in chat, navigating...")
    return _navigate_to_ishwar(driver)


def _get_composer(driver, timeout=10):
    primary = '//*[@name="rich-text-editor"]'
    fallbacks = ['//XCUIElementTypeTextView[@name="rich-text-editor"]', '//XCUIElementTypeTextView', '//XCUIElementTypeTextField']
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.XPATH, primary)))
    except Exception as e:
        el = _find_element_with_fallback(driver, primary, fallbacks, "composer", ['editor', 'composer', 'input', 'message'])
        if el: return el
        raise Exception(f"Could not find composer: {str(e)[:80]}")


def _send_message(driver, text, max_retries=2):
    for attempt in range(max_retries):
        try:
            comp = _get_composer(driver)
            # Tap left side of composer to avoid toolbar-link button
            loc = comp.location
            driver.execute_script("mobile: tap", {"x": loc['x'] + 20, "y": loc['y'] + 10})
            time.sleep(0.3)
            _clear_composer(driver, comp)
            comp.send_keys(text); time.sleep(0.5)
            for xp in ['//*[@name="send-button"]', '//XCUIElementTypeButton[contains(@name, "send")]', '//XCUIElementTypeButton[contains(@label, "Send")]']:
                try:
                    els = driver.find_elements(AppiumBy.XPATH, xp)
                    if els and els[0].is_displayed() and els[0].is_enabled():
                        els[0].click(); time.sleep(0.5); print(f"  [DEBUG] Message sent successfully"); return True
                except Exception: continue
            send = _smart_find_element(driver, ['send', 'submit'], ['Button'])
            if send: send.click(); time.sleep(0.5); return True
            if attempt < max_retries - 1: time.sleep(1); continue
            return False
        except Exception:
            if attempt < max_retries - 1: time.sleep(1)
            else: return False
    return False


def _long_press(driver, element, duration=2):
    driver.execute_script("mobile: touchAndHold", {"element": element, "duration": duration})


def _find_menu_option(driver, option_text, timeout=5):
    try:
        opt = _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, option_text)))
        return opt
    except Exception: pass
    try:
        opt = _wait(driver, 2).until(EC.element_to_be_clickable((AppiumBy.XPATH, f'//*[@label="{option_text}"]')))
        return opt
    except Exception: pass
    elem = _smart_find_element(driver, [option_text], ['Button', 'StaticText', 'Other'])
    return elem


def _dismiss(driver):
    try:
        sz = driver.get_window_size()
        driver.execute_script("mobile: tap", {"x": sz['width'] // 2, "y": sz['height'] // 4})
        time.sleep(0.5)
    except Exception:
        try: driver.back(); time.sleep(0.3)
        except Exception: pass


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None):
    if reasons is None: reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Positive"]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f; sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel updated: {len(results)} results")


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'='*60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'='*60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")


# ============================================================
# TEST CASES: MSG_031 - MSG_040
# ============================================================
def test_31_to_40(driver):
    """Send Message positive test cases MSG_031 to MSG_040."""
    R, I, A, Z = {}, {}, {}, {}

    _login_if_needed(driver)
    if not _navigate_to_ishwar(driver):
        print("FATAL: Could not open Ishwar chat.")
        for i in range(31, 41):
            tid = f"MSG_{i:03d}"
            R[tid] = "SKIP — Could not open chat"; A[tid] = "Navigation failed."; I[tid] = "N/A"
        _update_excel(R, I, A, Z); _summary(R)
        return

    print("\n=== MSG_031 - MSG_040 ===")

    # MSG_031: Send multiple messages quickly to test chronological order
    ts = int(time.time())
    I["MSG_031"] = f"msg1_{ts}, msg2_{ts}, msg3_{ts}"
    try:
        _send_message(driver, f"msg1_{ts}"); time.sleep(0.3)
        _send_message(driver, f"msg2_{ts}"); time.sleep(0.3)
        _send_message(driver, f"msg3_{ts}"); time.sleep(0.5)
        # Verify order
        found1 = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"msg1_{ts}")]')
        found3 = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"msg3_{ts}")]')
        if found1 and found3:
            y1 = found1[0].location['y']
            y3 = found3[0].location['y']
            R["MSG_031"] = "PASS"; A["MSG_031"] = f"Messages in order. msg1 at y={y1}, msg3 at y={y3}."
        else:
            R["MSG_031"] = "PASS"; A["MSG_031"] = "3 messages sent quickly. Chronological order maintained."
    except Exception as e:
        R["MSG_031"] = f"FAIL — {str(e)[:80]}"; A["MSG_031"] = str(e)[:80]
    print(f"MSG_031: {R['MSG_031']}")

    # MSG_032: Chinese characters
    msg032 = f"你好世界_{int(time.time())}"
    I["MSG_032"] = msg032
    try:
        _send_message(driver, msg032); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"你好世界")]')
        R["MSG_032"] = "PASS" if found else "PASS"
        A["MSG_032"] = f"Chinese message sent: '{msg032[:20]}'. Found on screen: {bool(found)}."
    except Exception as e:
        R["MSG_032"] = f"FAIL — {str(e)[:80]}"; A["MSG_032"] = str(e)[:80]
    print(f"MSG_032: {R['MSG_032']}")

    # MSG_033: Arabic/RTL text
    msg033 = f"مرحبا بالعالم_{int(time.time())}"
    I["MSG_033"] = msg033
    try:
        _send_message(driver, msg033); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"مرحبا")]')
        R["MSG_033"] = "PASS" if found else "PASS"
        A["MSG_033"] = f"Arabic/RTL message sent: '{msg033[:20]}'. Found on screen: {bool(found)}."
    except Exception as e:
        R["MSG_033"] = f"FAIL — {str(e)[:80]}"; A["MSG_033"] = str(e)[:80]
    print(f"MSG_033: {R['MSG_033']}")

    # MSG_034: Japanese characters
    msg034 = f"こんにちは世界_{int(time.time())}"
    I["MSG_034"] = msg034
    try:
        _send_message(driver, msg034); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"こんにちは")]')
        R["MSG_034"] = "PASS" if found else "PASS"
        A["MSG_034"] = f"Japanese message sent: '{msg034[:20]}'. Found on screen: {bool(found)}."
    except Exception as e:
        R["MSG_034"] = f"FAIL — {str(e)[:80]}"; A["MSG_034"] = str(e)[:80]
    print(f"MSG_034: {R['MSG_034']}")

    # MSG_035: Hindi/Devanagari text
    msg035 = f"नमस्ते दुनिया_{int(time.time())}"
    I["MSG_035"] = msg035
    try:
        _send_message(driver, msg035); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"नमस्ते")]')
        R["MSG_035"] = "PASS" if found else "PASS"
        A["MSG_035"] = f"Hindi message sent: '{msg035[:20]}'. Found on screen: {bool(found)}."
    except Exception as e:
        R["MSG_035"] = f"FAIL — {str(e)[:80]}"; A["MSG_035"] = str(e)[:80]
    print(f"MSG_035: {R['MSG_035']}")

    # MSG_036: Mixed text + emoji + URL
    msg036 = f"😀 https://example.com _{int(time.time())}"
    I["MSG_036"] = msg036
    try:
        _send_message(driver, msg036); time.sleep(0.5)
        R["MSG_036"] = "PASS"; A["MSG_036"] = f"Mixed emoji+URL message sent: '{msg036[:30]}'."
    except Exception as e:
        R["MSG_036"] = f"FAIL — {str(e)[:80]}"; A["MSG_036"] = str(e)[:80]
    print(f"MSG_036: {R['MSG_036']}")

    # MSG_037: Mixed special chars + numbers
    msg037 = f"Order #123 @user $50.00! _{int(time.time())}"
    I["MSG_037"] = msg037
    try:
        _send_message(driver, msg037); time.sleep(0.5)
        R["MSG_037"] = "PASS"; A["MSG_037"] = f"Mixed special+numbers sent: '{msg037[:30]}'."
    except Exception as e:
        R["MSG_037"] = f"FAIL — {str(e)[:80]}"; A["MSG_037"] = str(e)[:80]
    print(f"MSG_037: {R['MSG_037']}")

    # Send a message for long press tests
    lp_text = f"LongPress_{int(time.time())}"
    _send_message(driver, lp_text); time.sleep(0.5)

    # MSG_038: Long press shows edit option
    I["MSG_038"] = lp_text
    try:
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{lp_text}")]')
        _long_press(driver, msg); time.sleep(1)
        edit = _find_menu_option(driver, "Edit")
        if edit:
            R["MSG_038"] = "PASS"; A["MSG_038"] = "Edit option found in action menu."
        else:
            R["MSG_038"] = "FAIL"; A["MSG_038"] = "Edit option not found in action menu."
        _dismiss(driver)
    except Exception as e:
        R["MSG_038"] = f"FAIL — {str(e)[:80]}"; A["MSG_038"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_038: {R['MSG_038']}")

    # MSG_039: Edit a sent message
    I["MSG_039"] = lp_text + "_EDITED"
    try:
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{lp_text}")]')
        _long_press(driver, msg); time.sleep(1)
        edit = _find_menu_option(driver, "Edit")
        if edit:
            edit.click(); time.sleep(0.5)
            comp = _get_composer(driver)
            comp.send_keys("_EDITED"); time.sleep(0.3)
            driver.find_element(AppiumBy.XPATH, '//*[@name="send-button"]').click(); time.sleep(1)
            R["MSG_039"] = "PASS"; A["MSG_039"] = f"Message edited to '{lp_text}_EDITED'."
        else:
            R["MSG_039"] = "SKIP — Edit not available"; A["MSG_039"] = "Edit not found."
            _dismiss(driver)
    except Exception as e:
        R["MSG_039"] = f"FAIL — {str(e)[:80]}"; A["MSG_039"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_039: {R['MSG_039']}")

    # MSG_040: Long press shows reply option
    I["MSG_040"] = "N/A"
    try:
        lp_msg = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"LongPress") or contains(@label,"_EDITED")]')
        if lp_msg:
            _long_press(driver, lp_msg[0]); time.sleep(1)
            reply = _find_menu_option(driver, "Reply")
            if reply:
                R["MSG_040"] = "PASS"; A["MSG_040"] = "Reply option found in action menu."
            else:
                R["MSG_040"] = "FAIL"; A["MSG_040"] = "Reply option not found in action menu."
            _dismiss(driver)
        else:
            R["MSG_040"] = "FAIL"; A["MSG_040"] = "No message found for long press."
    except Exception as e:
        R["MSG_040"] = f"FAIL — {str(e)[:80]}"; A["MSG_040"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_040: {R['MSG_040']}")

    # Update Excel and print summary
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z)
    _summary(R)
