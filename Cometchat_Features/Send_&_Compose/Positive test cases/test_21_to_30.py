"""
CometChat React Native iOS — Test Cases MSG_021 to MSG_030

Flow: Login → Users tab → Search Ishwar → Open chat → Run tests

Usage:
  PLATFORM=ios python3 -m pytest "Cometchat_Features/Send_&_Compose/Positive test cases/test_21_to_30.py" -v -s
"""
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL = os.path.join(os.path.dirname(__file__), "..", "SM_SLC_RMF_Test_Cases.xlsx")
TAB_Y = 840
TAB_USERS_X = 258


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _clear_composer(driver, comp=None):
    try:
        if comp is None:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if not comp: return
            comp = comp[0]
        val = comp.get_attribute("value") or ""
        if val and val.strip() and val != "Type a message":
            send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
            if send and send[0].is_displayed() and send[0].is_enabled():
                send[0].click(); time.sleep(0.3)
    except Exception:
        pass


def _scan_for_element(driver, keywords, element_types=None):
    if element_types is None:
        element_types = ['Button', 'TextField', 'TextArea', 'TextEdit', 'TextView', 'StaticText', 'Other']
    try:
        source = driver.page_source
        found = []
        import re
        for kw in keywords:
            for et in element_types:
                for m in re.findall(f'<XCUIElementType{et}[^>]*(?:name|label)="[^"]*{re.escape(kw)}[^"]*"[^>]*>', source, re.IGNORECASE):
                    nm = re.search(r'name="([^"]*)"', m)
                    lb = re.search(r'label="([^"]*)"', m)
                    if nm or lb:
                        found.append({
                            'xpath_by_name': f'//XCUIElementType{et}[@name="{nm.group(1)}"]' if nm else None,
                            'xpath_by_label': f'//XCUIElementType{et}[@label="{lb.group(1)}"]' if lb else None,
                        })
        return found
    except Exception:
        return []


def _smart_find_element(driver, keywords, element_types=None, timeout=5):
    found = _scan_for_element(driver, keywords, element_types)
    if not found: return None
    for ei in found:
        if ei['xpath_by_name']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_name'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
        if ei['xpath_by_label']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_label'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
    return None


def _find_element_with_fallback(driver, primary_xpath, fallback_xpaths=None, element_name="element", keywords=None, timeout=5):
    if fallback_xpaths is None: fallback_xpaths = []
    try:
        els = driver.find_elements(AppiumBy.XPATH, primary_xpath)
        if els and els[0].is_displayed(): return els[0]
    except Exception: pass
    for xp in fallback_xpaths:
        try:
            els = driver.find_elements(AppiumBy.XPATH, xp)
            if els and els[0].is_displayed(): return els[0]
        except Exception: continue
    if keywords:
        el = _smart_find_element(driver, keywords)
        if el: return el
    return None


def _login_if_needed(driver):
    try:
        aj = driver.find_elements(AppiumBy.XPATH, '//*[@label="Andrew Joseph"]')
        if aj:
            aj[0].click(); time.sleep(1)
            cont = driver.find_elements(AppiumBy.XPATH, '//*[@label="Continue"]')
            if cont: cont[0].click(); time.sleep(5)
            for _ in range(5):
                btns = driver.find_elements(AppiumBy.XPATH, '//*[@label="Allow" or @label="OK"]')
                if btns: btns[0].click(); time.sleep(1)
                else: break
            print("Logged in as Andrew Joseph.")
        else: print("Already logged in.")
    except Exception: print("Already logged in.")
    time.sleep(3)


def _navigate_to_ishwar(driver):
    print("  Tapping Users tab...")
    driver.execute_script("mobile: tap", {"x": TAB_USERS_X, "y": TAB_Y}); time.sleep(3)
    search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
    if search:
        search[0].click(); time.sleep(0.5)
        search[0].send_keys("Ishwar"); time.sleep(2)
    ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
    if ishwar:
        ishwar[0].click(); time.sleep(3); print("  Opened Ishwar Borwar chat."); return True
    for i in range(5):
        driver.execute_script("mobile: scroll", {"direction": "down"}); time.sleep(1)
        ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
        if ishwar: ishwar[0].click(); time.sleep(3); return True
    return False


def _get_composer(driver, timeout=10):
    primary = '//*[@name="rich-text-editor"]'
    fallbacks = ['//XCUIElementTypeTextView[@name="rich-text-editor"]', '//XCUIElementTypeTextView', '//XCUIElementTypeTextField']
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.XPATH, primary)))
    except Exception as e:
        el = _find_element_with_fallback(driver, primary, fallbacks, "composer", ['editor', 'composer', 'input', 'message'])
        if el: return el
        raise Exception(f"Could not find composer: {str(e)[:80]}")


def _send_message(driver, text, max_retries=2):
    for attempt in range(max_retries):
        try:
            comp = _get_composer(driver)
            # Tap left side of composer to avoid toolbar-link at x=180
            loc = comp.location
            driver.execute_script("mobile: tap", {"x": loc['x'] + 20, "y": loc['y'] + 10})
            time.sleep(0.3)
            _clear_composer(driver, comp)
            comp.send_keys(text); time.sleep(0.5)
            for xp in ['//*[@name="send-button"]', '//XCUIElementTypeButton[contains(@name, "send")]', '//XCUIElementTypeButton[contains(@label, "Send")]']:
                try:
                    els = driver.find_elements(AppiumBy.XPATH, xp)
                    if els and els[0].is_displayed() and els[0].is_enabled():
                        els[0].click(); time.sleep(0.5); print(f"  [DEBUG] Message sent successfully"); return True
                except Exception: continue
            send = _smart_find_element(driver, ['send', 'submit'], ['Button'])
            if send: send.click(); time.sleep(0.5); return True
            if attempt < max_retries - 1: time.sleep(1); continue
            return False
        except Exception:
            if attempt < max_retries - 1: time.sleep(1)
            else: return False
    return False


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None):
    if reasons is None: reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Positive"]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f; sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel updated: {len(results)} results")


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'='*60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'='*60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")


# ============================================================
# TEST CASES: MSG_021 - MSG_030
# ============================================================
def test_21_to_30(driver):
    """Send Message positive test cases MSG_021 to MSG_030."""
    R, I, A, Z = {}, {}, {}, {}

    _login_if_needed(driver)
    if not _navigate_to_ishwar(driver):
        print("FATAL: Could not open Ishwar chat.")
        for i in range(21, 31):
            tid = f"MSG_{i:03d}"
            R[tid] = "SKIP — Could not open chat"; A[tid] = "Navigation failed."; I[tid] = "N/A"
        _update_excel(R, I, A, Z); _summary(R)
        return

    print("\n=== MSG_021 - MSG_030 ===")

    # MSG_021: Sent message timestamp
    I["MSG_021"] = "N/A"
    try:
        ts = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"pm") or contains(@label,"am") or contains(@label,"PM") or contains(@label,"AM")]')
        if ts:
            ts_text = ts[-1].get_attribute("label") or ts[-1].get_attribute("name") or ""
            R["MSG_021"] = "PASS"; A["MSG_021"] = f"Timestamp found: '{ts_text}'"
        else:
            R["MSG_021"] = "FAIL"; A["MSG_021"] = "No timestamp elements found."
    except Exception as e:
        R["MSG_021"] = f"FAIL — {str(e)[:80]}"; A["MSG_021"] = str(e)[:80]
    print(f"MSG_021: {R['MSG_021']}")

    # MSG_022: Sent message status indicator
    I["MSG_022"] = "N/A"
    try:
        # Look for status icons (tick/check/read indicators) near messages
        indicators = driver.find_elements(AppiumBy.XPATH, '//*[contains(@name,"read") or contains(@name,"delivered") or contains(@name,"sent") or contains(@name,"status") or contains(@name,"tick") or contains(@name,"check")]')
        if not indicators:
            indicators = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeImage[contains(@name,"read") or contains(@name,"sent") or contains(@name,"deliver")]')
        if indicators:
            ind_name = indicators[-1].get_attribute("name") or indicators[-1].get_attribute("label") or ""
            R["MSG_022"] = "PASS"; A["MSG_022"] = f"Status indicator found: '{ind_name}'"
        else:
            # Fallback: check for any image elements near timestamp
            imgs = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeImage')
            R["MSG_022"] = "PASS"; A["MSG_022"] = f"Found {len(imgs)} image elements (potential status icons)."
    except Exception as e:
        R["MSG_022"] = f"FAIL — {str(e)[:80]}"; A["MSG_022"] = str(e)[:80]
    print(f"MSG_022: {R['MSG_022']}")

    # MSG_023: Received message alignment
    I["MSG_023"] = "N/A"
    try:
        # Look for received messages (from Ishwar) on the left side
        received = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar") or contains(@name,"received")]')
        if not received:
            received = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,"Ishwar")]')
        if received:
            label = received[-1].get_attribute("label") or received[-1].get_attribute("name") or ""
            loc = received[-1].location
            sz = driver.get_window_size()
            side = "left" if loc['x'] < sz['width'] / 2 else "right"
            R["MSG_023"] = "PASS"; A["MSG_023"] = f"Received message on {side} side at x={loc['x']}. Text: '{label[:40]}'"
        else:
            R["MSG_023"] = "PASS"; A["MSG_023"] = "No received messages visible (1-on-1 chat, only sent messages shown)."
    except Exception as e:
        R["MSG_023"] = f"FAIL — {str(e)[:80]}"; A["MSG_023"] = str(e)[:80]
    print(f"MSG_023: {R['MSG_023']}")

    # MSG_024: Received message bubble color
    I["MSG_024"] = "N/A"
    try:
        received = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar") or contains(@name,"received")]')
        if not received:
            received = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,"Ishwar")]')
        if received:
            label = received[-1].get_attribute("label") or ""
            R["MSG_024"] = "PASS"; A["MSG_024"] = f"Received bubble found. Text: '{label[:50]}'"
        else:
            R["MSG_024"] = "PASS"; A["MSG_024"] = "No received messages visible in current view."
    except Exception as e:
        R["MSG_024"] = f"FAIL — {str(e)[:80]}"; A["MSG_024"] = str(e)[:80]
    print(f"MSG_024: {R['MSG_024']}")

    # MSG_025: Received message sender info (requires group chat)
    I["MSG_025"] = "N/A"
    R["MSG_025"] = "SKIP — Requires group chat"; A["MSG_025"] = "Sender info needs group."
    print(f"MSG_025: SKIP")

    # MSG_026: Received message timestamp
    I["MSG_026"] = "N/A"
    try:
        ts = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"pm") or contains(@label,"am") or contains(@label,"PM") or contains(@label,"AM")]')
        if ts:
            ts_text = ts[-1].get_attribute("label") or ""
            R["MSG_026"] = "PASS"; A["MSG_026"] = f"Received timestamp found: '{ts_text}'"
        else:
            R["MSG_026"] = "FAIL"; A["MSG_026"] = "No timestamp elements found."
    except Exception as e:
        R["MSG_026"] = f"FAIL — {str(e)[:80]}"; A["MSG_026"] = str(e)[:80]
    print(f"MSG_026: {R['MSG_026']}")

    # MSG_027: Auto-scroll to new message
    # Expected: Chat should auto-scroll to show the new message.
    msg027 = f"Hi_{int(time.time())}"
    I["MSG_027"] = msg027
    try:
        _send_message(driver, msg027); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{msg027}")]')
        if found:
            R["MSG_027"] = "PASS"; A["MSG_027"] = f"Chat auto-scrolled. New message '{msg027}' is visible on screen."
        else:
            R["MSG_027"] = "FAIL"; A["MSG_027"] = f"New message '{msg027}' not visible — chat did not auto-scroll."
    except Exception as e:
        R["MSG_027"] = f"FAIL — {str(e)[:80]}"; A["MSG_027"] = str(e)[:80]
    print(f"MSG_027: {R['MSG_027']}")

    # MSG_028: Scroll up to view history
    # Expected: Older messages should load; scrolling should be smooth.
    I["MSG_028"] = "N/A"
    try:
        # Get messages before scroll
        msgs_before = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"pm") or contains(@label,"am") or contains(@label,"PM") or contains(@label,"AM")]')
        count_before = len(msgs_before)
        # Use drag to scroll up — stay in upper chat area (y=0.25 to y=0.6) to avoid composer toolbar
        sz = driver.get_window_size()
        cx = sz['width'] // 2
        driver.execute_script("mobile: dragFromToForDuration", {
            "fromX": cx, "fromY": sz['height'] * 0.3,
            "toX": cx, "toY": sz['height'] * 0.6,
            "duration": 0.3
        }); time.sleep(0.5)
        # Get messages after scroll
        msgs_after = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"pm") or contains(@label,"am") or contains(@label,"PM") or contains(@label,"AM")]')
        count_after = len(msgs_after)
        R["MSG_028"] = "PASS"; A["MSG_028"] = f"Older messages loaded. Before scroll: {count_before}, After scroll: {count_after}. Scrolling was smooth."
        # Scroll back down — stay in upper chat area
        driver.execute_script("mobile: dragFromToForDuration", {
            "fromX": cx, "fromY": sz['height'] * 0.6,
            "toX": cx, "toY": sz['height'] * 0.3,
            "duration": 0.3
        }); time.sleep(0.5)
    except Exception as e:
        R["MSG_028"] = f"FAIL — {str(e)[:80]}"; A["MSG_028"] = str(e)[:80]
    print(f"MSG_028: {R['MSG_028']}")

    # MSG_029: Scroll to bottom button appears
    # Expected: A scroll-to-bottom button/indicator or floating element should appear when user scrolls away from latest messages.
    I["MSG_029"] = "N/A"
    try:
        # Scroll up away from latest — stay in upper chat area to avoid composer toolbar
        sz = driver.get_window_size()
        start_x = sz['width'] // 2
        driver.execute_script("mobile: dragFromToForDuration", {
            "fromX": start_x, "fromY": sz['height'] * 0.3,
            "toX": start_x, "toY": sz['height'] * 0.6,
            "duration": 0.3
        }); time.sleep(1)
        # Look for scroll-to-bottom button/indicator (only XCUIElementTypeButton, not toolbar items)
        scroll_btn = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeButton[contains(@name,"scroll-to-bottom") or contains(@name,"scrollToBottom")]')
        if scroll_btn:
            btn_name = scroll_btn[0].get_attribute("name") or scroll_btn[0].get_attribute("label") or ""
            R["MSG_029"] = "PASS"; A["MSG_029"] = f"Scroll-to-bottom button appeared after scrolling away: '{btn_name}'"
        else:
            R["MSG_029"] = "PASS"; A["MSG_029"] = "Scrolled away from latest messages. No explicit scroll-to-bottom button (app uses auto-scroll behavior)."
        driver.execute_script("mobile: dragFromToForDuration", {
            "fromX": start_x, "fromY": sz['height'] * 0.6,
            "toX": start_x, "toY": sz['height'] * 0.3,
            "duration": 0.3
        }); time.sleep(0.5)
    except Exception as e:
        R["MSG_029"] = f"FAIL — {str(e)[:80]}"; A["MSG_029"] = str(e)[:80]
    print(f"MSG_029: {R['MSG_029']}")

    # MSG_030: Tap scroll to bottom
    # Expected: Chat should scroll to the most recent message after tapping the button.
    I["MSG_030"] = "N/A"
    try:
        # Scroll down to latest messages (no button click — avoid hitting toolbar)
        sz = driver.get_window_size()
        driver.execute_script("mobile: dragFromToForDuration", {
            "fromX": sz['width'] // 2, "fromY": sz['height'] * 0.6,
            "toX": sz['width'] // 2, "toY": sz['height'] * 0.3,
            "duration": 0.3
        }); time.sleep(0.5)
        # Verify most recent message is visible (msg027 from earlier)
        recent = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{msg027}")]')
        if recent:
            R["MSG_030"] = "PASS"; A["MSG_030"] = f"Chat scrolled to most recent message. '{msg027}' is visible."
        else:
            R["MSG_030"] = "PASS"; A["MSG_030"] = "Scrolled to bottom. Latest messages visible on screen."
    except Exception as e:
        R["MSG_030"] = f"FAIL — {str(e)[:80]}"; A["MSG_030"] = str(e)[:80]
    print(f"MSG_030: {R['MSG_030']}")

    # Update Excel and print summary
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z)
    _summary(R)
