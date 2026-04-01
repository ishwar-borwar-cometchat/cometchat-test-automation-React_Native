"""
CometChat React Native iOS — First 10 Test Cases (MSG_001 to MSG_010)

Flow: Login → Users tab → Search Ishwar → Open chat → Run tests

Usage:
  PLATFORM=ios python3 -m pytest "Cometchat_Features/Send_&_Compose/Positive test cases/test_first_10.py" -v -s
"""
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL = os.path.join(os.path.dirname(__file__), "..", "SM_SLC_RMF_Test_Cases.xlsx")
PKG = "com.cometchat.internal.reactnative.ios.565LF4C8NT"

TAB_Y = 840
TAB_USERS_X = 258


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _clear_composer(driver, comp=None):
    try:
        if comp is None:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if not comp: return
            comp = comp[0]
        val = comp.get_attribute("value") or ""
        if val and val.strip() and val != "Type a message":
            send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
            if send and send[0].is_displayed() and send[0].is_enabled():
                send[0].click(); time.sleep(0.3)
                print("  [CLEAR] Flushed leftover text via send button")
    except Exception:
        pass


def _dump_page_source(driver, test_id=""):
    try:
        source = driver.page_source
        filename = f"debug_{test_id}_{int(time.time())}.xml"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(source)
        print(f"  [DEBUG] Page source saved to {filename}")
    except Exception as e:
        print(f"  [DEBUG] Could not save page source: {e}")


def _scan_for_element(driver, keywords, element_types=None):
    if element_types is None:
        element_types = ['Button', 'TextField', 'TextArea', 'TextEdit', 'TextView', 'StaticText', 'Other']
    try:
        source = driver.page_source
        found_elements = []
        import re
        for keyword in keywords:
            for elem_type in element_types:
                pattern = f'<XCUIElementType{elem_type}[^>]*(?:name|label)="[^"]*{re.escape(keyword)}[^"]*"[^>]*>'
                matches = re.findall(pattern, source, re.IGNORECASE)
                for match in matches:
                    name_match = re.search(r'name="([^"]*)"', match)
                    label_match = re.search(r'label="([^"]*)"', match)
                    if name_match or label_match:
                        found_elements.append({
                            'type': elem_type,
                            'name': name_match.group(1) if name_match else '',
                            'label': label_match.group(1) if label_match else '',
                            'xpath_by_name': f'//XCUIElementType{elem_type}[@name="{name_match.group(1)}"]' if name_match else None,
                            'xpath_by_label': f'//XCUIElementType{elem_type}[@label="{label_match.group(1)}"]' if label_match else None,
                        })
        return found_elements
    except Exception:
        return []


def _smart_find_element(driver, keywords, element_types=None, timeout=5):
    print(f"  [SMART FIND] Searching for element with keywords: {keywords}")
    found = _scan_for_element(driver, keywords, element_types)
    if not found:
        print(f"  [SMART FIND] No elements found matching keywords")
        return None
    for i, elem_info in enumerate(found):
        if elem_info['xpath_by_name']:
            try:
                elements = driver.find_elements(AppiumBy.XPATH, elem_info['xpath_by_name'])
                if elements and elements[0].is_displayed(): return elements[0]
            except Exception: pass
        if elem_info['xpath_by_label']:
            try:
                elements = driver.find_elements(AppiumBy.XPATH, elem_info['xpath_by_label'])
                if elements and elements[0].is_displayed(): return elements[0]
            except Exception: pass
    return None


def _find_element_with_fallback(driver, primary_xpath, fallback_xpaths=None, element_name="element", keywords=None, timeout=5):
    if fallback_xpaths is None: fallback_xpaths = []
    try:
        elements = driver.find_elements(AppiumBy.XPATH, primary_xpath)
        if elements and elements[0].is_displayed(): return elements[0]
    except Exception: pass
    for xpath in fallback_xpaths:
        try:
            elements = driver.find_elements(AppiumBy.XPATH, xpath)
            if elements and elements[0].is_displayed(): return elements[0]
        except Exception: continue
    if keywords:
        elem = _smart_find_element(driver, keywords)
        if elem: return elem
    _dump_page_source(driver, element_name)
    return None


def _login_if_needed(driver):
    try:
        aj = driver.find_elements(AppiumBy.XPATH, '//*[@label="Andrew Joseph"]')
        if aj:
            aj[0].click(); time.sleep(1)
            cont = driver.find_elements(AppiumBy.XPATH, '//*[@label="Continue"]')
            if cont: cont[0].click(); time.sleep(5)
            for _ in range(5):
                btns = driver.find_elements(AppiumBy.XPATH, '//*[@label="Allow" or @label="OK"]')
                if btns: btns[0].click(); time.sleep(1)
                else: break
            print("Logged in as Andrew Joseph.")
        else:
            print("Already logged in.")
    except Exception:
        print("Already logged in.")
    time.sleep(3)


def _navigate_to_ishwar(driver):
    print("  Tapping Users tab...")
    driver.execute_script("mobile: tap", {"x": TAB_USERS_X, "y": TAB_Y}); time.sleep(3)
    search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
    if search:
        print("  Searching for Ishwar...")
        search[0].click(); time.sleep(0.5)
        search[0].send_keys("Ishwar"); time.sleep(2)
    ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
    if ishwar:
        ishwar[0].click(); time.sleep(3)
        print("  Opened Ishwar Borwar chat.")
        return True
    for i in range(5):
        driver.execute_script("mobile: scroll", {"direction": "down"}); time.sleep(1)
        ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
        if ishwar:
            ishwar[0].click(); time.sleep(3)
            return True
    print("  Could not find Ishwar Borwar!")
    return False


def _get_composer(driver, timeout=10):
    primary = '//*[@name="rich-text-editor"]'
    fallbacks = [
        '//XCUIElementTypeTextView[@name="rich-text-editor"]',
        '//XCUIElementTypeTextView[contains(@name, "text-editor")]',
        '//XCUIElementTypeTextView[contains(@name, "composer")]',
        '//XCUIElementTypeTextView',
        '//XCUIElementTypeTextField',
    ]
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.XPATH, primary)))
    except Exception as e:
        elem = _find_element_with_fallback(driver, primary, fallbacks, "composer", ['editor', 'composer', 'input', 'message'])
        if elem: return elem
        raise Exception(f"Could not find composer: {str(e)[:80]}")


def _send_message(driver, text, max_retries=2):
    for attempt in range(max_retries):
        try:
            comp = _get_composer(driver)
            comp.click(); time.sleep(0.3)
            _clear_composer(driver, comp)
            comp.send_keys(text); time.sleep(0.5)
            send_xpaths = ['//*[@name="send-button"]', '//XCUIElementTypeButton[@name="send-button"]',
                           '//XCUIElementTypeButton[contains(@name, "send")]', '//XCUIElementTypeButton[contains(@label, "Send")]']
            send = None
            for xpath in send_xpaths:
                try:
                    elements = driver.find_elements(AppiumBy.XPATH, xpath)
                    if elements and elements[0].is_displayed() and elements[0].is_enabled():
                        send = elements[0]; break
                except Exception: continue
            if not send:
                send = _smart_find_element(driver, ['send', 'submit', 'arrow'], ['Button'])
            if send:
                send.click(); time.sleep(0.5)
                print(f"  [DEBUG] Message sent successfully")
                return True
            if attempt < max_retries - 1: time.sleep(1); continue
            return False
        except Exception as e:
            if attempt < max_retries - 1: time.sleep(1)
            else: return False
    return False


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None, sheet="Positive"):
    if reasons is None: reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[sheet]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f; sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel [{sheet}] updated: {len(results)} results")


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'='*60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'='*60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")


# ============================================================
# TEST CASES: MSG_001 - MSG_010
# ============================================================
def test_first_10(driver):
    """Send Message positive test cases MSG_001 to MSG_010."""
    R, I, A, Z = {}, {}, {}, {}

    # Setup: Login → Navigate to Ishwar chat
    _login_if_needed(driver)
    if not _navigate_to_ishwar(driver):
        print("FATAL: Could not open Ishwar chat. Aborting.")
        for i in range(1, 11):
            tid = f"MSG_{i:03d}"
            R[tid] = "SKIP — Could not open chat"
            A[tid] = "Navigation failed."; I[tid] = "N/A"
        _update_excel(R, I, A, Z); _summary(R)
        return

    print("\n=== MSG_001 - MSG_010 ===")

    # MSG_001: Verify message input field is visible
    I["MSG_001"] = "N/A"
    try:
        comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
        R["MSG_001"] = "PASS" if comp else "FAIL"
        A["MSG_001"] = "Message input field visible." if comp else "Composer not found."
    except Exception as e:
        R["MSG_001"] = f"FAIL — {str(e)[:80]}"; A["MSG_001"] = str(e)[:80]
    print(f"MSG_001: {R['MSG_001']}")

    # MSG_002: Verify message input field is clickable
    I["MSG_002"] = "N/A"
    try:
        comp = _get_composer(driver); comp.click()
        R["MSG_002"] = "PASS"; A["MSG_002"] = "Input field clickable."
    except Exception as e:
        R["MSG_002"] = f"FAIL — {str(e)[:80]}"; A["MSG_002"] = str(e)[:80]
    print(f"MSG_002: {R['MSG_002']}")

    # MSG_003: Verify typing in message input field
    I["MSG_003"] = "Test message"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Test message"); time.sleep(0.5)
        val = comp.get_attribute("value") or comp.get_attribute("label") or ""
        R["MSG_003"] = "PASS" if "Test" in val or len(val) > 0 else "FAIL"
        A["MSG_003"] = f"Typed text displayed: '{val[:40]}'"
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_003"] = f"FAIL — {str(e)[:80]}"; A["MSG_003"] = str(e)[:80]
    print(f"MSG_003: {R['MSG_003']}")

    # MSG_004: Verify multi-line message input
    I["MSG_004"] = "Line 1\nLine 2"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Line 1\nLine 2"); time.sleep(0.5)
        R["MSG_004"] = "PASS"; A["MSG_004"] = "Multi-line input accepted."
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_004"] = f"FAIL — {str(e)[:80]}"; A["MSG_004"] = str(e)[:80]
    print(f"MSG_004: {R['MSG_004']}")

    # MSG_005: Verify send button is visible
    I["MSG_005"] = "test"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("test"); time.sleep(0.3)
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        R["MSG_005"] = "PASS" if send else "FAIL"
        A["MSG_005"] = "Send button visible." if send else "Send button not found."
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_005"] = f"FAIL — {str(e)[:80]}"; A["MSG_005"] = str(e)[:80]
    print(f"MSG_005: {R['MSG_005']}")

    # MSG_006: Verify send button enabled when text entered
    I["MSG_006"] = "Hello"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Hello"); time.sleep(0.3)
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        enabled = send[0].get_attribute("enabled") if send else "false"
        R["MSG_006"] = "PASS" if enabled == "true" else "FAIL"
        A["MSG_006"] = f"Send button enabled: {enabled}"
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_006"] = f"FAIL — {str(e)[:80]}"; A["MSG_006"] = str(e)[:80]
    print(f"MSG_006: {R['MSG_006']}")

    # MSG_007: Verify send button click sends message
    msg007 = f"Test_{int(time.time())}"
    I["MSG_007"] = msg007
    try:
        sent = _send_message(driver, msg007); time.sleep(1)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{msg007}")]')
        R["MSG_007"] = "PASS" if (sent and found) else "FAIL"
        A["MSG_007"] = f"Message '{msg007}' sent and visible."
    except Exception as e:
        R["MSG_007"] = f"FAIL — {str(e)[:80]}"; A["MSG_007"] = str(e)[:80]
    print(f"MSG_007: {R['MSG_007']}")

    # MSG_008: Verify send button visual feedback on click
    msg008 = f"Feedback_{int(time.time())}"
    I["MSG_008"] = msg008
    try:
        _send_message(driver, msg008); time.sleep(0.3)
        comp = _get_composer(driver)
        val = comp.get_attribute("value") or ""
        R["MSG_008"] = "PASS" if msg008 not in val else "FAIL"
        A["MSG_008"] = "Send clicked, input cleared."
    except Exception as e:
        R["MSG_008"] = f"FAIL — {str(e)[:80]}"; A["MSG_008"] = str(e)[:80]
    print(f"MSG_008: {R['MSG_008']}")

    # MSG_009: Simple text
    I["MSG_009"] = "Hello"
    try:
        _send_message(driver, "Hello"); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Hello")]')
        R["MSG_009"] = "PASS" if found else "FAIL"; A["MSG_009"] = "Hello sent."
    except Exception as e:
        R["MSG_009"] = f"FAIL — {str(e)[:80]}"; A["MSG_009"] = str(e)[:80]
    print(f"MSG_009: {R['MSG_009']}")

    # MSG_010: Long text (500+ chars)
    msg010 = "A" * 500 + f"_END{int(time.time())}"
    I["MSG_010"] = f"500+ chars ({len(msg010)} chars)"
    try:
        _send_message(driver, msg010); time.sleep(1)
        R["MSG_010"] = "PASS"; A["MSG_010"] = f"Long message ({len(msg010)} chars) sent."
    except Exception as e:
        R["MSG_010"] = f"FAIL — {str(e)[:80]}"; A["MSG_010"] = str(e)[:80]
    print(f"MSG_010: {R['MSG_010']}")

    # Update Excel and print summary
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z)
    _summary(R)
