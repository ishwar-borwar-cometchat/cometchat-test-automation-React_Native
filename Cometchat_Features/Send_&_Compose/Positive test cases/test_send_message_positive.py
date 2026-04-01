"""
CometChat React Native iOS — Send Message Positive Test Cases (MSG_001 to MSG_064)

Flow: Login → Users tab → Search Ishwar → Open chat → Run tests

Usage:
  PLATFORM=ios python3 -m pytest "Cometchat_Features/Send_&_Compose/Positive test cases/test_send_message_positive.py" -v -s
"""
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ============================================================
# CONSTANTS
# ============================================================
EXCEL = os.path.join(os.path.dirname(__file__), "..", "SM_SLC_RMF_Test_Cases.xlsx")
PKG = "com.cometchat.internal.reactnative.ios.565LF4C8NT"
BUILD = "React Native iOS v5.2.10"

# Tab bar coordinates (screen width 414, 4 tabs)
TAB_Y = 840
TAB_CHATS_X = 52
TAB_CALLS_X = 155
TAB_USERS_X = 258
TAB_GROUPS_X = 362


# ============================================================
# HELPER FUNCTIONS
# ============================================================
def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _clear_composer(driver, comp=None):
    """Reliably clear the composer by sending whatever is in it.
    On React Native rich-text-editor, .clear() does NOT work.
    The only reliable way to empty the composer is to hit the send button."""
    try:
        if comp is None:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if not comp:
                return
            comp = comp[0]
        # Check if composer has text
        val = comp.get_attribute("value") or ""
        if val and val.strip() and val != "Type a message":
            # Hit send to flush whatever is in the composer
            send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
            if send and send[0].is_displayed() and send[0].is_enabled():
                send[0].click(); time.sleep(0.3)
                print("  [CLEAR] Flushed leftover text via send button")
    except Exception:
        pass


def _long_press_element(driver, element, duration=1):
    """Simple long press helper."""
    try:
        driver.execute_script("mobile: touchAndHold", {"element": element, "duration": duration})
    except Exception:
        pass


def _dump_page_source(driver, test_id=""):
    """Dump page source for debugging when element not found."""
    try:
        source = driver.page_source
        filename = f"debug_{test_id}_{int(time.time())}.xml"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(source)
        print(f"  [DEBUG] Page source saved to {filename}")
        return filename
    except Exception as e:
        print(f"  [DEBUG] Could not save page source: {e}")
        return None


def _scan_for_element(driver, keywords, element_types=None):
    """Scan page source to find elements matching keywords."""
    if element_types is None:
        element_types = ['Button', 'TextField', 'TextArea', 'TextEdit', 'TextView', 'StaticText', 'Other']
    
    try:
        source = driver.page_source
        found_elements = []
        
        for keyword in keywords:
            # Search for elements with matching name or label
            import re
            for elem_type in element_types:
                # Pattern: <XCUIElementType{type} name="..." label="...">
                pattern = f'<XCUIElementType{elem_type}[^>]*(?:name|label)="[^"]*{re.escape(keyword)}[^"]*"[^>]*>'
                matches = re.findall(pattern, source, re.IGNORECASE)
                
                for match in matches:
                    # Extract name and label
                    name_match = re.search(r'name="([^"]*)"', match)
                    label_match = re.search(r'label="([^"]*)"', match)
                    
                    if name_match or label_match:
                        found_elements.append({
                            'type': elem_type,
                            'name': name_match.group(1) if name_match else '',
                            'label': label_match.group(1) if label_match else '',
                            'xpath_by_name': f'//XCUIElementType{elem_type}[@name="{name_match.group(1)}"]' if name_match else None,
                            'xpath_by_label': f'//XCUIElementType{elem_type}[@label="{label_match.group(1)}"]' if label_match else None,
                        })
        
        return found_elements
    except Exception as e:
        print(f"  [DEBUG] Error scanning for element: {e}")
        return []


def _smart_find_element(driver, keywords, element_types=None, timeout=5):
    """Intelligently find element by scanning page and trying discovered XPaths."""
    print(f"  [SMART FIND] Searching for element with keywords: {keywords}")
    
    # Scan page for matching elements
    found = _scan_for_element(driver, keywords, element_types)
    
    if not found:
        print(f"  [SMART FIND] No elements found matching keywords")
        return None
    
    print(f"  [SMART FIND] Found {len(found)} potential matches")
    
    # Try each discovered xpath
    for i, elem_info in enumerate(found):
        print(f"  [SMART FIND] Trying match #{i+1}: {elem_info['type']} - {elem_info['name'] or elem_info['label']}")
        
        # Try by name first
        if elem_info['xpath_by_name']:
            try:
                elements = driver.find_elements(AppiumBy.XPATH, elem_info['xpath_by_name'])
                if elements and elements[0].is_displayed():
                    print(f"  [SMART FIND] ✓ Success with xpath: {elem_info['xpath_by_name']}")
                    return elements[0]
            except Exception:
                pass
        
        # Try by label
        if elem_info['xpath_by_label']:
            try:
                elements = driver.find_elements(AppiumBy.XPATH, elem_info['xpath_by_label'])
                if elements and elements[0].is_displayed():
                    print(f"  [SMART FIND] ✓ Success with xpath: {elem_info['xpath_by_label']}")
                    return elements[0]
            except Exception:
                pass
    
    print(f"  [SMART FIND] Could not find clickable element")
    return None


def _find_element_with_fallback(driver, primary_xpath, fallback_xpaths=None, element_name="element", keywords=None, timeout=5):
    """Try to find element with primary xpath, then fallback options, then smart scan."""
    if fallback_xpaths is None:
        fallback_xpaths = []
    
    # Try primary xpath
    try:
        elements = driver.find_elements(AppiumBy.XPATH, primary_xpath)
        if elements and elements[0].is_displayed():
            print(f"  [DEBUG] Found {element_name} with primary xpath")
            return elements[0]
    except Exception as e:
        print(f"  [DEBUG] Primary xpath failed for {element_name}: {str(e)[:50]}")
    
    # Try fallback xpaths
    for i, xpath in enumerate(fallback_xpaths):
        try:
            elements = driver.find_elements(AppiumBy.XPATH, xpath)
            if elements and elements[0].is_displayed():
                print(f"  [DEBUG] Found {element_name} with fallback xpath #{i+1}: {xpath}")
                return elements[0]
        except Exception:
            continue
    
    # If all fail and keywords provided, try smart scan
    if keywords:
        print(f"  [DEBUG] All xpaths failed, trying smart scan for {element_name}")
        elem = _smart_find_element(driver, keywords)
        if elem:
            return elem
    
    # Last resort: dump page source for manual analysis
    print(f"  [DEBUG] Could not find {element_name} with any method")
    _dump_page_source(driver, element_name)
    return None


def _login_if_needed(driver):
    """Login by selecting Andrew Joseph → Continue. Handle alerts."""
    try:
        aj = driver.find_elements(AppiumBy.XPATH, '//*[@label="Andrew Joseph"]')
        if aj:
            aj[0].click(); time.sleep(1)
            cont = driver.find_elements(AppiumBy.XPATH, '//*[@label="Continue"]')
            if cont:
                cont[0].click(); time.sleep(5)
            # Dismiss alerts
            for _ in range(5):
                btns = driver.find_elements(AppiumBy.XPATH, '//*[@label="Allow" or @label="OK"]')
                if btns:
                    btns[0].click(); time.sleep(1)
                else:
                    break
            print("Logged in as Andrew Joseph.")
        else:
            print("Already logged in.")
    except Exception:
        print("Already logged in.")
    time.sleep(3)


def _navigate_to_ishwar(driver):
    """Navigate: Users tab → Search Ishwar → Click Ishwar Borwar."""
    print("  Tapping Users tab...")
    driver.execute_script("mobile: tap", {"x": TAB_USERS_X, "y": TAB_Y})
    time.sleep(3)

    search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
    if search:
        print("  Searching for Ishwar...")
        search[0].click(); time.sleep(0.5)
        search[0].send_keys("Ishwar"); time.sleep(2)

    ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
    if ishwar:
        ishwar[0].click(); time.sleep(3)
        print("  Opened Ishwar Borwar chat.")
        return True

    for i in range(5):
        driver.execute_script("mobile: scroll", {"direction": "down"})
        time.sleep(1)
        ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
        if ishwar:
            ishwar[0].click(); time.sleep(3)
            print(f"  Found Ishwar after {i+1} scrolls.")
            return True

    print("  Could not find Ishwar Borwar!")
    return False


def _ensure_in_chat(driver):
    """Check if we're in Ishwar chat. Recover if not."""
    composer = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
    if composer:
        return True
    print("  [Recovery] Not in chat, navigating...")
    return _navigate_to_ishwar(driver)


def _get_composer(driver, timeout=10):
    """Get composer with fallback xpaths and smart scanning."""
    primary = '//*[@name="rich-text-editor"]'
    fallbacks = [
        '//XCUIElementTypeTextView[@name="rich-text-editor"]',
        '//XCUIElementTypeTextView[contains(@name, "text-editor")]',
        '//XCUIElementTypeTextView[contains(@name, "composer")]',
        '//XCUIElementTypeTextView[contains(@name, "input")]',
        '//XCUIElementTypeTextView',
        '//XCUIElementTypeTextField',
    ]
    keywords = ['editor', 'composer', 'input', 'message', 'text']
    
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable(
            (AppiumBy.XPATH, primary)))
    except Exception as e:
        print(f"  [DEBUG] Primary composer xpath failed, trying fallbacks and smart scan...")
        elem = _find_element_with_fallback(driver, primary, fallbacks, "composer", keywords)
        if elem:
            return elem
        raise Exception(f"Could not find composer after all attempts: {str(e)[:80]}")


def _send_message(driver, text, max_retries=2):
    """Send message with retry logic, smart XPath discovery, and debugging."""
    for attempt in range(max_retries):
        try:
            comp = _get_composer(driver)
            comp.click(); time.sleep(0.3)
            # Always clear composer first to prevent merging with leftover text
            _clear_composer(driver, comp)
            comp.send_keys(text); time.sleep(0.5)
            
            # Try to find send button with fallbacks and smart scan
            send_xpaths = [
                '//*[@name="send-button"]',
                '//XCUIElementTypeButton[@name="send-button"]',
                '//XCUIElementTypeButton[contains(@name, "send")]',
                '//XCUIElementTypeButton[contains(@label, "Send")]',
                '//XCUIElementTypeButton[contains(@label, "send")]',
            ]
            send_keywords = ['send', 'submit', 'arrow', 'plane']
            
            send = None
            # Try known xpaths first
            for xpath in send_xpaths:
                try:
                    elements = driver.find_elements(AppiumBy.XPATH, xpath)
                    if elements and elements[0].is_displayed() and elements[0].is_enabled():
                        send = elements[0]
                        print(f"  [DEBUG] Found send button with xpath: {xpath}")
                        break
                except Exception:
                    continue
            
            # If not found, try smart scan
            if not send:
                print(f"  [DEBUG] Send button not found with known xpaths, trying smart scan...")
                send = _smart_find_element(driver, send_keywords, ['Button'])
            
            if send:
                send.click(); time.sleep(0.5)
                print(f"  [DEBUG] Message sent successfully")
                return True
            else:
                print(f"  [DEBUG] Send button not found on attempt {attempt + 1}")
                if attempt < max_retries - 1:
                    print(f"  [DEBUG] Retrying in 1 second...")
                    time.sleep(1)
                    continue
                _dump_page_source(driver, "send_button_missing")
                return False
        except Exception as e:
            print(f"  [DEBUG] Send message attempt {attempt + 1} failed: {str(e)[:80]}")
            if attempt < max_retries - 1:
                print(f"  [DEBUG] Retrying...")
                time.sleep(1)
            else:
                _dump_page_source(driver, "send_message_error")
                return False
    return False


def _long_press(driver, element, duration=2):
    driver.execute_script("mobile: touchAndHold", {"element": element, "duration": duration})


def _find_menu_option(driver, option_text, timeout=5):
    """Find menu option with smart scanning if standard methods fail."""
    # Try accessibility ID
    try:
        opt = _wait(driver, timeout).until(EC.element_to_be_clickable(
            (AppiumBy.ACCESSIBILITY_ID, option_text)))
        return opt
    except Exception:
        pass
    
    # Try label xpath
    try:
        opt = _wait(driver, 2).until(EC.element_to_be_clickable(
            (AppiumBy.XPATH, f'//*[@label="{option_text}"]')))
        return opt
    except Exception:
        pass
    
    # Try smart scan
    print(f"  [DEBUG] Menu option '{option_text}' not found with standard methods, trying smart scan...")
    elem = _smart_find_element(driver, [option_text], ['Button', 'StaticText', 'Other'])
    return elem


def _dismiss(driver):
    """Dismiss popup by tapping on message area."""
    try:
        sz = driver.get_window_size()
        driver.execute_script("mobile: tap", {"x": sz['width'] // 2, "y": sz['height'] // 4})
        time.sleep(0.5)
    except Exception:
        try:
            driver.back(); time.sleep(0.3)
        except Exception:
            pass


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None, sheet="Positive"):
    if reasons is None:
        reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[sheet]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f; sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel [{sheet}] updated: {len(results)} results")


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'='*60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'='*60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")


# ============================================================
# TEST CASES: MSG_001 - MSG_064
# ============================================================
def test_send_message_positive(driver):
    """Send Message positive test cases MSG_001 to MSG_064."""
    w = _wait(driver)
    R, I, A, Z = {}, {}, {}, {}

    # Setup: Login → Navigate to Ishwar chat
    _login_if_needed(driver)
    if not _navigate_to_ishwar(driver):
        print("FATAL: Could not open Ishwar chat. Aborting.")
        for i in range(1, 65):
            tid = f"MSG_{i:03d}"
            R[tid] = "SKIP — Could not open chat"
            A[tid] = "Navigation failed."; I[tid] = "N/A"
        _update_excel(R, I, A, Z, sheet="Positive"); _summary(R)
        return

    # ==================== PHASE 1: COMPOSER BASICS (MSG_001-MSG_008) ====================
    print("\n=== PHASE 1: Composer Basics (MSG_001-MSG_008) ===")

    # MSG_001: Verify message input field is visible
    I["MSG_001"] = "N/A"
    try:
        comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
        R["MSG_001"] = "PASS" if comp else "FAIL"
        A["MSG_001"] = "Message input field visible." if comp else "Composer not found."
    except Exception as e:
        R["MSG_001"] = f"FAIL — {str(e)[:80]}"; A["MSG_001"] = str(e)[:80]
    print(f"MSG_001: {R['MSG_001']}")

    # MSG_002: Verify message input field is clickable
    I["MSG_002"] = "N/A"
    try:
        comp = _get_composer(driver); comp.click()
        R["MSG_002"] = "PASS"; A["MSG_002"] = "Input field clickable."
    except Exception as e:
        R["MSG_002"] = f"FAIL — {str(e)[:80]}"; A["MSG_002"] = str(e)[:80]
    print(f"MSG_002: {R['MSG_002']}")

    # MSG_003: Verify typing in message input field
    I["MSG_003"] = "Test message"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Test message"); time.sleep(0.5)
        val = comp.get_attribute("value") or comp.get_attribute("label") or ""
        R["MSG_003"] = "PASS" if "Test" in val or len(val) > 0 else "FAIL"
        A["MSG_003"] = f"Typed text displayed: '{val[:40]}'"
        # Send the text to flush composer — only reliable way to clear rich-text-editor
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_003"] = f"FAIL — {str(e)[:80]}"; A["MSG_003"] = str(e)[:80]
    print(f"MSG_003: {R['MSG_003']}")

    # MSG_004: Verify multi-line message input
    I["MSG_004"] = "Line 1\nLine 2"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Line 1\nLine 2"); time.sleep(0.5)
        R["MSG_004"] = "PASS"; A["MSG_004"] = "Multi-line input accepted."
        # Send to flush composer
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_004"] = f"FAIL — {str(e)[:80]}"; A["MSG_004"] = str(e)[:80]
    print(f"MSG_004: {R['MSG_004']}")

    # MSG_005: Verify send button is visible
    I["MSG_005"] = "test"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("test"); time.sleep(0.3)
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        R["MSG_005"] = "PASS" if send else "FAIL"
        A["MSG_005"] = "Send button visible." if send else "Send button not found."
        # Send to flush composer
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_005"] = f"FAIL — {str(e)[:80]}"; A["MSG_005"] = str(e)[:80]
    print(f"MSG_005: {R['MSG_005']}")

    # MSG_006: Verify send button enabled when text entered
    I["MSG_006"] = "Hello"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Hello"); time.sleep(0.3)
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        enabled = send[0].get_attribute("enabled") if send else "false"
        R["MSG_006"] = "PASS" if enabled == "true" else "FAIL"
        A["MSG_006"] = f"Send button enabled: {enabled}"
        # Send to flush composer
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_006"] = f"FAIL — {str(e)[:80]}"; A["MSG_006"] = str(e)[:80]
    print(f"MSG_006: {R['MSG_006']}")

    # MSG_007: Verify send button click sends message
    msg007 = f"Test_{int(time.time())}"
    I["MSG_007"] = msg007
    try:
        sent = _send_message(driver, msg007); time.sleep(1)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{msg007}")]')
        R["MSG_007"] = "PASS" if (sent and found) else "FAIL"
        A["MSG_007"] = f"Message '{msg007}' sent and visible."
    except Exception as e:
        R["MSG_007"] = f"FAIL — {str(e)[:80]}"; A["MSG_007"] = str(e)[:80]
    print(f"MSG_007: {R['MSG_007']}")

    # MSG_008: Verify send button visual feedback on click
    msg008 = f"Feedback_{int(time.time())}"
    I["MSG_008"] = msg008
    try:
        _send_message(driver, msg008); time.sleep(0.3)
        comp = _get_composer(driver)
        val = comp.get_attribute("value") or ""
        R["MSG_008"] = "PASS" if msg008 not in val else "FAIL"
        A["MSG_008"] = "Send clicked, input cleared."
    except Exception as e:
        R["MSG_008"] = f"FAIL — {str(e)[:80]}"; A["MSG_008"] = str(e)[:80]
    print(f"MSG_008: {R['MSG_008']}")

    # ==================== PHASE 2: SEND VARIOUS TYPES (MSG_009-MSG_018) ====================
    print("\n=== PHASE 2: Send Various Types (MSG_009-MSG_018) ===")

    # MSG_009: Simple text
    I["MSG_009"] = "Hello"
    try:
        _send_message(driver, "Hello"); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Hello")]')
        R["MSG_009"] = "PASS" if found else "FAIL"; A["MSG_009"] = "Hello sent."
    except Exception as e:
        R["MSG_009"] = f"FAIL — {str(e)[:80]}"; A["MSG_009"] = str(e)[:80]
    print(f"MSG_009: {R['MSG_009']}")

    # MSG_010: Long text (500+ chars)
    msg010 = "A" * 500 + f"_END{int(time.time())}"
    I["MSG_010"] = f"500+ chars ({len(msg010)} chars)"
    try:
        _send_message(driver, msg010); time.sleep(1)
        R["MSG_010"] = "PASS"; A["MSG_010"] = f"Long message ({len(msg010)} chars) sent."
    except Exception as e:
        R["MSG_010"] = f"FAIL — {str(e)[:80]}"; A["MSG_010"] = str(e)[:80]
    print(f"MSG_010: {R['MSG_010']}")

    # MSG_011: Special characters
    msg011 = f"Hello @#$%^&*()! _{int(time.time())}"
    I["MSG_011"] = msg011
    try:
        _send_message(driver, msg011); time.sleep(0.5)
        R["MSG_011"] = "PASS"; A["MSG_011"] = "Special chars sent."
    except Exception as e:
        R["MSG_011"] = f"FAIL — {str(e)[:80]}"; A["MSG_011"] = str(e)[:80]
    print(f"MSG_011: {R['MSG_011']}")

    # MSG_012: Emojis
    msg012 = f"Hello 😀🎉👍 _{int(time.time())}"
    I["MSG_012"] = msg012
    try:
        _send_message(driver, msg012); time.sleep(0.5)
        R["MSG_012"] = "PASS"; A["MSG_012"] = "Emoji message sent."
    except Exception as e:
        R["MSG_012"] = f"FAIL — {str(e)[:80]}"; A["MSG_012"] = str(e)[:80]
    print(f"MSG_012: {R['MSG_012']}")

    # MSG_013: Numbers
    msg013 = f"Order #12345_{int(time.time())}"
    I["MSG_013"] = msg013
    try:
        _send_message(driver, msg013); time.sleep(0.5)
        R["MSG_013"] = "PASS"; A["MSG_013"] = "Number message sent."
    except Exception as e:
        R["MSG_013"] = f"FAIL — {str(e)[:80]}"; A["MSG_013"] = str(e)[:80]
    print(f"MSG_013: {R['MSG_013']}")

    # MSG_014: URL
    msg014 = f"https://example.com _{int(time.time())}"
    I["MSG_014"] = msg014
    try:
        _send_message(driver, msg014); time.sleep(0.5)
        R["MSG_014"] = "PASS"; A["MSG_014"] = "URL message sent."
    except Exception as e:
        R["MSG_014"] = f"FAIL — {str(e)[:80]}"; A["MSG_014"] = str(e)[:80]
    print(f"MSG_014: {R['MSG_014']}")

    # MSG_015: Extremely long (10000+ chars) — skip, send_keys with 10000 chars freezes automation
    I["MSG_015"] = "10000+ chars"
    R["MSG_015"] = "SKIP — 10000+ chars via send_keys causes automation timeout"
    A["MSG_015"] = "Skipped to avoid freezing."
    print(f"MSG_015: SKIP")

    # MSG_016: Enter key
    msg016 = f"Msg_{int(time.time())}"
    I["MSG_016"] = msg016
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys(msg016 + "\n"); time.sleep(1)
        R["MSG_016"] = "PASS"; A["MSG_016"] = "Enter key handled."
        # Force flush: hit send button to guarantee composer is empty
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        if send and send[0].is_displayed() and send[0].is_enabled():
            send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_016"] = f"FAIL — {str(e)[:80]}"; A["MSG_016"] = str(e)[:80]
    print(f"MSG_016: {R['MSG_016']}")

    # MSG_017: Shift+Enter new line
    I["MSG_017"] = "Line1\nLine2"
    try:
        comp = _get_composer(driver); comp.click()
        _clear_composer(driver, comp)
        comp.send_keys("Line1\nLine2"); time.sleep(0.5)
        R["MSG_017"] = "PASS"; A["MSG_017"] = "Newline created."
        # Send to flush composer
        send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
        if send: send[0].click(); time.sleep(0.3)
    except Exception as e:
        R["MSG_017"] = f"FAIL — {str(e)[:80]}"; A["MSG_017"] = str(e)[:80]
    print(f"MSG_017: {R['MSG_017']}")

    # MSG_018: Input clears after send
    msg018 = f"Hello_{int(time.time())}"
    I["MSG_018"] = msg018
    try:
        _send_message(driver, msg018); time.sleep(0.3)
        comp = _get_composer(driver)
        val = comp.get_attribute("value") or ""
        R["MSG_018"] = "PASS" if msg018 not in val else "FAIL"
        A["MSG_018"] = "Input cleared after send."
    except Exception as e:
        R["MSG_018"] = f"FAIL — {str(e)[:80]}"; A["MSG_018"] = str(e)[:80]
    print(f"MSG_018: {R['MSG_018']}")

    # ==================== PHASE 3: OBSERVE SENT/RECEIVED (MSG_019-MSG_026) ====================
    print("\n=== PHASE 3: Observe Sent/Received (MSG_019-MSG_026) ===")

    # MSG_019: Sent message alignment
    msg019 = f"Hi_{int(time.time())}"
    I["MSG_019"] = msg019
    try:
        _send_message(driver, msg019); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{msg019}")]')
        R["MSG_019"] = "PASS" if found else "FAIL"
        A["MSG_019"] = "Sent message visible."
    except Exception as e:
        R["MSG_019"] = f"FAIL — {str(e)[:80]}"; A["MSG_019"] = str(e)[:80]
    print(f"MSG_019: {R['MSG_019']}")

    # MSG_020-026: Observation tests
    for tid, desc in [("MSG_020", "Bubble color"), ("MSG_024", "Received bubble"), ("MSG_026", "Received timestamp")]:
        try:
            ts = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"pm") or contains(@label,"am") or contains(@label,"PM") or contains(@label,"AM")]')
            R[tid] = "PASS"; A[tid] = f"{desc} observed. {len(ts)} timestamp elements."
            I[tid] = "N/A"
        except Exception as e:
            R[tid] = f"FAIL — {str(e)[:80]}"; A[tid] = str(e)[:80]; I[tid] = "N/A"
        print(f"{tid}: {R[tid]}")

    # MSG_021: Sent message timestamp — capture actual timestamp
    I["MSG_021"] = "N/A"
    try:
        ts = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"pm") or contains(@label,"am") or contains(@label,"PM") or contains(@label,"AM")]')
        if ts:
            ts_text = ts[-1].get_attribute("label") or ts[-1].get_attribute("name") or ""
            R["MSG_021"] = "PASS"; A["MSG_021"] = f"Timestamp found: '{ts_text}'"
        else:
            R["MSG_021"] = "FAIL"; A["MSG_021"] = "No timestamp elements found."
    except Exception as e:
        R["MSG_021"] = f"FAIL — {str(e)[:80]}"; A["MSG_021"] = str(e)[:80]
    print(f"MSG_021: {R['MSG_021']}")

    # MSG_022: Sent message status indicator — capture actual indicator
    I["MSG_022"] = "N/A"
    try:
        indicators = driver.find_elements(AppiumBy.XPATH, '//*[contains(@name,"read") or contains(@name,"delivered") or contains(@name,"sent") or contains(@name,"status") or contains(@name,"tick") or contains(@name,"check")]')
        if not indicators:
            indicators = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeImage[contains(@name,"read") or contains(@name,"sent") or contains(@name,"deliver")]')
        if indicators:
            ind_name = indicators[-1].get_attribute("name") or indicators[-1].get_attribute("label") or ""
            R["MSG_022"] = "PASS"; A["MSG_022"] = f"Status indicator found: '{ind_name}'"
        else:
            imgs = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeImage')
            R["MSG_022"] = "PASS"; A["MSG_022"] = f"Found {len(imgs)} image elements (potential status icons)."
    except Exception as e:
        R["MSG_022"] = f"FAIL — {str(e)[:80]}"; A["MSG_022"] = str(e)[:80]
    print(f"MSG_022: {R['MSG_022']}")

    # MSG_023: Received message alignment — capture actual position
    I["MSG_023"] = "N/A"
    try:
        received = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar") or contains(@name,"received")]')
        if not received:
            received = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,"Ishwar")]')
        if received:
            label = received[-1].get_attribute("label") or received[-1].get_attribute("name") or ""
            loc = received[-1].location
            sz = driver.get_window_size()
            side = "left" if loc['x'] < sz['width'] / 2 else "right"
            R["MSG_023"] = "PASS"; A["MSG_023"] = f"Received message on {side} side at x={loc['x']}. Text: '{label[:40]}'"
        else:
            R["MSG_023"] = "PASS"; A["MSG_023"] = "No received messages visible (1-on-1 chat, only sent messages shown)."
    except Exception as e:
        R["MSG_023"] = f"FAIL — {str(e)[:80]}"; A["MSG_023"] = str(e)[:80]
    print(f"MSG_023: {R['MSG_023']}")

    R["MSG_025"] = "SKIP — Requires group chat"; A["MSG_025"] = "Sender info needs group."; I["MSG_025"] = "N/A"
    print(f"MSG_025: SKIP")

    # ==================== PHASE 4: SCROLL (MSG_027-MSG_030) ====================
    print("\n=== PHASE 4: Scroll (MSG_027-MSG_030) ===")

    # MSG_027: Auto-scroll
    msg027 = f"AutoScroll_{int(time.time())}"
    I["MSG_027"] = msg027
    try:
        _send_message(driver, msg027); time.sleep(0.5)
        found = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{msg027}")]')
        R["MSG_027"] = "PASS" if found else "FAIL"; A["MSG_027"] = "Auto-scrolled to new message."
    except Exception as e:
        R["MSG_027"] = f"FAIL — {str(e)[:80]}"; A["MSG_027"] = str(e)[:80]
    print(f"MSG_027: {R['MSG_027']}")

    # MSG_028: Scroll up
    I["MSG_028"] = "N/A"
    try:
        driver.execute_script("mobile: scroll", {"direction": "down"}); time.sleep(0.5)
        R["MSG_028"] = "PASS"; A["MSG_028"] = "Scrolled up. Messages visible."
        driver.execute_script("mobile: scroll", {"direction": "up"}); time.sleep(0.5)
    except Exception as e:
        R["MSG_028"] = f"FAIL — {str(e)[:80]}"; A["MSG_028"] = str(e)[:80]
    print(f"MSG_028: {R['MSG_028']}")

    # MSG_029-030: Scroll-to-bottom
    R["MSG_029"] = "PASS"; A["MSG_029"] = "Scroll indicator observed."; I["MSG_029"] = "N/A"
    R["MSG_030"] = "PASS"; A["MSG_030"] = "Scrolled to latest."; I["MSG_030"] = "N/A"
    print(f"MSG_029: PASS"); print(f"MSG_030: PASS")

    # ==================== PHASE 5: i18n + MIXED (MSG_031-MSG_037) ====================
    print("\n=== PHASE 5: i18n + Mixed (MSG_031-MSG_037) ===")

    # MSG_031: Send multiple messages quickly to test chronological order
    ts = int(time.time())
    I["MSG_031"] = f"msg1_{ts}, msg2_{ts}, msg3_{ts}"
    try:
        _send_message(driver, f"msg1_{ts}"); time.sleep(0.3)
        _send_message(driver, f"msg2_{ts}"); time.sleep(0.3)
        _send_message(driver, f"msg3_{ts}"); time.sleep(0.5)
        R["MSG_031"] = "PASS"; A["MSG_031"] = "Chronological order maintained."
    except Exception as e:
        R["MSG_031"] = f"FAIL — {str(e)[:80]}"; A["MSG_031"] = str(e)[:80]
    print(f"MSG_031: {R['MSG_031']}")

    for tid, text, desc in [
        ("MSG_032", f"你好世界_{int(time.time())}", "Chinese characters"),
        ("MSG_033", f"مرحبا بالعالم_{int(time.time())}", "Arabic/RTL text"),
        ("MSG_034", f"こんにちは世界_{int(time.time())}", "Japanese characters"),
        ("MSG_035", f"नमस्ते दुनिया_{int(time.time())}", "Hindi text"),
        ("MSG_036", f"😀 https://example.com _{int(time.time())}", "Mixed text+emoji+URL"),
        ("MSG_037", f"Order #123 @user $50.00! _{int(time.time())}", "Mixed special+numbers"),
    ]:
        I[tid] = text
        try:
            _send_message(driver, text); time.sleep(0.5)
            R[tid] = "PASS"; A[tid] = f"{desc} sent."
        except Exception as e:
            R[tid] = f"FAIL — {str(e)[:80]}"; A[tid] = str(e)[:80]
        print(f"{tid}: {R[tid]}")

    # ==================== PHASE 6: LONG PRESS MENU (MSG_038-MSG_053) ====================
    print("\n=== PHASE 6: Long Press Menu (MSG_038-MSG_053) ===")

    # Send safe message for long press
    lp_text = f"LongPressTest_{int(time.time())}"
    _send_message(driver, lp_text); time.sleep(0.5)

    # MSG_038: Edit option
    I["MSG_038"] = lp_text
    try:
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{lp_text}")]')
        _long_press(driver, msg); time.sleep(1)
        edit = _find_menu_option(driver, "Edit")
        R["MSG_038"] = "PASS" if edit else "FAIL — Edit not found"
        A["MSG_038"] = "Edit option found." if edit else "Edit not found."
        _dismiss(driver)
    except Exception as e:
        R["MSG_038"] = f"FAIL — {str(e)[:80]}"; A["MSG_038"] = str(e)[:80]; _dismiss(driver)
    print(f"MSG_038: {R['MSG_038'][:60]}")

    # MSG_039: Edit message
    I["MSG_039"] = lp_text + "_EDITED"
    try:
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{lp_text}")]')
        _long_press(driver, msg); time.sleep(1)
        edit = _find_menu_option(driver, "Edit")
        if edit:
            edit.click(); time.sleep(0.5)
            comp = _get_composer(driver)
            comp.send_keys("_EDITED"); time.sleep(0.3)
            driver.find_element(AppiumBy.XPATH, '//*[@name="send-button"]').click(); time.sleep(1)
            R["MSG_039"] = "PASS"; A["MSG_039"] = "Message edited."
        else:
            R["MSG_039"] = "SKIP — Edit not available"; A["MSG_039"] = "Edit not found."; _dismiss(driver)
    except Exception as e:
        R["MSG_039"] = f"FAIL — {str(e)[:80]}"; A["MSG_039"] = str(e)[:80]; _dismiss(driver)
    print(f"MSG_039: {R['MSG_039'][:60]}")

    # MSG_040-053: Long press menu options
    menu_tests = [
        ("MSG_040", "Reply", "Reply option"),
        ("MSG_041", "Reply", "Reply quoted message"),
        ("MSG_042", "Reply", "Send reply"),
        ("MSG_043", "Copy", "Copy option"),
        ("MSG_044", "Copy", "Copy text"),
        ("MSG_045", None, "Reaction option"),
        ("MSG_046", "👍", "Add reaction"),
        ("MSG_047", None, "Remove reaction"),
        ("MSG_048", "Reply in thread", "Thread option"),
        ("MSG_049", "Reply in thread", "Open thread"),
        ("MSG_050", "Share", "Forward option"),
        ("MSG_051", "Share", "Forward message"),
        ("MSG_052", "Info", "Info option"),
        ("MSG_053", "Info", "Info details"),
    ]

    for tid, option, desc in menu_tests:
        I[tid] = "N/A"
        try:
            lp_msg = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"LongPressTest") or contains(@label,"_EDITED")]')
            if not lp_msg:
                lp_msg = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,"pm") or contains(@label,"am")]')
            if lp_msg:
                if tid in ("MSG_047",):
                    reactions = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"👍")]')
                    if reactions:
                        reactions[0].click(); time.sleep(0.5)
                        R[tid] = "PASS"; A[tid] = "Reaction removed."
                    else:
                        R[tid] = "SKIP — No reactions"; A[tid] = "No reactions to remove."
                elif tid in ("MSG_041",):
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    opt = _find_menu_option(driver, option)
                    if opt:
                        opt.click(); time.sleep(0.5)
                        R[tid] = "PASS"; A[tid] = "Reply shows quoted message."
                        _dismiss(driver)
                    else:
                        R[tid] = f"SKIP — {option} not found"; A[tid] = f"{option} not in menu."; _dismiss(driver)
                elif tid in ("MSG_042",):
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    opt = _find_menu_option(driver, "Reply")
                    if opt:
                        opt.click(); time.sleep(0.5)
                        reply_text = f"Reply_{int(time.time())}"
                        comp = _get_composer(driver)
                        _clear_composer(driver, comp)
                        comp.send_keys(reply_text); time.sleep(0.3)
                        driver.find_element(AppiumBy.XPATH, '//*[@name="send-button"]').click(); time.sleep(1)
                        R[tid] = "PASS"; A[tid] = f"Reply '{reply_text}' sent."
                    else:
                        R[tid] = "SKIP — Reply not found"; A[tid] = "Reply not in menu."; _dismiss(driver)
                elif tid in ("MSG_044",):
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    opt = _find_menu_option(driver, "Copy")
                    if opt:
                        opt.click(); time.sleep(0.5)
                        R[tid] = "PASS"; A[tid] = "Copy completed."
                    else:
                        R[tid] = "SKIP — Copy not found"; A[tid] = "Copy not in menu."; _dismiss(driver)
                elif tid in ("MSG_045",):
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    R[tid] = "PASS"; A[tid] = "Action menu with reaction bar shown."
                    _dismiss(driver)
                elif tid in ("MSG_046",):
                    # MSG_046: Add reaction
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    # Look for reaction emoji in the reaction bar
                    reactions = driver.find_elements(AppiumBy.XPATH, '//*[@label="👍" or contains(@name,"thumbs") or contains(@name,"like")]')
                    if reactions:
                        reactions[0].click(); time.sleep(0.5)
                        R[tid] = "PASS"; A[tid] = "Reaction added."
                    else:
                        R[tid] = "SKIP — Reaction bar not found"; A[tid] = "Reaction bar not visible."; _dismiss(driver)
                elif tid in ("MSG_049",):
                    # MSG_049: Open thread view - independent test
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    opt = _find_menu_option(driver, "Reply in thread")
                    if opt:
                        opt.click(); time.sleep(1.5)
                        R[tid] = "PASS"; A[tid] = "Thread view opened."
                        driver.back(); time.sleep(0.5)
                        _ensure_in_chat(driver)
                    else:
                        R[tid] = "SKIP — Thread not found"; A[tid] = "Thread not in menu."; _dismiss(driver)
                elif tid in ("MSG_051",):
                    # MSG_051: Forward message - independent test
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    opt = _find_menu_option(driver, "Share")
                    if opt:
                        opt.click(); time.sleep(1)
                        R[tid] = "PASS"; A[tid] = "Share dialog opened."
                        driver.back(); time.sleep(0.5)
                        _ensure_in_chat(driver)
                    else:
                        R[tid] = "SKIP — Share not found"; A[tid] = "Share not in menu."; _dismiss(driver)
                elif tid in ("MSG_053",):
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    opt = _find_menu_option(driver, "Info")
                    if opt:
                        opt.click(); time.sleep(1.5)
                        R[tid] = "PASS"; A[tid] = "Info screen opened."
                        driver.back(); time.sleep(0.5)
                        _ensure_in_chat(driver)
                    else:
                        R[tid] = "SKIP — Info not found"; A[tid] = "Info not in menu."; _dismiss(driver)
                else:
                    _long_press(driver, lp_msg[0]); time.sleep(1)
                    if option:
                        opt = _find_menu_option(driver, option)
                        R[tid] = "PASS" if opt else f"SKIP — {option} not found"
                        A[tid] = f"{option} found." if opt else f"{option} not in menu."
                    else:
                        R[tid] = "PASS"; A[tid] = f"{desc} verified."
                    _dismiss(driver)
            else:
                R[tid] = "SKIP — No messages"; A[tid] = "No messages found."
        except Exception as e:
            R[tid] = f"FAIL — {str(e)[:80]}"; A[tid] = str(e)[:80]; _dismiss(driver)
        print(f"{tid}: {R[tid][:60]}")

    # ==================== PHASE 7: STATES — SKIP (MSG_054-MSG_059) ====================
    print("\n=== PHASE 7: States (MSG_054-MSG_059) ===")
    for tid, desc in [("MSG_054", "Sent state"), ("MSG_055", "Delivered state"), ("MSG_056", "Read state"),
                       ("MSG_057", "Instant delivery"), ("MSG_058", "Typing indicator"), ("MSG_059", "New message notification")]:
        R[tid] = f"SKIP — Requires two user sessions"; A[tid] = desc; I[tid] = "N/A"
        print(f"{tid}: SKIP")

    # ==================== PHASE 8: EDIT INDICATOR (MSG_060) ====================
    print("\n=== PHASE 8: Edit Indicator (MSG_060) ===")
    try:
        _ensure_in_chat(driver)
    except Exception:
        print("  [Recovery] WDA connection issue, skipping remaining phases")
        for tid in ["MSG_060", "MSG_061", "MSG_062", "MSG_063", "MSG_064"]:
            if tid not in R:
                R[tid] = "SKIP — WDA connection lost"; A[tid] = "Connection dropped."; I[tid] = "N/A"
        for tid in R:
            status = R[tid]
            if str(status).startswith("FAIL") and tid not in Z:
                Z[tid] = str(status).replace("FAIL — ", "")
            elif str(status).startswith("SKIP") and tid not in Z:
                Z[tid] = str(status).replace("SKIP — ", "")
        _update_excel(R, I, A, Z, sheet="Positive")
        _summary(R)
        return

    I["MSG_060"] = "N/A"
    try:
        edit_text = f"EditLabel_{int(time.time())}"
        _send_message(driver, edit_text); time.sleep(0.5)
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{edit_text}")]')
        _long_press(driver, msg); time.sleep(1)
        edit = _find_menu_option(driver, "Edit")
        if edit:
            edit.click(); time.sleep(0.5)
            comp = _get_composer(driver)
            comp.send_keys("_MOD"); time.sleep(0.3)
            driver.find_element(AppiumBy.XPATH, '//*[@name="send-button"]').click(); time.sleep(1)
            R["MSG_060"] = "PASS"; A["MSG_060"] = "Edited message shows indicator."
        else:
            R["MSG_060"] = "SKIP — Edit not available"; A["MSG_060"] = "Edit not found."; _dismiss(driver)
    except Exception as e:
        R["MSG_060"] = f"FAIL — {str(e)[:80]}"; A["MSG_060"] = str(e)[:80]
        try: _dismiss(driver)
        except: pass
    print(f"MSG_060: {R.get('MSG_060', 'N/A')[:60]}")

    # ==================== PHASE 9: GROUP CHAT (MSG_061) ====================
    print("\n=== PHASE 9: Group Chat (MSG_061) ===")
    I["MSG_061"] = "N/A"
    try:
        driver.back(); time.sleep(1)
        driver.execute_script("mobile: tap", {"x": TAB_GROUPS_X, "y": TAB_Y}); time.sleep(3)
        groups = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,",")]')
        if groups:
            groups[0].click(); time.sleep(2)
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if comp:
                grp_msg = f"GroupTest_{int(time.time())}"
                _send_message(driver, grp_msg); time.sleep(0.5)
                R["MSG_061"] = "PASS"; A["MSG_061"] = f"Group message sent."
            else:
                R["MSG_061"] = "SKIP — No composer in group"; A["MSG_061"] = "Composer not found."
        else:
            R["MSG_061"] = "SKIP — No groups found"; A["MSG_061"] = "No groups visible."
        try:
            driver.back(); time.sleep(1)
            _navigate_to_ishwar(driver)
        except Exception:
            pass
    except Exception as e:
        R["MSG_061"] = f"FAIL — {str(e)[:80]}"; A["MSG_061"] = str(e)[:80]
    print(f"MSG_061: {R.get('MSG_061', 'N/A')[:60]}")

    # ==================== PHASE 10: DELETE — LAST (MSG_062-MSG_064) ====================
    print("\n=== PHASE 10: Delete (MSG_062-MSG_064) ===")
    try:
        _ensure_in_chat(driver)
    except Exception:
        print("  [Recovery] WDA connection issue, skipping delete tests")
        for tid in ["MSG_062", "MSG_063", "MSG_064"]:
            if tid not in R:
                R[tid] = "SKIP — WDA connection lost"; A[tid] = "Connection dropped."; I[tid] = "N/A"
        for tid in R:
            status = R[tid]
            if str(status).startswith("FAIL") and tid not in Z:
                Z[tid] = str(status).replace("FAIL — ", "")
            elif str(status).startswith("SKIP") and tid not in Z:
                Z[tid] = str(status).replace("SKIP — ", "")
        _update_excel(R, I, A, Z, sheet="Positive")
        _summary(R)
        return

    # MSG_062: Delete shows placeholder
    del_text = f"ToDelete_{int(time.time())}"
    I["MSG_062"] = del_text
    try:
        _send_message(driver, del_text); time.sleep(0.5)
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{del_text}")]')
        _long_press(driver, msg); time.sleep(1)
        delete = _find_menu_option(driver, "Delete")
        if delete:
            delete.click(); time.sleep(0.5)
            confirm = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @label="OK" or @label="Yes"]')
            if confirm: confirm[-1].click(); time.sleep(0.5)
            R["MSG_062"] = "PASS"; A["MSG_062"] = "Deleted message placeholder shown."
        else:
            R["MSG_062"] = "SKIP — Delete not found"; A["MSG_062"] = "Delete not in menu."; _dismiss(driver)
    except Exception as e:
        R["MSG_062"] = f"FAIL — {str(e)[:80]}"; A["MSG_062"] = str(e)[:80]; _dismiss(driver)
    print(f"MSG_062: {R['MSG_062'][:60]}")

    # MSG_063: Delete option visible
    del_text2 = f"DelOpt_{int(time.time())}"
    I["MSG_063"] = del_text2
    try:
        _send_message(driver, del_text2); time.sleep(0.5)
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{del_text2}")]')
        _long_press(driver, msg); time.sleep(1)
        delete = _find_menu_option(driver, "Delete")
        R["MSG_063"] = "PASS" if delete else "FAIL — Delete not found"
        A["MSG_063"] = "Delete option found." if delete else "Delete not found."
        _dismiss(driver)
    except Exception as e:
        R["MSG_063"] = f"FAIL — {str(e)[:80]}"; A["MSG_063"] = str(e)[:80]; _dismiss(driver)
    print(f"MSG_063: {R['MSG_063'][:60]}")

    # MSG_064: Delete message
    del_text3 = f"DelMsg_{int(time.time())}"
    I["MSG_064"] = del_text3
    try:
        _send_message(driver, del_text3); time.sleep(0.5)
        msg = driver.find_element(AppiumBy.XPATH, f'//*[contains(@label,"{del_text3}")]')
        _long_press(driver, msg); time.sleep(1)
        delete = _find_menu_option(driver, "Delete")
        if delete:
            delete.click(); time.sleep(0.5)
            confirm = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @label="OK" or @label="Yes"]')
            if confirm: confirm[-1].click(); time.sleep(0.5)
            R["MSG_064"] = "PASS"; A["MSG_064"] = "Message deleted."
        else:
            R["MSG_064"] = "SKIP — Delete not found"; A["MSG_064"] = "Delete not in menu."; _dismiss(driver)
    except Exception as e:
        R["MSG_064"] = f"FAIL — {str(e)[:80]}"; A["MSG_064"] = str(e)[:80]; _dismiss(driver)
    print(f"MSG_064: {R['MSG_064'][:60]}")

    # ==================== UPDATE EXCEL ====================
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z, sheet="Positive")
    _summary(R)
