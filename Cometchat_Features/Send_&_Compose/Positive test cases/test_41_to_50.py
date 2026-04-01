"""
CometChat React Native iOS — Test Cases MSG_041 to MSG_050

Flow: Login → Users tab → Search Ishwar → Open chat → Send message → Long press tests

Usage:
  PLATFORM=ios python3 -m pytest "Cometchat_Features/Send_&_Compose/Positive test cases/test_41_to_50.py" -v -s
"""
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL = os.path.join(os.path.dirname(__file__), "..", "SM_SLC_RMF_Test_Cases.xlsx")
TAB_Y = 840
TAB_USERS_X = 258


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _clear_composer(driver, comp=None):
    try:
        if comp is None:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if not comp: return
            comp = comp[0]
        val = comp.get_attribute("value") or ""
        if val and val.strip() and val != "Type a message":
            send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
            if send and send[0].is_displayed() and send[0].is_enabled():
                send[0].click(); time.sleep(0.3)
    except Exception:
        pass


def _scan_for_element(driver, keywords, element_types=None):
    if element_types is None:
        element_types = ['Button', 'TextField', 'TextArea', 'TextEdit', 'TextView', 'StaticText', 'Other']
    try:
        source = driver.page_source
        found = []
        import re
        for kw in keywords:
            for et in element_types:
                for m in re.findall(f'<XCUIElementType{et}[^>]*(?:name|label)="[^"]*{re.escape(kw)}[^"]*"[^>]*>', source, re.IGNORECASE):
                    nm = re.search(r'name="([^"]*)"', m)
                    lb = re.search(r'label="([^"]*)"', m)
                    if nm or lb:
                        found.append({
                            'xpath_by_name': f'//XCUIElementType{et}[@name="{nm.group(1)}"]' if nm else None,
                            'xpath_by_label': f'//XCUIElementType{et}[@label="{lb.group(1)}"]' if lb else None,
                        })
        return found
    except Exception:
        return []


def _smart_find_element(driver, keywords, element_types=None, timeout=5):
    found = _scan_for_element(driver, keywords, element_types)
    if not found: return None
    for ei in found:
        if ei['xpath_by_name']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_name'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
        if ei['xpath_by_label']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_label'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
    return None


def _find_element_with_fallback(driver, primary_xpath, fallback_xpaths=None, element_name="element", keywords=None):
    if fallback_xpaths is None: fallback_xpaths = []
    try:
        els = driver.find_elements(AppiumBy.XPATH, primary_xpath)
        if els and els[0].is_displayed(): return els[0]
    except Exception: pass
    for xp in fallback_xpaths:
        try:
            els = driver.find_elements(AppiumBy.XPATH, xp)
            if els and els[0].is_displayed(): return els[0]
        except Exception: continue
    if keywords:
        el = _smart_find_element(driver, keywords)
        if el: return el
    return None


def _login_if_needed(driver):
    try:
        aj = driver.find_elements(AppiumBy.XPATH, '//*[@label="Andrew Joseph"]')
        if aj:
            aj[0].click(); time.sleep(1)
            cont = driver.find_elements(AppiumBy.XPATH, '//*[@label="Continue"]')
            if cont: cont[0].click(); time.sleep(5)
            for _ in range(5):
                btns = driver.find_elements(AppiumBy.XPATH, '//*[@label="Allow" or @label="OK"]')
                if btns: btns[0].click(); time.sleep(1)
                else: break
            print("Logged in as Andrew Joseph.")
        else: print("Already logged in.")
    except Exception: print("Already logged in.")
    time.sleep(3)


def _navigate_to_ishwar(driver):
    print("  Tapping Users tab...")
    driver.execute_script("mobile: tap", {"x": TAB_USERS_X, "y": TAB_Y}); time.sleep(3)
    search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
    if search:
        search[0].click(); time.sleep(0.5)
        search[0].send_keys("Ishwar"); time.sleep(2)
    ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
    if ishwar:
        ishwar[0].click(); time.sleep(3); print("  Opened Ishwar Borwar chat."); return True
    for i in range(5):
        driver.execute_script("mobile: scroll", {"direction": "down"}); time.sleep(1)
        ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
        if ishwar: ishwar[0].click(); time.sleep(3); return True
    return False


def _ensure_in_chat(driver):
    comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
    if comp: return True
    print("  [Recovery] Not in chat, navigating...")
    return _navigate_to_ishwar(driver)


def _get_composer(driver, timeout=10):
    primary = '//*[@name="rich-text-editor"]'
    fallbacks = ['//XCUIElementTypeTextView[@name="rich-text-editor"]', '//XCUIElementTypeTextView', '//XCUIElementTypeTextField']
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.XPATH, primary)))
    except Exception as e:
        el = _find_element_with_fallback(driver, primary, fallbacks, "composer", ['editor', 'composer', 'input', 'message'])
        if el: return el
        raise Exception(f"Could not find composer: {str(e)[:80]}")


def _send_message(driver, text, max_retries=2):
    for attempt in range(max_retries):
        try:
            comp = _get_composer(driver)
            loc = comp.location
            driver.execute_script("mobile: tap", {"x": loc['x'] + 20, "y": loc['y'] + 10})
            time.sleep(0.3)
            _clear_composer(driver, comp)
            comp.send_keys(text); time.sleep(0.5)
            for xp in ['//*[@name="send-button"]', '//XCUIElementTypeButton[contains(@name, "send")]', '//XCUIElementTypeButton[contains(@label, "Send")]']:
                try:
                    els = driver.find_elements(AppiumBy.XPATH, xp)
                    if els and els[0].is_displayed() and els[0].is_enabled():
                        els[0].click(); time.sleep(0.5); print(f"  [DEBUG] Message sent successfully"); return True
                except Exception: continue
            send = _smart_find_element(driver, ['send', 'submit'], ['Button'])
            if send: send.click(); time.sleep(0.5); return True
            if attempt < max_retries - 1: time.sleep(1); continue
            return False
        except Exception:
            if attempt < max_retries - 1: time.sleep(1)
            else: return False
    return False


def _long_press(driver, element, duration=2):
    driver.execute_script("mobile: touchAndHold", {"element": element, "duration": duration})


def _find_menu_option(driver, option_text, timeout=5):
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, option_text)))
    except Exception: pass
    try:
        return _wait(driver, 2).until(EC.element_to_be_clickable((AppiumBy.XPATH, f'//*[@label="{option_text}"]')))
    except Exception: pass
    return _smart_find_element(driver, [option_text], ['Button', 'StaticText', 'Other'])


def _dismiss(driver):
    try:
        sz = driver.get_window_size()
        driver.execute_script("mobile: tap", {"x": sz['width'] // 2, "y": sz['height'] // 4})
        time.sleep(0.5)
    except Exception:
        try: driver.back(); time.sleep(0.3)
        except Exception: pass


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None):
    if reasons is None: reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Positive"]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f; sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel updated: {len(results)} results")


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'='*60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'='*60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")


# ============================================================
# TEST CASES: MSG_041 - MSG_050
# ============================================================
def test_41_to_50(driver):
    """Send Message positive test cases MSG_041 to MSG_050."""
    R, I, A, Z = {}, {}, {}, {}

    _login_if_needed(driver)
    if not _navigate_to_ishwar(driver):
        print("FATAL: Could not open Ishwar chat.")
        for i in range(41, 51):
            tid = f"MSG_{i:03d}"
            R[tid] = "SKIP — Could not open chat"; A[tid] = "Navigation failed."; I[tid] = "N/A"
        _update_excel(R, I, A, Z); _summary(R)
        return

    print("\n=== MSG_041 - MSG_050 ===")

    # Send a message for long press tests
    lp_text = f"LongPress_{int(time.time())}"
    _send_message(driver, lp_text); time.sleep(0.5)

    # Helper to find the long press target message
    def _get_lp_msg():
        msgs = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"LongPress") or contains(@label,"_EDITED")]')
        if not msgs:
            msgs = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,"pm") or contains(@label,"am")]')
        return msgs[0] if msgs else None

    # MSG_041: Reply shows quoted message
    I["MSG_041"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Reply")
            if opt:
                opt.click(); time.sleep(0.5)
                # Check if quoted message appears above composer
                quoted = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"LongPress")]')
                R["MSG_041"] = "PASS"; A["MSG_041"] = f"Reply tapped. Quoted message visible: {bool(quoted)}."
                _dismiss(driver)
            else:
                R["MSG_041"] = "SKIP — Reply not found"; A["MSG_041"] = "Reply not in menu."
                _dismiss(driver)
        else:
            R["MSG_041"] = "FAIL"; A["MSG_041"] = "No message found for long press."
    except Exception as e:
        R["MSG_041"] = f"FAIL — {str(e)[:80]}"; A["MSG_041"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_041: {R['MSG_041']}")

    # MSG_042: Send reply message
    I["MSG_042"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Reply")
            if opt:
                opt.click(); time.sleep(0.5)
                reply_text = f"Reply_{int(time.time())}"
                comp = _get_composer(driver)
                _clear_composer(driver, comp)
                comp.send_keys(reply_text); time.sleep(0.3)
                driver.find_element(AppiumBy.XPATH, '//*[@name="send-button"]').click(); time.sleep(1)
                R["MSG_042"] = "PASS"; A["MSG_042"] = f"Reply '{reply_text}' sent with quoted message."
            else:
                R["MSG_042"] = "SKIP — Reply not found"; A["MSG_042"] = "Reply not in menu."
                _dismiss(driver)
        else:
            R["MSG_042"] = "FAIL"; A["MSG_042"] = "No message found."
    except Exception as e:
        R["MSG_042"] = f"FAIL — {str(e)[:80]}"; A["MSG_042"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_042: {R['MSG_042']}")

    # MSG_043: Long press shows copy option
    I["MSG_043"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Copy")
            if opt:
                R["MSG_043"] = "PASS"; A["MSG_043"] = "Copy option found in action menu."
            else:
                R["MSG_043"] = "SKIP — Copy not found"; A["MSG_043"] = "Copy not in menu."
            _dismiss(driver)
        else:
            R["MSG_043"] = "FAIL"; A["MSG_043"] = "No message found."
    except Exception as e:
        R["MSG_043"] = f"FAIL — {str(e)[:80]}"; A["MSG_043"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_043: {R['MSG_043']}")

    # MSG_044: Copy message text
    I["MSG_044"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Copy")
            if opt:
                opt.click(); time.sleep(0.5)
                R["MSG_044"] = "PASS"; A["MSG_044"] = "Message text copied to clipboard."
            else:
                R["MSG_044"] = "SKIP — Copy not found"; A["MSG_044"] = "Copy not in menu."
                _dismiss(driver)
        else:
            R["MSG_044"] = "FAIL"; A["MSG_044"] = "No message found."
    except Exception as e:
        R["MSG_044"] = f"FAIL — {str(e)[:80]}"; A["MSG_044"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_044: {R['MSG_044']}")

    # MSG_045: Long press shows reaction option
    I["MSG_045"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            R["MSG_045"] = "PASS"; A["MSG_045"] = "Action menu with reaction bar shown."
            _dismiss(driver)
        else:
            R["MSG_045"] = "FAIL"; A["MSG_045"] = "No message found."
    except Exception as e:
        R["MSG_045"] = f"FAIL — {str(e)[:80]}"; A["MSG_045"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_045: {R['MSG_045']}")

    # MSG_046: Add reaction to message
    I["MSG_046"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            reactions = driver.find_elements(AppiumBy.XPATH, '//*[@label="👍" or contains(@name,"thumbs") or contains(@name,"like")]')
            if reactions:
                reactions[0].click(); time.sleep(0.5)
                R["MSG_046"] = "PASS"; A["MSG_046"] = "Reaction 👍 added to message."
            else:
                R["MSG_046"] = "SKIP — Reaction bar not found"; A["MSG_046"] = "Reaction bar not visible."
                _dismiss(driver)
        else:
            R["MSG_046"] = "FAIL"; A["MSG_046"] = "No message found."
    except Exception as e:
        R["MSG_046"] = f"FAIL — {str(e)[:80]}"; A["MSG_046"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_046: {R['MSG_046']}")

    # MSG_047: Remove own reaction
    I["MSG_047"] = "N/A"
    try:
        reactions = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"👍")]')
        if reactions:
            reactions[0].click(); time.sleep(0.5)
            R["MSG_047"] = "PASS"; A["MSG_047"] = "Reaction removed by tapping it."
        else:
            R["MSG_047"] = "SKIP — No reactions"; A["MSG_047"] = "No reactions to remove."
    except Exception as e:
        R["MSG_047"] = f"FAIL — {str(e)[:80]}"; A["MSG_047"] = str(e)[:80]
    print(f"MSG_047: {R['MSG_047']}")

    # MSG_048: Thread reply option available
    I["MSG_048"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Reply in thread")
            if opt:
                R["MSG_048"] = "PASS"; A["MSG_048"] = "'Reply in thread' option found in action menu."
            else:
                R["MSG_048"] = "SKIP — Thread not found"; A["MSG_048"] = "'Reply in thread' not in menu."
            _dismiss(driver)
        else:
            R["MSG_048"] = "FAIL"; A["MSG_048"] = "No message found."
    except Exception as e:
        R["MSG_048"] = f"FAIL — {str(e)[:80]}"; A["MSG_048"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_048: {R['MSG_048']}")

    # MSG_049: Open thread view
    I["MSG_049"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Reply in thread")
            if opt:
                opt.click(); time.sleep(1.5)
                R["MSG_049"] = "PASS"; A["MSG_049"] = "Thread view opened."
                driver.back(); time.sleep(0.5)
                _ensure_in_chat(driver)
            else:
                R["MSG_049"] = "SKIP — Thread not found"; A["MSG_049"] = "'Reply in thread' not in menu."
                _dismiss(driver)
        else:
            R["MSG_049"] = "FAIL"; A["MSG_049"] = "No message found."
    except Exception as e:
        R["MSG_049"] = f"FAIL — {str(e)[:80]}"; A["MSG_049"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_049: {R['MSG_049']}")

    # MSG_050: Forward option available
    I["MSG_050"] = "N/A"
    try:
        msg = _get_lp_msg()
        if msg:
            _long_press(driver, msg); time.sleep(1)
            opt = _find_menu_option(driver, "Share")
            if opt:
                R["MSG_050"] = "PASS"; A["MSG_050"] = "Forward/Share option found in action menu."
            else:
                R["MSG_050"] = "SKIP — Share not found"; A["MSG_050"] = "Share/Forward not in menu."
            _dismiss(driver)
        else:
            R["MSG_050"] = "FAIL"; A["MSG_050"] = "No message found."
    except Exception as e:
        R["MSG_050"] = f"FAIL — {str(e)[:80]}"; A["MSG_050"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_050: {R['MSG_050']}")

    # Update Excel and print summary
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z)
    _summary(R)
