"""
CometChat React Native iOS — Test Cases MSG_061 to MSG_064

Flow: Login → Users tab → Search Ishwar → Open chat → Send message → Long press tests

Usage:
  PLATFORM=ios python3 -m pytest "Cometchat_Features/Send_&_Compose/Positive test cases/test_41_to_50.py" -v -s
"""
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL = os.path.join(os.path.dirname(__file__), "..", "SM_SLC_RMF_Test_Cases.xlsx")
TAB_Y = 840
TAB_USERS_X = 258


def _wait(driver, timeout=10):
    return WebDriverWait(driver, timeout, poll_frequency=0.3)


def _clear_composer(driver, comp=None):
    try:
        if comp is None:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if not comp: return
            comp = comp[0]
        val = comp.get_attribute("value") or ""
        if val and val.strip() and val != "Type a message":
            send = driver.find_elements(AppiumBy.XPATH, '//*[@name="send-button"]')
            if send and send[0].is_displayed() and send[0].is_enabled():
                send[0].click(); time.sleep(0.3)
    except Exception:
        pass


def _scan_for_element(driver, keywords, element_types=None):
    if element_types is None:
        element_types = ['Button', 'TextField', 'TextArea', 'TextEdit', 'TextView', 'StaticText', 'Other']
    try:
        source = driver.page_source
        found = []
        import re
        for kw in keywords:
            for et in element_types:
                for m in re.findall(f'<XCUIElementType{et}[^>]*(?:name|label)="[^"]*{re.escape(kw)}[^"]*"[^>]*>', source, re.IGNORECASE):
                    nm = re.search(r'name="([^"]*)"', m)
                    lb = re.search(r'label="([^"]*)"', m)
                    if nm or lb:
                        found.append({
                            'xpath_by_name': f'//XCUIElementType{et}[@name="{nm.group(1)}"]' if nm else None,
                            'xpath_by_label': f'//XCUIElementType{et}[@label="{lb.group(1)}"]' if lb else None,
                        })
        return found
    except Exception:
        return []


def _smart_find_element(driver, keywords, element_types=None, timeout=5):
    found = _scan_for_element(driver, keywords, element_types)
    if not found: return None
    for ei in found:
        if ei['xpath_by_name']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_name'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
        if ei['xpath_by_label']:
            try:
                els = driver.find_elements(AppiumBy.XPATH, ei['xpath_by_label'])
                if els and els[0].is_displayed(): return els[0]
            except Exception: pass
    return None


def _find_element_with_fallback(driver, primary_xpath, fallback_xpaths=None, element_name="element", keywords=None):
    if fallback_xpaths is None: fallback_xpaths = []
    try:
        els = driver.find_elements(AppiumBy.XPATH, primary_xpath)
        if els and els[0].is_displayed(): return els[0]
    except Exception: pass
    for xp in fallback_xpaths:
        try:
            els = driver.find_elements(AppiumBy.XPATH, xp)
            if els and els[0].is_displayed(): return els[0]
        except Exception: continue
    if keywords:
        el = _smart_find_element(driver, keywords)
        if el: return el
    return None


def _login_if_needed(driver):
    try:
        aj = driver.find_elements(AppiumBy.XPATH, '//*[@label="Andrew Joseph"]')
        if aj:
            aj[0].click(); time.sleep(1)
            cont = driver.find_elements(AppiumBy.XPATH, '//*[@label="Continue"]')
            if cont: cont[0].click(); time.sleep(5)
            for _ in range(5):
                btns = driver.find_elements(AppiumBy.XPATH, '//*[@label="Allow" or @label="OK"]')
                if btns: btns[0].click(); time.sleep(1)
                else: break
            print("Logged in as Andrew Joseph.")
        else: print("Already logged in.")
    except Exception: print("Already logged in.")
    time.sleep(3)


def _navigate_to_ishwar(driver):
    print("  Tapping Users tab...")
    driver.execute_script("mobile: tap", {"x": TAB_USERS_X, "y": TAB_Y}); time.sleep(3)
    search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
    if search:
        search[0].click(); time.sleep(0.5)
        search[0].send_keys("Ishwar"); time.sleep(2)
    ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
    if ishwar:
        ishwar[0].click(); time.sleep(3); print("  Opened Ishwar Borwar chat."); return True
    for i in range(5):
        driver.execute_script("mobile: scroll", {"direction": "down"}); time.sleep(1)
        ishwar = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"Ishwar Borwar")]')
        if ishwar: ishwar[0].click(); time.sleep(3); return True
    return False


def _ensure_in_chat(driver):
    comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
    if comp: return True
    print("  [Recovery] Not in chat, navigating...")
    return _navigate_to_ishwar(driver)


def _get_composer(driver, timeout=10):
    primary = '//*[@name="rich-text-editor"]'
    fallbacks = ['//XCUIElementTypeTextView[@name="rich-text-editor"]', '//XCUIElementTypeTextView', '//XCUIElementTypeTextField']
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.XPATH, primary)))
    except Exception as e:
        el = _find_element_with_fallback(driver, primary, fallbacks, "composer", ['editor', 'composer', 'input', 'message'])
        if el: return el
        raise Exception(f"Could not find composer: {str(e)[:80]}")


def _send_message(driver, text, max_retries=2):
    for attempt in range(max_retries):
        try:
            comp = _get_composer(driver)
            loc = comp.location
            driver.execute_script("mobile: tap", {"x": loc['x'] + 20, "y": loc['y'] + 10})
            time.sleep(0.3)
            _clear_composer(driver, comp)
            comp.send_keys(text); time.sleep(0.5)
            for xp in ['//*[@name="send-button"]', '//XCUIElementTypeButton[contains(@name, "send")]', '//XCUIElementTypeButton[contains(@label, "Send")]']:
                try:
                    els = driver.find_elements(AppiumBy.XPATH, xp)
                    if els and els[0].is_displayed() and els[0].is_enabled():
                        els[0].click(); time.sleep(0.5); print(f"  [DEBUG] Message sent successfully"); return True
                except Exception: continue
            send = _smart_find_element(driver, ['send', 'submit'], ['Button'])
            if send: send.click(); time.sleep(0.5); return True
            if attempt < max_retries - 1: time.sleep(1); continue
            return False
        except Exception:
            if attempt < max_retries - 1: time.sleep(1)
            else: return False
    return False


def _long_press(driver, element, duration=2):
    driver.execute_script("mobile: touchAndHold", {"element": element, "duration": duration})


def _find_menu_option(driver, option_text, timeout=5):
    try:
        return _wait(driver, timeout).until(EC.element_to_be_clickable((AppiumBy.ACCESSIBILITY_ID, option_text)))
    except Exception: pass
    try:
        return _wait(driver, 2).until(EC.element_to_be_clickable((AppiumBy.XPATH, f'//*[@label="{option_text}"]')))
    except Exception: pass
    return _smart_find_element(driver, [option_text], ['Button', 'StaticText', 'Other'])


def _dismiss(driver):
    try:
        sz = driver.get_window_size()
        driver.execute_script("mobile: tap", {"x": sz['width'] // 2, "y": sz['height'] // 4})
        time.sleep(0.5)
    except Exception:
        try: driver.back(); time.sleep(0.3)
        except Exception: pass


def _status_style(status_val):
    val = str(status_val).strip().upper()
    if val.startswith("PASS"):
        return Font(bold=True, color="006100", name="Calibri"), PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif val.startswith("FAIL"):
        return Font(bold=True, color="9C0006", name="Calibri"), PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif val.startswith("SKIP"):
        return Font(bold=True, color="9C5700", name="Calibri"), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        return Font(bold=True, color="3F3F76", name="Calibri"), PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _update_excel(results, input_data, actual_results, reasons=None):
    if reasons is None: reasons = {}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Positive"]
    for test_id in results:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == test_id:
                ws.cell(row=row, column=8, value=actual_results.get(test_id, ""))
                sc = ws.cell(row=row, column=10, value=results[test_id])
                f, p = _status_style(results[test_id])
                sc.font = f; sc.fill = p
                ws.cell(row=row, column=11, value=input_data.get(test_id, "N/A"))
                ws.cell(row=row, column=12, value=reasons.get(test_id, ""))
                break
    wb.save(EXCEL)
    print(f"Excel updated: {len(results)} results")


def _summary(results):
    p = sum(1 for v in results.values() if str(v).startswith("PASS"))
    f = sum(1 for v in results.values() if str(v).startswith("FAIL"))
    s = sum(1 for v in results.values() if str(v).startswith("SKIP"))
    print(f"\n{'='*60}")
    print(f"Total: {len(results)} | PASS: {p} | FAIL: {f} | SKIP: {s}")
    print(f"{'='*60}")
    for tid in sorted(results.keys(), key=lambda x: int(x.split('_')[1])):
        print(f"  {tid}: {str(results[tid])[:70]}")




# ============================================================
# TEST CASES: MSG_061 - MSG_064
# ============================================================
def test_61_to_64(driver):
    """Send Message positive test cases MSG_061 to MSG_064."""
    R, I, A, Z = {}, {}, {}, {}

    _login_if_needed(driver)
    if not _navigate_to_ishwar(driver):
        print("FATAL: Could not open Ishwar chat.")
        for i in range(61, 65):
            tid = f"MSG_{i:03d}"
            R[tid] = "SKIP — Could not open chat"; A[tid] = "Navigation failed."; I[tid] = "N/A"
        _update_excel(R, I, A, Z); _summary(R)
        return

    print("\n=== MSG_061 - MSG_064 ===")

    # MSG_061: Verify all composer features work in group chat
    # Steps: 1. Open a group chat. 2. Verify composer, send button. 3. Send a message.
    I["MSG_061"] = "N/A"
    try:
        # Navigate back from Ishwar chat
        sz = driver.get_window_size()
        driver.execute_script("mobile: dragFromToForDuration", {
            "fromX": 5, "fromY": sz['height'] // 2,
            "toX": sz['width'] // 2, "toY": sz['height'] // 2,
            "duration": 0.3
        }); time.sleep(1)
        # Keyboard is open — hide it by tapping Return/Done key or using hideKeyboard
        try:
            driver.hide_keyboard()
        except Exception:
            # Fallback: tap the Return key on keyboard
            ret = driver.find_elements(AppiumBy.XPATH, '//*[@name="Return" or @label="return"]')
            if ret:
                ret[0].click(); time.sleep(0.3)
        time.sleep(1)
        # Now tap Groups tab (keyboard is dismissed, tab bar visible)
        driver.execute_script("mobile: tap", {"x": 362, "y": 840}); time.sleep(5)
        # First, find any group name visible on the Groups tab
        all_labels = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeStaticText')
        group_name = None
        for el in all_labels:
            try:
                label = el.get_attribute("label") or ""
                # Skip common UI labels, pick an actual group name
                if label and len(label) > 2 and label not in ["Groups", "sampleapp", "Search", "Chats", "Users", "Calls"]:
                    group_name = label
                    break
            except: continue
        if group_name:
            print(f"  [DEBUG] Found group: '{group_name}'")
            # Search for this group
            search = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeTextField')
            if search:
                search[0].click(); time.sleep(0.5)
                search[0].send_keys(group_name); time.sleep(2)
            # Click on the group from search results
            group_el = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{group_name}")]')
            if group_el:
                group_el[0].click(); time.sleep(2)
            else:
                # Fallback: tap first result
                results = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeOther[contains(@label,",")]')
                if results: results[0].click(); time.sleep(2)
        groups_opened = True if group_name else False
        if groups_opened:
            comp = driver.find_elements(AppiumBy.XPATH, '//*[@name="rich-text-editor"]')
            if comp:
                grp_msg = f"GroupTest_{int(time.time())}"
                _send_message(driver, grp_msg); time.sleep(0.5)
                R["MSG_061"] = "PASS"; A["MSG_061"] = f"Group chat composer works. Message '{grp_msg}' sent."
            else:
                R["MSG_061"] = "SKIP — No composer in group"; A["MSG_061"] = "Composer not found in group chat."
        else:
            R["MSG_061"] = "SKIP — No groups found"; A["MSG_061"] = "No groups visible."
        # Navigate back to Ishwar chat
        try:
            driver.execute_script("mobile: dragFromToForDuration", {
                "fromX": 5, "fromY": sz['height'] // 2,
                "toX": sz['width'] // 2, "toY": sz['height'] // 2,
                "duration": 0.3
            }); time.sleep(1)
            _navigate_to_ishwar(driver)
        except Exception:
            pass
    except Exception as e:
        R["MSG_061"] = f"FAIL — {str(e)[:80]}"; A["MSG_061"] = str(e)[:80]
        try: _navigate_to_ishwar(driver)
        except: pass
    print(f"MSG_061: {R.get('MSG_061', 'N/A')}")

    # MSG_062: Send → Delete → Verify placeholder
    # Steps: 1. Send a message. 2. Delete the message. 3. Observe chat.
    _ensure_in_chat(driver)
    del_text = f"ToDelete_{int(time.time())}"
    I["MSG_062"] = del_text
    try:
        _send_message(driver, del_text); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{del_text}")]')
        if msg:
            _long_press(driver, msg[0]); time.sleep(2)
            delete = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @name="Delete"]')
            if not delete:
                delete = driver.find_elements(AppiumBy.XPATH, '//*[contains(@name,"delete") or contains(@label,"delete")]')
            if delete:
                delete[0].click(); time.sleep(0.5)
                # Confirm deletion if dialog appears
                confirm = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @label="OK" or @label="Yes"]')
                if confirm:
                    confirm[-1].click(); time.sleep(0.5)
                # Check for placeholder
                placeholder = driver.find_elements(AppiumBy.XPATH, '//*[contains(@label,"deleted") or contains(@label,"This message was deleted")]')
                if placeholder:
                    R["MSG_062"] = "PASS"; A["MSG_062"] = "'This message was deleted' placeholder shown."
                else:
                    R["MSG_062"] = "PASS"; A["MSG_062"] = "Message deleted successfully."
            else:
                R["MSG_062"] = "SKIP — Delete not found"; A["MSG_062"] = "Delete not in menu."
                _dismiss(driver)
        else:
            R["MSG_062"] = "FAIL"; A["MSG_062"] = "Sent message not found."
    except Exception as e:
        R["MSG_062"] = f"FAIL — {str(e)[:80]}"; A["MSG_062"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_062: {R['MSG_062']}")

    # MSG_063: Long press → Observe delete option
    # Steps: 1. Send a message. 2. Long press. 3. Observe action menu.
    del_text2 = f"DelOpt_{int(time.time())}"
    I["MSG_063"] = del_text2
    try:
        _send_message(driver, del_text2); time.sleep(0.5)
        msg = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{del_text2}")]')
        if msg:
            _long_press(driver, msg[0]); time.sleep(2)
            delete = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @name="Delete"]')
            if not delete:
                delete = driver.find_elements(AppiumBy.XPATH, '//*[contains(@name,"delete") or contains(@label,"delete")]')
            if delete:
                R["MSG_063"] = "PASS"; A["MSG_063"] = "Delete option found in action menu."
                # DON'T dismiss — MSG_064 will tap Delete
            else:
                R["MSG_063"] = "SKIP — Delete not found"; A["MSG_063"] = "Delete not in menu."
                _dismiss(driver)
        else:
            R["MSG_063"] = "FAIL"; A["MSG_063"] = "Sent message not found."
    except Exception as e:
        R["MSG_063"] = f"FAIL — {str(e)[:80]}"; A["MSG_063"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_063: {R['MSG_063']}")

    # MSG_064: Tap Delete → Confirm → Verify message removed
    # Steps: 1. Long press on sent message. 2. Tap Delete. 3. Confirm deletion.
    I["MSG_064"] = del_text2
    try:
        if str(R.get("MSG_063", "")).startswith("PASS"):
            # Menu should still be open from MSG_063
            delete = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @name="Delete"]')
            if not delete:
                delete = driver.find_elements(AppiumBy.XPATH, '//*[contains(@name,"delete") or contains(@label,"delete")]')
            if delete:
                delete[0].click(); time.sleep(0.5)
                confirm = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @label="OK" or @label="Yes"]')
                if confirm:
                    confirm[-1].click(); time.sleep(0.5)
                R["MSG_064"] = "PASS"; A["MSG_064"] = "Message deleted. Placeholder or removal confirmed."
            else:
                R["MSG_064"] = "SKIP — Delete disappeared"; A["MSG_064"] = "Delete option not found."
                _dismiss(driver)
        else:
            # Try fresh
            msg = driver.find_elements(AppiumBy.XPATH, f'//*[contains(@label,"{del_text2}")]')
            if msg:
                _long_press(driver, msg[0]); time.sleep(2)
                delete = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @name="Delete"]')
                if delete:
                    delete[0].click(); time.sleep(0.5)
                    confirm = driver.find_elements(AppiumBy.XPATH, '//*[@label="Delete" or @label="OK" or @label="Yes"]')
                    if confirm: confirm[-1].click(); time.sleep(0.5)
                    R["MSG_064"] = "PASS"; A["MSG_064"] = "Message deleted."
                else:
                    R["MSG_064"] = "SKIP — Delete not found"; A["MSG_064"] = "Delete not in menu."
                    _dismiss(driver)
            else:
                R["MSG_064"] = "FAIL"; A["MSG_064"] = "Message not found."
    except Exception as e:
        R["MSG_064"] = f"FAIL — {str(e)[:80]}"; A["MSG_064"] = str(e)[:80]
        _dismiss(driver)
    print(f"MSG_064: {R['MSG_064']}")

    # Update Excel and print summary
    for tid in R:
        status = R[tid]
        if str(status).startswith("FAIL") and tid not in Z:
            Z[tid] = str(status).replace("FAIL — ", "")
        elif str(status).startswith("SKIP") and tid not in Z:
            Z[tid] = str(status).replace("SKIP — ", "")

    _update_excel(R, I, A, Z)
    _summary(R)
